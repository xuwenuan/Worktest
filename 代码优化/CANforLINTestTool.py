import configparser
import os
import re
import logging
import sys
import traceback

import openpyxl
import xlrd2
import csv

from PyQt5.QtGui import QCursor
from xlsxwriter import Workbook
from PyQt5 import QtWidgets, QtCore
from PyQt5.QtWidgets import QMessageBox, QApplication, QMainWindow, QFileDialog, QTableWidgetItem, QHeaderView, \
    QPushButton, QComboBox, QAction, QMenu
from xlsxwriter import Workbook
from CanLinConfig import CanLinConfig
import tkinter as tk
from tkinter import messagebox

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1300, 600)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout.setObjectName("gridLayout")
        self.tableWidget = QtWidgets.QTableWidget(self.centralwidget)
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)
        self.gridLayout.addWidget(self.tableWidget, 0, 0, 1, 5)
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setObjectName("pushButton")
        self.gridLayout.addWidget(self.pushButton, 1, 1, 1, 1)
        self.pushButton_2 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_2.setObjectName("pushButton_2")
        self.gridLayout.addWidget(self.pushButton_2, 1, 0, 1, 1)
        self.pushButton3 = QtWidgets.QPushButton(self.centralwidget)
        self.gridLayout.addWidget(self.pushButton3, 1, 2, 1, 1)

        # 清空数据按钮
        self.pushButton_clear = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_clear.setObjectName("pushButton_clear")
        self.gridLayout.addWidget(self.pushButton_clear, 1, 3, 1, 1)

        # 导出到Excel按钮
        self.pushButton_export = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_export.setObjectName("pushButton_export")
        self.gridLayout.addWidget(self.pushButton_export, 1, 4, 1, 1)

        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 800, 23))
        self.menubar.setObjectName("menubar")
        self.menu = QtWidgets.QMenu(self.menubar)
        self.menu.setObjectName("menu")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "空调测试用例编写工具"))
        self.pushButton.setText(_translate("MainWindow", "导出自动测试文件"))
        self.pushButton_2.setText(_translate("MainWindow", "导入文件"))
        self.pushButton3.setText('添加行')
        self.pushButton_clear.setText(_translate("MainWindow", "清空数据"))  # 设置按钮文本
        self.pushButton_export.setText(_translate("MainWindow", "导出测试用例"))


class RoutingTest(QMainWindow,Ui_MainWindow):
    logging.basicConfig(
            filename="log.log",
            filemode="w",
            datefmt="%Y-%m-%d %H:%M:%S %p",
            format="%(asctime)s - %(name)s - %(levelname)s - %(module)s: %(message)s",
            level=40
        )
    routingList = []
    CANList = []
    logging.basicConfig(
        filename="log.log",
        filemode="w",
        datefmt="%Y-%m-%d %H:%M:%S %p",
        format="%(asctime)s - %(name)s - %(levelname)s - %(module)s: %(message)s",
        level=logging.DEBUG
    )

    def __init__(self, parent = None):
        super(RoutingTest, self).__init__(parent)
        self.setupUi(self)
        self.pushButton_2.clicked.connect(self.exceltoTable)
        self.settable()
        # self.pushButton.clicked.connect(lambda :self.readMesseage(path2,path3))
        self.pushButton3.clicked.connect(lambda: self.addrow(0))
        self.pushButton_clear.clicked.connect(self.clear_table_data)  # 绑定清空数据按钮
        self.pushButton_export.clicked.connect(self.export_to_excel)

    def clear_table_data(self):
        """
        清空表格中的所有数据和行
        """
        try:
            response = QMessageBox.question(self, '确认', '确定清空数据吗？')
            if response == QMessageBox.Yes:
                self.tableWidget.clearContents()  # 清空表格内容
                self.tableWidget.setRowCount(0)  # 清空表格行数
                self.settable()  # 重新设置表格头
                # 清空外部缓存数据（如果有）
                self.routingList.clear()
                self.CANList.clear()

                # QMessageBox.information(self, "提示", "数据已清空！")
            else:
                pass
        except Exception as e:
            logging.error(traceback.format_exc() + "\n")

    def export_to_excel(self):
        try:
            # 打开文件保存对话框，让用户选择保存路径
            filename, _ = QFileDialog.getSaveFileName(self, "导出测试用例", "", "Excel Files (*.xlsx)")
            if not filename:
                return  # 如果用户取消选择，直接返回

            # 创建 Excel 工作簿和工作表
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "导出数据"

            # 写入表头
            headers = ['源网段', '源节点', '源ID', '源端是否CANFD', '源DLC', '源周期', '源字段', '目标网段', '目标节点',
                        '目标ID','目标端是否CANFD','目标DLC','目标周期','目标字段','报文路由类型','MsgRoutType','测试使能','可选项1','操作']
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)

            # 写入表格数据
            for row_num in range(self.tableWidget.rowCount()):
                for col_num in range(self.tableWidget.columnCount()):
                    item = self.tableWidget.item(row_num, col_num)
                    if item is not None:
                        sheet.cell(row=row_num + 2, column=col_num + 1, value=item.text())

            # 保存文件
            workbook.save(filename)
            QMessageBox.information(self, "成功", f"数据已成功导出到 {filename}")
        except Exception as e:
            logging.error(f"导出到Excel时发生错误: {traceback.format_exc()}")
            QMessageBox.critical(self, "错误", "导出失败，请查看日志！")

    def exceltoTable(self):
        try:
            self.openfile = self.exceltoTable or os.getcwd()
            path = QFileDialog.getOpenFileName(self, "选择文件", '/', "xlsx Files(*.xlsx)")
            if path[0]:
                book = xlrd2.open_workbook(path[0])
                sheet = (book.sheets()[0])
                row = self.tableWidget.rowCount()
                for i in range(1, sheet.nrows):
                    values = sheet.row_values(i)
                    self.addrow(0)
                    for j in range(len(values)):
                        try:
                            text = str(int(values[j]))
                        except:
                            text = str(values[j])
                        self.tableWidget.setItem(row + i - 1, j, QTableWidgetItem(text))
        except:
            logging.error(traceback.format_exc() + "\n")



    def addrow(self, row):
        try:
            if row == 0:
                rowPosition = self.tableWidget.rowCount()
            else:
                rowPosition = row
            self.tableWidget.insertRow(rowPosition)
            deleteButton = QPushButton("删除".format(rowPosition))
            deleteButton.clicked.connect(self.delete_clicked)
            self.tableWidget.setCellWidget(rowPosition, 18, deleteButton)
        except:
            logging.error(traceback.format_exc() + "\n")

    def delete_clicked(self):
        button = self.sender()
        if button:
            row = self.tableWidget.indexAt(button.pos()).row()
            self.tableWidget.removeRow(row)
            self.tableWidget.verticalScrollBar().setSliderPosition(row)

    def settable(self):
        headnameList = ['源网段', '源节点', '源ID', '源端是否CANFD', '源DLC', '源周期', '源字段', '目标网段', '目标节点',
                        '目标ID','目标端是否CANFD','目标DLC','目标周期','目标字段','报文路由类型','MsgRoutType','测试使能','可选项1','操作']
        self.tableWidget.setColumnCount(len(headnameList))
        self.tableWidget.setRowCount(0)
        self.tableWidget.verticalHeader().setVisible(False)
        self.tableWidget.setHorizontalHeaderLabels(headnameList)
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # self.tableWidget.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        # self.tableWidget.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.tableWidget.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        # self.tableWidget.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeToContents)
        # self.tableWidget.horizontalHeader().setSectionResizeMode(9, QHeaderView.ResizeToContents)
        self.tableWidget.horizontalHeader().setSectionResizeMode(11, QHeaderView.ResizeToContents)
        # self.tableWidget.horizontalHeader().setSectionResizeMode(18, QHeaderView.ResizeToContents)
    def show_success_dialog(self,state,message):
        # 创建一个简单的Tkinter窗口
        success_dialog = tk.Tk()
        # 隐藏主窗口
        success_dialog.withdraw()
        # 显示成功消息
        messagebox.showinfo(state, message)


    #添加的复制粘贴模块
    def contextMenuEvent(self, event):
        try:
            # 创建右键菜单
            menu = QMenu(self)
            row = self.tableWidget.rowAt(event.y())
            # 添加复制和粘贴功能
            copy_action = QAction("复制 (Ctrl+C)", self)
            paste_action = QAction("粘贴 (Ctrl+V)", self)

            copy_action.triggered.connect(self.copy_selected_cells)
            paste_action.triggered.connect(self.paste_cells)

            menu.addAction(copy_action)
            menu.addAction(paste_action)

            # 显示菜单
            menu.exec_(QCursor.pos())

        except Exception as e:
            pass

    def copy_selected_cells(self):
        """复制选中的单元格内容到剪贴板"""
        selected_ranges = self.tableWidget.selectedRanges()
        if not selected_ranges:
            return

        clipboard_text = ""
        for selected_range in selected_ranges:
            top_row, bottom_row = selected_range.topRow(), selected_range.bottomRow()
            left_col, right_col = selected_range.leftColumn(), selected_range.rightColumn()

            # 构造复制内容
            for row in range(top_row, bottom_row + 1):
                row_data = []
                for col in range(left_col, right_col + 1):
                    item = self.tableWidget.item(row, col)
                    if item:
                        row_data.append(item.text())
                    else:
                        widget = self.tableWidget.cellWidget(row, col)
                        if isinstance(widget, QComboBox):
                            row_data.append(widget.currentText())
                        else:
                            row_data.append("")
                clipboard_text += "\t".join(row_data) + "\n"

        # 将内容复制到剪贴板
        clipboard = QApplication.clipboard()
        clipboard.setText(clipboard_text.strip())

    def paste_cells(self):
        """将剪贴板内容粘贴到选中的单元格"""
        clipboard = QApplication.clipboard()
        clipboard_text = clipboard.text()
        if not clipboard_text:
            # logging.info("No text in clipboard to paste.")
            return

        # 解析剪贴板内容为二维数组
        rows = clipboard_text.split("\n")
        data = [row.split("\t") for row in rows]

        # 获取目标起始位置
        selected_ranges = self.tableWidget.selectedRanges()
        if not selected_ranges:
            # logging.info("No cells selected for pasting.")
            return
        start_row = selected_ranges[0].topRow()
        start_col = selected_ranges[0].leftColumn()

        # 填充表格
        for i, row_data in enumerate(data):
            for j, cell_value in enumerate(row_data):
                target_row = start_row + i
                target_col = start_col + j

                # 动态扩展表格
                if target_row >= self.tableWidget.rowCount():
                    self.tableWidget.insertRow(target_row)
                if target_col >= self.tableWidget.columnCount():
                    self.tableWidget.insertColumn(target_col)

                # 设置单元格内容
                widget = self.tableWidget.cellWidget(target_row, target_col)
                if isinstance(widget, QComboBox):
                    widget.setCurrentText(cell_value)
                    # logging.info(f"Pasted text '{cell_value}' to QComboBox at ({target_row}, {target_col})")
                else:
                    item = QTableWidgetItem(cell_value)
                    self.tableWidget.setItem(target_row, target_col, item)

    def readMesseage(self,path2):
        try:
            filename = QFileDialog.getSaveFileName(self, '导出测试用例', '', 'Excel File(*.xlsx)')
            if filename[0]:
                path3=filename[0]
                for i in range(0, self.tableWidget.rowCount()):
                    datalist = []
                    for j in range(0, self.tableWidget.columnCount() - 1):
                        data = self.tableWidget.item(i, j).text()
                        datalist.append(data)
                    self.routingList.append(datalist)
                if path2:
                    # book1 = xlrd2.open_workbook(path1)
                    # sheetbook1 = book1.sheet_by_name('Sheet1')
                    # for i in range(1,sheetbook1.nrows):
                    #     values = sheetbook1.row_values(i)
                    #     self.routingList.append(values)
                    book2 = xlrd2.open_workbook(path2)
                    sheetbook2 = book2.sheet_by_name('硬件配置表')
                    self.LINList=sheetbook2.col_values(3)[3:]
                    self.CANList = sheetbook2.col_values(2)[3:]
                    if self.genautotable(path3):
                        self.show_success_dialog('Success', '生成成功！')
                    else:
                        self.show_success_dialog('FAIL', '生成失败，没有找到相关属性！')
        except:
            logging.error(traceback.format_exc() + "\n")
            self.show_success_dialog('FAIL', '生成失败，查看log！')

    # def genautotable(self,path3):
    #     # self.routingList = self.routingList[1:]
    #     dataList=[]
    #     # print(self.routingList)
    #     for item in self.routingList:
    #         preview = []
    #         try:
    #             data1 = '0'+str(self.LINList.index(item[0]))+'01'
    #         except:
    #             data1 = '0000'
    #         if int(item[3]) == 2:
    #             ID = '4'+item[2].replace('0x', '').replace('0X', '').zfill(3)
    #         else:
    #             ID = item[2].replace('0x', '').replace('0X', '').zfill(4)
    #         if int(item[5]) < 20:
    #             data3 = '0014'
    #             cycle = '20'
    #         else:
    #             data3 = str(hex(int(item[5]))).upper()[2:].zfill(4)
    #             cycle = item[5]
    #         data2 = ID + str(hex(int(cycle) * 5)).upper()[2:].zfill(4)
    #         data4 = str(hex((int(item[4])-1)*8)).upper()[2:].zfill(4)
    #         data5 = str(hex(int(item[4])*8)).upper()[2:].zfill(4)
    #         message1 = data1+data2+data3+data4+data5
    #         Mac = ' '.join(message1[i:i + 2] for i in range(0, len(message1), 2)).upper()
    #         if item[1] != '/':
    #             name = item[0]+'_'+item[2]+"_"+item[1]
    #         else:
    #             name = item[0]+'_'+item[2]+"_"
    #         data = ['event', name, Mac, cycle, '3', '0x700', '', '', '--', 'Motorola']
    #         dataList.append(data)
    #         preview.append(data)
    #         seconditem = ['check', '单条测试用例处理状态', '0', cycle, '100', '0x710', '16', '8', '--', 'Motorola']
    #         dataList.append(seconditem)
    #         preview.append(seconditem)
    #         s = 'FF'*(int(item[4])+1)
    #         string = ' '.join(s[i:i + 2] for i in range(2, len(s), 2)).upper()
    #         thirditem = ['event', name, string, cycle, '3', item[2], '', '', '--', 'Motorola']
    #         anthirditem = ['event', name, string.replace('FF', '00'), cycle, '3', item[2], '', '', '--', 'Motorola']
    #         dataList.append(thirditem)
    #         preview.append(anthirditem)
    #         if item[8] != '/':
    #             name = item[7]+'CAN_'+item[9]+"_"+item[8]
    #         else:
    #             name = item[7]+'CAN_'+item[9]+"_"
    #         try:
    #             channel = '0' + str(self.CANList.index(item[7] + '_CAN')) + '00'
    #         except:
    #             channel = '0000'
    #         if int(item[10]) ==1:
    #             ID = '8'+item[9].replace('0x', '').replace('0X', '').zfill(3)
    #         else:
    #             ID = item[9].replace('0x', '').replace('0X', '').zfill(4)
    #         if int(item[12]) < 20:
    #             data3 = '0014'
    #             cycle2 = '20'
    #         else:
    #             data3 = str(hex(int(item[12]))).upper()[2:].zfill(4)
    #             cycle2 = item[12]
    #
    #         data4 = str(hex((int(item[11]) - 1) * 8)).upper()[2:].zfill(4)
    #         data5 = str(hex(int(item[11]) * 8)).upper()[2:].zfill(4)
    #         message2 = channel+ID+str(hex(int(cycle2)*5)).upper()[2:].zfill(4)+data3+data4+data5
    #         Mac2 = ' '.join(message2[i:i + 2] for i in range(0, len(message2), 2)).upper()
    #         forthitem = ['event', name, Mac2, cycle2, '3', '0x701', '', '', '--', 'Motorola']
    #         dataList.append(forthitem)
    #         preview.append(forthitem)
    #         fifthitem = ['check', '单条测试用例处理状态', '1', int(cycle2)*5, '100', '0x710', '16', '8', '--', 'Motorola']
    #         dataList.append(fifthitem)
    #         preview.append(fifthitem)
    #         t = 'FF' * (int(item[11])+1)
    #         text = ' '.join(t[i:i + 2] for i in range(2, len(t), 2)).upper()
    #         sixth = ['routercheck', name, text, cycle2, '100', item[9], '', '', '--', 'Motorola']
    #         ansixth = ['routercheck', name, text.replace('FF', '00'), cycle2, '100', item[9], '', '', '--', 'Motorola']
    #         dataList.append(sixth)
    #         preview.append(ansixth)
    #         dataList = dataList+preview
    #         text=item[0]+' '+item[1]+' '+item[2]+'      '+item[7]+' '+item[8]+' '+item[9]
    #         note = ['测试用例说明', text, '12', '', '', '', '', '', '', '']
    #         dataList.append(note)
    #     if not dataList:
    #         return False
    #     else:
    #         headnameList = ['序号', '操作类型', '操作名称', '操作值', '间隔（ms）', 'Cycle（ms）', 'canID', 'Start', 'Length',
    #                         'flag', 'format']
    #         workbook = Workbook(path3)
    #         sheet1 = workbook.add_worksheet('CAN')
    #         sheet1.set_column('C:D', 50)
    #         sheet1.set_column('B:B', 15)
    #         sheet1.set_column('E:K', 10)
    #         font = workbook.add_format({'font_name': '等线', 'font_size': 12, 'align': 'center'})
    #         for k in range(len(headnameList)):
    #             sheet1.write(0, k, headnameList[k], font)
    #         row = 1
    #         for l in range(len(dataList)):
    #             sheet1.write(row, 0, str(row), font)
    #             for j in range(len(dataList[l])):
    #                 sheet1.write(row, j+1, dataList[l][j], font)
    #             row = row+1
    #         workbook.close()
    #         return True

    def genautotable(self, path3):
        dataList = []
        # 遍历 tableWidget 中的每一行，动态生成 dataList
        for row in range(self.tableWidget.rowCount()):
            item = []
            for col in range(self.tableWidget.columnCount() - 1):  # 排除最后一列的操作按钮
                cell_item = self.tableWidget.item(row, col)
                if cell_item is not None:
                    item.append(cell_item.text())
                else:
                    item.append("")  # 如果单元格为空，使用空字符串作为默认值

            # 以下逻辑保持不变，仅将数据来源从 routingList 替换为 tableWidget
            preview = []
            try:
                data1 = '0' + str(self.LINList.index(item[0])) + '01'
            except ValueError:
                data1 = '0000'
            if int(item[3]) == 2:
                ID = '4' + item[2].replace('0x', '').replace('0X', '').zfill(3)
            else:
                ID = item[2].replace('0x', '').replace('0X', '').zfill(4)
            if int(item[5]) < 20:
                data3 = '0014'
                cycle = '20'
            else:
                data3 = str(hex(int(item[5]))).upper()[2:].zfill(4)
                cycle = item[5]
            data2 = ID + str(hex(int(cycle) * 5)).upper()[2:].zfill(4)
            data4 = str(hex((int(item[4]) - 1) * 8)).upper()[2:].zfill(4)
            data5 = str(hex(int(item[4]) * 8)).upper()[2:].zfill(4)
            message1 = data1 + data2 + data3 + data4 + data5
            Mac = ' '.join(message1[i:i + 2] for i in range(0, len(message1), 2)).upper()
            if item[1] != '/':
                name = item[0] + '_' + item[2] + "_" + item[1]
            else:
                name = item[0] + '_' + item[2] + "_"
            data = ['event', name, Mac, cycle, '3', '0x700', '', '', '--', 'Motorola']
            dataList.append(data)
            preview.append(data)

            seconditem = ['check', '单条测试用例处理状态', '0', cycle, '100', '0x710', '16', '8', '--', 'Motorola']
            dataList.append(seconditem)
            preview.append(seconditem)

            s = 'FF' * (int(item[4]) + 1)
            string = ' '.join(s[i:i + 2] for i in range(2, len(s), 2)).upper()
            thirditem = ['event', name, string, cycle, '3', item[2], '', '', '--', 'Motorola']
            anthirditem = ['event', name, string.replace('FF', '00'), cycle, '3', item[2], '', '', '--', 'Motorola']
            dataList.append(thirditem)
            preview.append(anthirditem)

            if item[8] != '/':
                name = item[7] + 'CAN_' + item[9] + "_" + item[8]
            else:
                name = item[7] + 'CAN_' + item[9] + "_"
            try:
                channel = '0' + str(self.CANList.index(item[7] + '_CAN')) + '00'
            except ValueError:
                channel = '0000'
            if int(item[10]) == 1:
                ID = '8' + item[9].replace('0x', '').replace('0X', '').zfill(3)
            else:
                ID = item[9].replace('0x', '').replace('0X', '').zfill(4)
            if int(item[12]) < 20:
                data3 = '0014'
                cycle2 = '20'
            else:
                data3 = str(hex(int(item[12]))).upper()[2:].zfill(4)
                cycle2 = item[12]

            data4 = str(hex((int(item[11]) - 1) * 8)).upper()[2:].zfill(4)
            data5 = str(hex(int(item[11]) * 8)).upper()[2:].zfill(4)
            message2 = channel + ID + str(hex(int(cycle2) * 5)).upper()[2:].zfill(4) + data3 + data4 + data5
            Mac2 = ' '.join(message2[i:i + 2] for i in range(0, len(message2), 2)).upper()
            forthitem = ['event', name, Mac2, cycle2, '3', '0x701', '', '', '--', 'Motorola']
            dataList.append(forthitem)
            preview.append(forthitem)

            fifthitem = ['check', '单条测试用例处理状态', '1', int(cycle2) * 5, '100', '0x710', '16', '8', '--', 'Motorola']
            dataList.append(fifthitem)
            preview.append(fifthitem)

            t = 'FF' * (int(item[11]) + 1)
            text = ' '.join(t[i:i + 2] for i in range(2, len(t), 2)).upper()
            sixth = ['routercheck', name, text, cycle2, '100', item[9], '', '', '--', 'Motorola']
            ansixth = ['routercheck', name, text.replace('FF', '00'), cycle2, '100', item[9], '', '', '--', 'Motorola']
            dataList.append(sixth)
            preview.append(ansixth)

            dataList = dataList + preview
            text = item[0] + ' ' + item[1] + ' ' + item[2] + '      ' + item[7] + ' ' + item[8] + ' ' + item[9]
            note = ['测试用例说明', text, '12', '', '', '', '', '', '', '']
            dataList.append(note)

        has_dataList = bool(dataList)
        if not has_dataList:
            QMessageBox.warning(self, "警告", "没有可保存的数据！")
            return False
        else:
            headnameList = ['序号', '操作类型', '操作名称', '操作值', '间隔（ms）', 'Cycle（ms）', 'canID', 'Start', 'Length',
                            'flag', 'format']
            workbook = Workbook(path3)
            sheet1 = workbook.add_worksheet('CAN')
            sheet1.set_column('C:D', 50)
            sheet1.set_column('B:B', 15)
            sheet1.set_column('E:K', 10)
            font = workbook.add_format({'font_name': '等线', 'font_size': 12, 'align': 'center'})
            for k in range(len(headnameList)):
                sheet1.write(0, k, headnameList[k], font)
            row = 1
            for l in range(len(dataList)):
                sheet1.write(row, 0, str(row), font)
                for j in range(len(dataList[l])):
                    sheet1.write(row, j + 1, dataList[l][j], font)
                row = row + 1
            workbook.close()
            return True


if __name__ == '__main__':
    # path1= r"C:\Users\pc\Downloads\MessageRoutingTestTable_V01.csv"
    # path2 = r"C:\Users\pc\Downloads\自动化测试盒协议_BDM_KQC2_R.xlsx"
    # path3='./路由测试.xlsx'
    app =QApplication(sys.argv)
    # 初始化
    mainwindow = QMainWindow()
    ui_components = Ui_MainWindow()
    ui_components.setupUi(mainwindow)
    aa = RoutingTest()
    aa.show()
    sys.exit(app.exec_())



















