import math
import os
import re
import logging
import traceback
import PyQt5
from PyQt5 import QtGui, QtWidgets, QtCore
from PyQt5.QtCore import Qt, QEvent, QThread, pyqtSignal
from PyQt5.QtGui import QBrush, QColor, QCursor, QPalette, QFont, QDesktopServices
from PyQt5.QtWidgets import QAbstractItemView, QApplication, QMainWindow, QHeaderView, QFileDialog, QTableWidgetItem, \
    QPushButton, \
    QCompleter, QMessageBox, QLineEdit, QMenu, QComboBox, QTableWidget, QAction, QProgressDialog
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.units import points_to_pixels
from win32comext.shell.demos.servers import context_menu
from xlsxwriter.workbook import Workbook
from ELFAnalysis import ELFAnalysis
# from openpyxl import Workbook
import openpyxl
from openpyxl.utils import get_column_letter
import sys
import xlrd2
from CanLinConfig import CanLinConfig
from tabletest_copy import Ui_MainWindow
# from tabletest import Ui_MainWindow
from GetFileData import GetFileData
# from AddChapter import NewWindow4
# from docx import Document
from InterfaceTest import InterfaceTest
from CANforLINTestTool import RoutingTest
from SignalRoutingTestTool import SignalRoutingTest
import atexit
import threading
import time
from PyQt5.QtCore import QUrl

class LoadELFThread(QThread):
    finished = pyqtSignal(str)  # 成功信号，传递成功消息
    error = pyqtSignal(str)     # 错误信号，传递错误消息

    def __init__(self, elf_analysis, elf_path):
        super().__init__()
        self.elf_analysis = elf_analysis
        self.elf_path = elf_path

    def run(self):
        try:
            self.elf_analysis.loadPath(self.elf_path)
            self.finished.emit(f"文件 '{self.elf_path}' 导入成功！")
        except FileNotFoundError:
            self.error.emit(f"文件 '{self.elf_path}' 不存在，请检查路径是否正确。")
        except PermissionError:
            self.error.emit(f"无法访问文件 '{self.elf_path}'，请检查文件权限。")
        except Exception as e:
            self.error.emit(f"文件导入失败：{str(e)}")


class QComboboxEx(QComboBox):
    comboBoxID = ""
    canOrLinValue = ""
    index = 0

    def wheelEvent(self, QWheelEvent) -> None:
        pass

class MainWindow(QMainWindow,Ui_MainWindow):
    diclist,lindic = [],[]
    diclist1,lindic1= [],[]
    datalist,linlist=[],[]
    signalD,linD = [],[]
    signalD1,linD1=[],[]
    CAN = []
    LIN =[]
    openclose=[]
    temp=[]
    CV=[]
    PWM=[]
    comboBoxList3=[]
    comboBoxList1=['前提','触发','期望','配置']
    interfaceList=[]
    interfaceDic=[]
    structDict=[]
    RTlist = []
    ID=[]
    banqiao=[]
    Canname=''
    Linname=''
    filename =''
    resolutionnum=''
    num = 0
    offset=0
    resolution=0
    chapterList = []
    chapterLevel1=[]
    chapterLevel2 = []
    chapterLevel3 = []
    chapterLevel4 = []
    config = CanLinConfig()
    getfile = GetFileData()
    elfAnalysis = ELFAnalysis()
    # InterfaceTest = InterfaceTest()

    first_open_file_call = True

    logging.basicConfig(
        filename="log.log",
        filemode="w",
        datefmt="%Y-%m-%d %H:%M:%S %p",
        format="%(asctime)s - %(name)s - %(levelname)s - %(module)s: %(message)s",
        level=40
    )

    def __init__(self, parent=None):
        try:
            super(MainWindow, self).__init__(parent)
            self.setupUi(self)
            self.settable()
            self.pushButton_2.clicked.connect(lambda: self.addrow(0))
            self.action_xlsx.triggered.connect(self.open_file)
            self.pushButton.clicked.connect(self.genmessage)
            self.pushButton_4.clicked.connect(self.saveToexcel)
            self.pushButton_5.clicked.connect(lambda: self.addnote1(0))
            self.action_xlsx2.triggered.connect(self.getELFAnalysis)
            self.pushButton_3.clicked.connect(self.clear_row)
            self.action_xlsx1.triggered.connect(self.excelTotable)
            self.action_xlsx4.triggered.connect(self.genAutoExcel)
            self.tableWidget.setContextMenuPolicy(Qt.CustomContextMenu)
            self.tableWidget.customContextMenuRequested.connect(self.contextMenuEvent)

            self.addwidget()
            thread = threading.Thread(target=self.save_file_periodically)
            thread.daemon = True
            thread.start()
            # self.getELFAnalysis()
            logging.basicConfig(
                filename="log.log",
                filemode="w",
                datefmt="%Y-%m-%d %H:%M:%S %p",
                format="%(asctime)s - %(name)s - %(levelname)s - %(module)s: %(message)s",
                level=40
            )
        except:
            logging.error(traceback.format_exc() + "\n")

    def addwidget(self):
        layout1=InterfaceTest()
        layout1.elfAnalysis = self.elfAnalysis
        layout1.pushButton.clicked.connect(lambda: layout1.readMesseage(self.filename,self.elfpath))
        self.tabWidget.addTab(layout1, '接口测试')

        layout2=RoutingTest()
        layout2.pushButton.clicked.connect(lambda: layout2.readMesseage(self.filename))
        self.tabWidget.addTab(layout2, 'id路由测试')

        layout3 = SignalRoutingTest()
        layout3.pushButton.clicked.connect(lambda: layout3.readMesseage(self.filename, './信号路由自动测试.xlsx', './硬件输入的信号路由测试.xlsx'))
        self.tabWidget.addTab(layout3, '信号路由测试')
        self.tabWidget.setCurrentIndex(0)


    # 保存文件的函数
    def save_file_periodically(self):
        try:
        # 初始化上次保存时间
            last_save_time = time.time()
            while True:
                # 等待一秒钟
                time.sleep(60)
                # 检查是否到了保存时间点
                current_time = time.time()
                if current_time - last_save_time >= 120:
                    self.tableToexcel(filepath='./测试用例（自动保存）.xlsx')
        except:
            logging.error(traceback.format_exc() + "\n")



    def insertChapter(self):
        self.win.setwin()
        if self.win.newchapter.startswith('Level 1:'):
            self.chapterList.append(self.win.newchapter)
            self.chapterList=sorted(self.chapterList, key=str.lower)
            self.datatotable()
            # self.tableWidget.verticalScrollBar().setSliderPosition(0)
        elif self.win.newchapter !='':
            msg_box = QMessageBox(QMessageBox.Information, "提示", "请输入正确的章节格式!")
            msg_box.exec_()

    # def getELFAnalysis(self):
    #     filter = "elf Files(*.elf);;out files (*.out)"
    #     path = QFileDialog.getOpenFileName(self, "选择文件", '/', filter)
    #     self.elfpath = path[0]
    #     if self.elfpath:
    #         # elfAnalysis = ELFAnalysis()
    #         self.elfAnalysis.loadPath(self.elfpath)
    #         # self.structDict = elfAnalysis.structDict
    #         msg_box = QMessageBox(QMessageBox.Information, "标题", "导入成功!")
    #         msg_box.exec_()

    def getELFAnalysis(self):
        # 提取文件过滤器和默认路径为类属性，提高灵活性
        file_filter = "elf Files(*.elf);;out files (*.out)"
        default_path = self.last_open_path if hasattr(self, 'last_open_path') else '/'

        # 打开文件选择对话框
        path = QFileDialog.getOpenFileName(self, "选择文件", default_path, file_filter)
        self.elfpath = path[0]

        # 如果用户取消选择，给出友好提示
        if not self.elfpath:
            msg_box = QMessageBox(QMessageBox.Warning, "提示", "未选择文件！")
            msg_box.exec_()
            return

        # 记住上次打开的路径
        self.last_open_path = os.path.dirname(self.elfpath)

        try:
            # 启动后台线程加载 ELF 文件
            self.thread = LoadELFThread(self.elfAnalysis, self.elfpath)
            self.thread.finished.connect(self.onLoadFinished)
            self.thread.error.connect(self.onLoadError)
            self.thread.start()

            # 使用 QProgressDialog 显示加载中的提示信息
            self.progress_dialog = QProgressDialog("正在加载文件，请稍候...", "取消", 0, 0, self)
            self.progress_dialog.setWindowModality(Qt.WindowModal)
            self.progress_dialog.setWindowTitle("加载中")
            self.progress_dialog.show()

            # 确保在加载完成后关闭进度对话框
            self.thread.finished.connect(self.progress_dialog.close)
            self.thread.error.connect(self.progress_dialog.close)

        except Exception as e:
            # 捕获异常并提示用户
            logging.error(f"文件导入失败：{str(e)}")
            error_title = "导入失败"
            error_message = f"文件导入失败：{str(e)}"
            msg_box = QMessageBox(QMessageBox.Critical, error_title, error_message)
            msg_box.exec_()

    def onLoadFinished(self, message):
        self.progress_dialog.close()
        QMessageBox.information(self, "导入成功", message)

    def onLoadError(self, message):
        self.progress_dialog.close()
        QMessageBox.critical(self, "导入失败", message)




    def contextMenuEvent(self, event):
        try:
        # 创建右键菜单
            menu = QMenu(self)
            row = self.tableWidget.rowAt(event.y())

        # 添加菜单项：插入步骤、插入备注
            insert_row_action = menu.addAction('插入步骤')
            note_row_action = menu.addAction('插入备注')

        # 添加复制和粘贴功能
            copy_action = QAction("复制 (Ctrl+C)", self)
            paste_action = QAction("粘贴 (Ctrl+V)", self)

            copy_action.triggered.connect(self.copy_selected_cells)
            paste_action.triggered.connect(self.paste_cells)

            menu.addAction(copy_action)
            menu.addAction(paste_action)

        # 绑定菜单项触发事件
            insert_row_action.triggered.connect(lambda: self.addrow1(row + 1))
            note_row_action.triggered.connect(lambda: self.addnote1(row + 1))

        # 显示菜单
            menu.exec_(QCursor.pos())

        except Exception as e:
            pass


    def copy_selected_cells(self):
        """复制选中的单元格内容到剪贴板"""
        selected_ranges = self.tableWidget.selectedRanges()
        if not selected_ranges:
            return

        clipboard_text = ""
        for selected_range in selected_ranges:
            top_row, bottom_row = selected_range.topRow(), selected_range.bottomRow()
            left_col, right_col = selected_range.leftColumn(), selected_range.rightColumn()

        # 构造复制内容
            for row in range(top_row, bottom_row + 1):
                row_data = []
                for col in range(left_col, right_col + 1):
                    item = self.tableWidget.item(row, col)
                    if item:
                        row_data.append(item.text())
                    else:
                        widget = self.tableWidget.cellWidget(row, col)
                        if isinstance(widget, QComboBox):
                            row_data.append(widget.currentText())
                        else:
                            row_data.append("")
                clipboard_text += "\t".join(row_data) + "\n"

    # 将内容复制到剪贴板
        clipboard = QApplication.clipboard()
        clipboard.setText(clipboard_text.strip())

    def paste_cells(self):
        """将剪贴板内容粘贴到选中的单元格"""
        clipboard = QApplication.clipboard()
        clipboard_text = clipboard.text()
        if not clipboard_text:
            return

    # 解析剪贴板内容为二维数组
        rows = clipboard_text.split("\n")
        data = [row.split("\t") for row in rows]

    # 获取目标起始位置
        selected_ranges = self.tableWidget.selectedRanges()
        if not selected_ranges:
            return
        start_row = selected_ranges[0].topRow()
        start_col = selected_ranges[0].leftColumn()

    # 填充表格
        for i, row_data in enumerate(data):
            for j, cell_value in enumerate(row_data):
                target_row = start_row + i
                target_col = start_col + j

            # 动态扩展表格
                if target_row >= self.tableWidget.rowCount():
                    self.tableWidget.insertRow(target_row)
                if target_col >= self.tableWidget.columnCount():
                    self.tableWidget.insertColumn(target_col)

            # 设置单元格内容
                widget = self.tableWidget.cellWidget(target_row, target_col)
                if isinstance(widget, QComboBox):
                    widget.setCurrentText(cell_value)
                else:
                    item = QTableWidgetItem(cell_value)
                    self.tableWidget.setItem(target_row, target_col, item)

        # 关闭窗口时保存文件提示
    def closeEvent(self, event):
        # 弹出提示框询问用户是否保存更改并退出
        reply = QtWidgets.QMessageBox.question(
            self, u'温馨提示', u'是否保存更改并退出?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No | QtWidgets.QMessageBox.Cancel
        )

        if reply == QtWidgets.QMessageBox.Yes:
            try:
                # 根据当前选项卡执行对应的保存操作
                current_tab_index = self.tabWidget.currentIndex()
                current_widget = self.tabWidget.currentWidget()

                if current_tab_index == 0:
                    self.saveToexcel()
                elif current_tab_index == 1:
                    current_widget.readMesseage(self.filename, self.elfpath)
                elif current_tab_index == 2:
                    current_widget.readMesseage(self.filename)
                elif current_tab_index == 3:
                    current_widget.readMesseage(self.filename, './信号路由自动测试.xlsx', './硬件输入的信号路由测试.xlsx')

                event.accept()  # 关闭窗口
            except Exception as e:
                logging.error(f"保存操作失败: {e}")
                QMessageBox.critical(self, "错误", "保存操作失败，请检查日志！")
                event.ignore()  # 取消关闭窗口
        elif reply == QtWidgets.QMessageBox.No:
            event.accept()  # 不保存直接关闭窗口
        else:
            event.ignore()  # 取消退出操作，保持窗口打开


        # 这是一个测试信息提示函数
    def show_test_message(self):
        msg_test = QMessageBox(QMessageBox.Information, "提示", "这是一条测试信息")
        msg_test.exec_()


    #这是添加菜单栏的帮助窗口
    def show_success_dialog(self, state, message):
        if state == 1:
            QDesktopServices.openUrl(QUrl.fromLocalFile("./caseTool使用说明.pdf"))
        elif state == 2:
            QMessageBox.information(self, "重要提示！", "第一步导入协议文件，第二步导入map文件，第三步才是导入对应的测试文件。这样才能正确导出测试用例")
        else:
            # 处理其他情况
            pass



    def getTabelCellData(self, row, section):
        date = ""
        try:
            date = self.tableWidget.cellWidget(row, section).currentText()
        except:
            try:
                date = self.tableWidget.item(row, section).text()
            except:
                pass
        return date

    def genmessage(self):
        try:
            row = self.tableWidget.rowCount()
            # 查找所有数字序列
            pattern = r'\d+'
            for i in range(row):
                parameter = self.getTabelCellData(i, 4)
                projectname = self.getTabelCellData(i, 2)
                combox_temp = self.tableWidget.cellWidget(i, 1)
                # print(parameter)
                if self.tableWidget.cellWidget(i,0) and self.tableWidget.cellWidget(i,1).currentText()!='软件接口测试' and self.tableWidget.cellWidget(i,0).currentText()!='配置':
                    if (self.tableWidget.cellWidget(i,0).currentText()=='前提' or self.tableWidget.cellWidget(i,0).currentText()=='触发') and self.tableWidget.cellWidget(i,1).currentText() in self.CAN+self.LIN:
                        data1 ='0x700: '
                        if self.tableWidget.cellWidget(i, 1).currentText() in self.CAN:
                            data2 = str(hex(self.CAN.index(self.tableWidget.cellWidget(i,1).currentText())+1).upper()[2:].zfill(2))
                        elif self.tableWidget.cellWidget(i, 1).currentText() in self.LIN:
                            data2 = str(hex(self.LIN.index(self.tableWidget.cellWidget(i, 1).currentText())+1).upper()[2:].zfill(2))
                        else:
                            data2 ='00'
                    elif self.tableWidget.cellWidget(i,0).currentText()=='期望' and self.tableWidget.cellWidget(i,1).currentText() in self.CAN+self.LIN:
                        data1='0x701: '
                        if self.tableWidget.cellWidget(i, 1).currentText() in self.CAN:
                            data2 = str(hex(self.CAN.index(self.tableWidget.cellWidget(i, 1).currentText())+1).upper()[2:].zfill(2))
                        elif self.tableWidget.cellWidget(i, 1).currentText() in self.LIN:
                            data2 = str(hex(self.LIN.index(self.tableWidget.cellWidget(i, 1).currentText())+1).upper()[2:].zfill(2))
                        else:
                            data2 ='00'
                    elif self.tableWidget.cellWidget(i,1).currentText()=='温度传感器':
                        data1= '0x'+str(int(self.ID[0]))+': '
                        if projectname in self.temp:
                            data2 = str(hex(self.temp.index(projectname)+1).upper()[2:].zfill(2))
                        else:
                            data2 ='00'
                    elif self.tableWidget.cellWidget(i,1).currentText()=='继电器输出' or self.tableWidget.cellWidget(i,1).currentText()=='半桥输出':
                        data1= '0x'+str(int(self.ID[3]))+': '
                        if self.tableWidget.cellWidget(i,1).currentText()=='继电器输出':
                            data2 = str(hex(self.PWM.index(projectname)+1).upper()[2:].zfill(2))
                        elif self.tableWidget.cellWidget(i,1).currentText()=='半桥输出':
                            data2 = str(hex(self.banqiao.index(projectname) + 1).upper()[2:].zfill(2))
                        else:
                            data2 ='00'
                    elif self.tableWidget.cellWidget(i,1).currentText()=='电压传感器':
                        data1= '0x'+str(int(self.ID[1]))+': '
                        if projectname in self.CV:
                            data2 = str(hex(self.CV.index(projectname)+1).upper()[2:].zfill(2))
                        else:
                            data2 ='00'
                    elif self.tableWidget.cellWidget(i,1).currentText()=='电压输入':
                        data1= '0x'+str(int(self.ID[2]))+': '
                        if projectname in self.Vinput:
                            data2 = str(hex(self.Vinput.index(projectname)+1).upper()[2:].zfill(2))
                        else:
                            data2 ='00'
                    elif self.tableWidget.cellWidget(i,1).currentText()=='开关输入':
                        data1= '0x'+str(int(self.ID[5]))+': '
                        if projectname in self.openclose:
                            data2 = str(hex(self.openclose.index(projectname)+1).upper()[2:].zfill(2))
                        else:
                            data2 ='00'
                    elif self.tableWidget.cellWidget(i,1).currentText()=='PWM输出':
                        data1= '0x'+str(int(self.ID[6]))+': '
                        if projectname in self.outputPWM:
                            data2 = str(hex(self.outputPWM.index(projectname)+1).upper()[2:].zfill(2))
                        else:
                            data2 ='00'
                    elif self.tableWidget.cellWidget(i,1).currentText()=='PWM输入':
                        data1= '0x'+str(int(self.ID[7]))+': '
                        if projectname in self.inputPWM:
                            data2 = str(hex(self.inputPWM.index(projectname)+1).upper()[2:].zfill(2))
                        else:
                            data2 ='00'
                    else:
                        data1='0x000: '
                        data2 ='00'
                    if  self.tableWidget.cellWidget(i,1).currentText() in self.LIN or self.tableWidget.cellWidget(i,1).currentText()=='半桥输出':
                        data3= ' 01'
                    else:
                        data3=' 00'
                    if parameter=='ON'or parameter=='有效':
                        data4 =' 00 01'
                        self.num=1
                    elif parameter=='OFF' or parameter=='无效':
                        data4 = ' 00 00'
                        self.num=0
                    elif self.tableWidget.cellWidget(i,1).currentText()=='开关输入' or self.tableWidget.cellWidget(i,1).currentText()=='半桥输出':
                        if parameter =='LSD':
                            data4=' 00 00'
                        elif parameter=='HSD':
                            data4=' 00 01'
                        elif parameter=='高阻':
                            data4=' 00 02'
                        else:
                            data4 = ' 00 00'
                    elif self.tableWidget.cellWidget(i,1).currentText()=='电压传感器':
                        number = 1/float(self.getfile.getresolution('电压值',self.filename))
                        newnum =int(float(parameter)*number)
                        print(number,parameter,newnum,hex(newnum))
                        newnum=str(hex(newnum)).upper()[2:].zfill(4)
                        num1 = newnum[:2]
                        num2 = newnum[2:]
                        data4 = ' ' + num1 + ' ' + num2
                    elif self.tableWidget.cellWidget(i,1).currentText()=='电压输入':
                        print(projectname)
                        number = 1 / float(self.getfile.getresolution(str(self.Vinput.index(projectname)+1)+'电压值',self.filename))
                        newnum = int(float(parameter) * number)
                        print(number, parameter, newnum, hex(newnum))
                        newnum = str(hex(newnum)).upper()[2:].zfill(4)
                        num1 = newnum[:2]
                        num2 = newnum[2:]
                        data4 = ' ' + num1 + ' ' + num2
                    elif self.tableWidget.cellWidget(i,1).currentText()=='PWM输入':
                        print(projectname)
                        number = 1 / float(self.getfile.getresolution(str(self.inputPWM.index(projectname)+1)+'占空比',self.filename))
                        newnum = int(float(parameter) * number)
                        print(number, parameter, newnum, hex(newnum))
                        newnum = str(hex(newnum)).upper()[2:].zfill(4)
                        num1 = newnum[:2]
                        num2 = newnum[2:]
                        data4 = ' ' + num1 + ' ' + num2
                    elif self.tableWidget.cellWidget(i,1).currentText()=='PWM输出':
                        print(projectname)
                        number = 1 / float(self.getfile.getresolution(str(self.outputPWM.index(projectname)+1)+'占空比',self.filename))
                        newnum = int(float(parameter) * number)
                        print(number, parameter, newnum, hex(newnum))
                        newnum = str(hex(newnum)).upper()[2:].zfill(4)
                        num1 = newnum[:2]
                        num2 = newnum[2:]
                        data4 = ' ' + num1 + ' ' + num2
                    # 参数在Cansignalname
                    elif projectname!='' and self.tableWidget.cellWidget(i,1).currentText()!='温度传感器':
                        name = projectname
                        # 参数为数字
                        # if parameter.isnumeric():
                        # 找偏移和精度
                        newnum = '0000'
                        #CAN用例解析矩阵
                        if name in self.datalist:
                            if parameter.startswith('0x') or parameter.startswith('0X'):
                                hex1 = int(parameter, 16)
                                self.num1 = hex1
                            elif parameter.isdecimal():
                                decimal = int(parameter)
                                self.num1 = decimal
                            for l in range(len(self.diclist1)):
                                if name in self.diclist1[l] and self.diclist1[l][2].startswith('CANFD'):
                                    nameid = self.diclist1[l][1].split('x')[1]
                                    if len(nameid)==2:
                                        newnum = '80' + nameid
                                    elif len(nameid)==3:
                                        newnum = '8'+nameid
                                elif name in self.diclist1[l]:
                                    nameid = self.diclist1[l][1].split('x')[1]
                                    if len(nameid) == 2:
                                        newnum = '00' + nameid
                                    elif len(nameid) == 3:
                                        newnum = '0' + nameid
                            for k in range(len(self.signalD1)):
                                if name == self.signalD1[k][0]:
                                    for j in range(len(self.signalD1[k][1:])):
                                        if parameter in self.signalD1[k][1:][j]:
                                            self.j = j
                                            # print(parameter, self.signalD1[k][j])
                                            if '~' in self.signalD1[k][1:][j].split(':')[0]:
                                                self.num = int(self.signalD1[k][1:][j].split(':')[0].split('~')[1], base=16)
                                                self.num1 = int(self.signalD1[k][1:][j].split(':')[0].split('~')[1],
                                                               base=16)
                                            else:
                                                self.num = int(self.signalD1[k][1:][j].split(':')[0], base=16)
                                                self.num1 = int(self.signalD1[k][1:][j].split(':')[0], base=16)
                                            newnum = str(hex(self.num).upper()[2:].zfill(4))
                        #测试盒CAN通讯矩阵
                        elif name in self.candiclist1[0]:
                            for l in range(len(self.candiclist1)):
                                if name in self.candiclist1[l]:
                                    nameid = self.candiclist1[l][2].split('x')[1]
                                    newnum = '0'+nameid
                        # LIN用例解析矩阵
                        elif name in self.linlist:
                            if parameter.startswith('0x') or parameter.startswith('0X'):
                                hex1 = int(parameter, 16)
                                self.num1 = hex1
                            elif parameter.isdecimal():
                                decimal = int(parameter)
                                self.num1=decimal
                            # else:
                            for l in range(len(self.lindic1)):
                                if name in self.lindic1[l]:
                                    nameid = self.lindic1[l][1].split('x')[1]
                                    if len(nameid)==2:
                                        newnum = '40' + nameid
                                    elif len(nameid)==3:
                                        newnum = '4'+nameid
                            for k in range(len(self.linD1)):
                                if name == self.linD1[k][0]:
                                    for j in range(len(self.linD1[k][1:])):
                                        if parameter in self.linD1[k][1:][j]:
                                            self.j = j
                                            # print(parameter, self.signalD1[k][j])
                                            if '~' in self.linD1[k][1:][j].split(':')[0]:
                                                self.num = int(self.linD1[k][1:][j].split(':')[0].split('~')[1], base=16)
                                                self.num1 = int(self.linD1[k][1:][j].split(':')[0].split('~')[1],
                                                               base=16)
                                            else:
                                                self.num = int(self.linD1[k][1:][j].split(':')[0], base=16)
                                                self.num1 = int(self.linD1[k][1:][j].split(':')[0], base=16)
                                            newnum = str(hex(self.num).upper()[2:].zfill(4))
                        num1 = newnum[:2]
                        num2 = newnum[2:]
                        data4 = ' '+num1+' '+num2
                    elif self.tableWidget.cellWidget(i,1) and self.tableWidget.cellWidget(i,1).currentText()=='温度传感器':
                        # if '度' in parameter:
                        index = self.temp.index(projectname)
                        tempvalue=int(self.RTlist[index][parameter])
                        newnum = str(hex(tempvalue).upper()[2:].zfill(4))
                        num1 = newnum[:2]
                        num2=newnum[2:]
                        data4 = ' '+num1+' '+num2
                        # else:
                        #     msg_box = QMessageBox(QMessageBox.Information, "提示", "温度传感器中\n请输入正确的参数格式!")
                        #     msg_box.exec_()
                    else:
                        data4=' 00 00'
                    if self.tableWidget.cellWidget(i,5) and self.tableWidget.cellWidget(i,5).text()!='':
                        text = self.tableWidget.cellWidget(i,5).text()
                        if '0x' in text or '0X' in text:
                            num = int(text, 16)
                        else:
                            num = int(re.findall(pattern, text)[0])
                        if self.tableWidget.cellWidget(i,3) and self.tableWidget.cellWidget(i,3).currentText()=='跳变':
                            if (text.startswith('0xF') or text.startswith('0XF')) and len(text)>=6:
                                data5 = ' FF FF'
                            else:
                                F000 = int('F000', 16)
                                result = str(hex(F000 + num).upper()[2:].zfill(4))
                                num1 = result[:2]
                                num2 = result[2:]
                                data5 = ' ' + num1 + ' ' + num2
                        else:
                            F000 = int('0000',16)
                            result = str(hex(F000+num).upper()[2:].zfill(4))
                            num1 = result[:2]
                            num2 = result[2:]
                            data5 = ' '+num1+' '+num2
                    else:
                        data5 =' FF FF'
                    if projectname in self.datalist:
                        time=0
                        for j in range(len(self.diclist1)):
                            if projectname in self.diclist1[j]:
                                time = int(self.diclist1[j][0])
                        result = str(hex(time).upper()[2:].zfill(4))
                        num1 = result[:2]
                        num2 = result[2:]
                        data6 = ' '+num1+' '+num2
                        for k in range(len(self.diclist)):
                            if projectname ==self.diclist[k][1]:
                                # BCAN,信号名称,精度，偏移量，起始位，信号长度，起始字节
                                self.length = self.diclist[k][5]
                                self.offset = self.diclist[k][3]
                                self.resolution=self.diclist[k][2]
                                self.startBit = self.diclist[k][4]
                                data6 = data6+self.config.getStartandLengthHex(self.startBit,self.length)
                                break
                    elif projectname in self.linlist:
                        time = 0
                        for j in range(len(self.lindic1)):
                            if projectname in self.lindic1[j]:
                                time = int(self.lindic1[j][0])
                        result = str(hex(time).upper()[2:].zfill(4))
                        num1 = result[:2]
                        num2 = result[2:]
                        data6 = ' ' + num1 + ' ' + num2
                        for k in range(len(self.lindic)):
                            if projectname ==self.lindic[k][1]:
                                self.startBit = self.lindic[k][4]
                                self.length = self.lindic[k][5]
                                self.offset=self.lindic[k][3]
                                self.resolution=self.lindic[k][2]
                                data6 = data6+self.config.getStartandLengthHex(self.startBit,self.length)
                                break
                        print(self.startBit,self.length)
                    else:
                        data6 = ' 00 00'
                    if self.tableWidget.cellWidget(i,0).currentText()!='配置' and (self.tableWidget.cellWidget(i,1).currentText() in self.CAN or self.tableWidget.cellWidget(i,1).currentText() in self.LIN):
                        # print(self.num,  self.diclist)
                        id = '0x000: '
                        data7 = '00 00 00 00 00 00 00 00'
                        if self.tableWidget.cellWidget(i,1).currentText() in self.CAN+self.LIN:
                            print(self.num1,self.offset,self.resolution)
                            if bool(re.match(r'^[-+]?[0-9]*\.?[0-9]+$', parameter)) is True:

                                num = int((self.num1-int(self.offset))/float(self.resolution))
                            else:
                                num = self.num1
                            combox_temp.canOrLinValue = num
                            # print(self.startBit,self.length,num,self.num)
                            data7 = self.config.getMessage(self.startBit,self.length,num,8,'Motorola')
                        if self.tableWidget.cellWidget(i,1).currentText() in self.CAN:
                            for k in range(len(self.diclist1)):
                                if projectname in self.diclist1[k]:
                                    id = self.diclist1[k][1]+": "
                                    break
                        elif self.tableWidget.cellWidget(i,1).currentText() in self.LIN:
                            for k in range(len(self.lindic1)):
                                if projectname in self.lindic1[k]:
                                    id = self.lindic1[k][1]+": "
                                    break
                        message=data1+data2+data3+data4+data5+data6+"\n"+id+data7
                    else:
                        message=data1+data2+data3+data4+data5+data6
                    if self.tableWidget.cellWidget(i,1).currentText()=='诊断服务':
                        if self.tableWidget.cellWidget(i, 0).currentText() == '期望':
                            id = '0x7F1: '
                        else:
                            id='0x7F0: '
                        LEN = str(hex(len(self.tableWidget.item(i,2).text().strip().split(' '))).upper()[2:].zfill(2))
                        NEXTLen='00'.zfill(8-1-len(self.tableWidget.item(i,2).text().strip().split(' ')))*2
                        NEXTText=re.sub(r"(?<=\w)(?=(?:\w\w)+$)"," ",NEXTLen)
                        message = id+self.tableWidget.item(i,2).text().strip()
                    self.tableWidget.setItem(i, 7, QTableWidgetItem(message))
                elif (self.tableWidget.cellWidget(i,0) and self.tableWidget.cellWidget(i,0).currentText()=='期望') and self.tableWidget.cellWidget(i,1).currentText()=='软件接口测试':
                    discribe = self.tableWidget.cellWidget(i,2).currentText().replace("\n", "").replace("\t", "").strip()
                    message = '00 00 00 00 00 00 00 00'
                    if bool(re.match(r'^[-+]?[0-9]*\.?[0-9]+$', parameter)) is True:
                        for k in range(len(self.interfaceDic)) :
                            # print(discribe,self.interfaceDic[k][2])
                            if discribe ==self.interfaceDic[k][2]:
                                signalName = self.interfaceDic[k][1]
                                signalMac = self.elfAnalysis.getAddressWithName(signalName)
                                Mac=' '.join(signalMac[i:i + 2] for i in range(2, len(signalMac), 2)).upper()
                                signaltype= self.interfaceDic[k][0]
                                signalLen,signalValue = self.config.getSignalLenAndSignalValue(signaltype,parameter)
                                message = '0x7E9: '+Mac+signalLen+'\n0x7EA: '+Mac+signalValue.upper()
                                break
                    else:
                        msg_box = QMessageBox(QMessageBox.Information, "提示", "软件接口测试中\n请输入正确的参数格式!")
                        msg_box.exec_()
                    self.tableWidget.setItem(i, 7, QTableWidgetItem(message))
                elif (self.tableWidget.cellWidget(i,0) and self.tableWidget.cellWidget(i,0).currentText()=='配置') and self.tableWidget.cellWidget(i,1).currentText()=='CAN':
                    message = '0x'+str('7FF')+': FF FF FF FF FF FF FF FF FF FF FF '
                    startbyte=''
                    signlen=''
                    startlength =''
                    dic = {'LowSpeedCAN':'0','CAN':'1','CANFD':'2','无效':'3','19.2kps':'0','9.6kps':'1'}
                    if parameter in dic:
                        value = dic[parameter]
                    else:
                        value=parameter
                    for k in range(len(self.candiclist)):
                        if projectname ==self.candiclist[k][0]:
                            startbyte = int(self.candiclist[k][1])
                            startlength = int(self.candiclist[k][2])
                            signlen = int(self.candiclist[k][3])
                    config = CanLinConfig()
                    # print(startbyte,startlength,signlen,value)
                    if startlength!="" and signlen!='' and startlength!='':
                        a = config.getConfigBytesString(startbyte, startlength, signlen, value, 12)
                        message = '0x'+str('7FF')+': '+a
                    self.tableWidget.setItem(i, 7, QTableWidgetItem(message))
                # (self.tableWidget.cellWidget(i,1).currentText(),message)
        except:
            logging.error(traceback.format_exc() + "\n")

    def genAutoExcel(self):
        try:
            if self.tabWidget.currentIndex()==0:
                headnameList = ['序号', '操作类型', '操作名称', '操作值', '间隔（ms）', 'Cycle（ms）', 'canID', 'Start', 'Length','flag','format']
                datalist = []
                self.j = 0
                for i in range(0,self.tableWidget.rowCount()):
                    combox_temp = self.tableWidget.cellWidget(i, 1)
                    waittime = self.getTabelCellData(i, 6)
                    param = self.getTabelCellData(i,4)
                    try:
                        projectname = self.tableWidget.cellWidget(i, 2).currentText()
                    except:
                        try:
                            projectname = self.tableWidget.item(i, 2).text()
                        except:
                            projectname = ''
                    startBit=''
                    length=''
                    if projectname in self.datalist:
                        for k in range(len(self.diclist)):
                            if projectname == self.diclist[k][1]:
                                # BCAN,信号名称,精度，偏移量，起始位，信号长度，起始字节
                                length = self.diclist[k][5]
                                startBit = self.diclist[k][4]
                    elif projectname in self.linlist:
                        for k in range(len(self.lindic)):
                            if projectname == self.lindic[k][1]:
                                startBit = self.lindic[k][4]
                                length = self.lindic[k][5]
                    check = self.getTabelCellData(i, 3)
                    if check =='大于':
                        checktxt='greaterCheck'
                    elif check=='小于':
                        checktxt='lessCheck'
                    elif check=='小于等于':
                        checktxt ='lessEqualCheck'
                    elif check =='大于等于':
                        checktxt='greaterEqualCheck'
                    else:
                        checktxt='check'
                    self.j =self.j+1
                    if projectname!='':
                        if self.tableWidget.cellWidget(i, 0).currentText()=='期望':
                            if self.tableWidget.item(i,8) and self.tableWidget.item(i,8).text()!='':
                                num = int(self.tableWidget.item(i,8).text())
                                messages = self.tableWidget.item(i, 7).text().split('\n')
                                backnum =messages[1].split(' ')[num]
                                data1 = ['event', projectname, messages[0].split(':')[1],
                                         waittime, '1', messages[0].split(':')[0], '', '', '--', 'Motorola'
                                         ]
                                datalist.append(data1)
                                data2 = ['check', projectname, backnum,
                                         waittime, '100', messages[1].split(':')[0], num*8, '8', '--', 'Motorola'
                                         ]
                                datalist.append(data2)
                            else:
                                needList=['电压输入','开关输入','PWM输入']
                                messages = self.tableWidget.item(i, 7).text().split('\n')
                                if self.tableWidget.cellWidget(i, 1).currentText() in needList:
                                    if self.tableWidget.cellWidget(i, 1).currentText()=='开关输入':
                                        num = self.openclose.index(projectname)+1
                                        num,startBit,length=self.getfile.getBit('开关输入',param,num,self.filename)
                                        data2 = [checktxt, projectname, num,
                                                 waittime, '100', messages[0].split(':')[0], startBit, length, '--', 'Motorola'
                                                 ]
                                    elif self.tableWidget.cellWidget(i, 1).currentText() == '电压输入':
                                        num = self.Vinput.index(projectname) + 1
                                        num, startBit, length = self.getfile.getBit('电压输入', param, num, self.filename)
                                        data2 = [checktxt, projectname, num,
                                                 waittime, '100', messages[0].split(':')[0], startBit, length, '--', 'Motorola'
                                                 ]
                                    elif self.tableWidget.cellWidget(i, 1).currentText() == 'PWM输入':
                                        num = self.inputPWM.index(projectname) + 1
                                        num, startBit, length = self.getfile.getBit('PWM输入', param, num, self.filename)
                                        data2 = [checktxt, projectname, num,
                                                 waittime, '100', messages[0].split(':')[0], startBit, length, '--', 'Motorola'
                                                 ]
                                    else:
                                        data2=[]
                                    if data2!=[]:
                                        datalist.append(data2)
                                else:
                                    data1 = ['event', projectname, messages[0].split(':')[1],
                                             waittime, '1', messages[0].split(':')[0], '', '', '--', 'Motorola'
                                             ]
                                    datalist.append(data1)
                                    if self.tableWidget.cellWidget(i, 1).currentText() == '软件接口测试':
                                        data2 = [checktxt, projectname, param,
                                                 waittime, '100', messages[1].split(':')[0], 64, 32, '--', 'Motorola'
                                                 ]
                                    elif projectname in self.datalist+self.linlist:
                                        data2 = [checktxt, projectname, combox_temp.canOrLinValue,
                                                 waittime, '100', messages[1].split(':')[0], startBit, length, '--', 'Motorola'
                                                 ]
                                    else:
                                        data2=[]
                                    if data2!=[]:
                                        datalist.append(data2)
                        elif '\n' in self.tableWidget.item(i,7).text():
                            messages = self.tableWidget.item(i, 7).text().split('\n')
                            data1 = [ 'event', projectname, messages[0].split(':')[1],
                                     waittime, '1', messages[0].split(':')[0], '','', '--', 'Motorola'
                                     ]
                            datalist.append(data1)
                            data2 = [ 'event', projectname, messages[1].split(':')[1],
                                     waittime, '1', messages[1].split(':')[0], '', '', '--', 'Motorola'
                                     ]
                            datalist.append(data2)
                        else:
                            data = ['event',projectname,self.tableWidget.item(i,7).text().split(':')[1],
                                    waittime,'1',self.tableWidget.item(i,7).text().split(':')[0],'','','--','Motorola']
                            datalist.append(data)
                # print(datalist)
                if datalist != []:
                    try:
                        workbook = Workbook('./自动测试.xlsx')
                        sheet1 = workbook.add_worksheet('CAN')
                        sheet1.set_column('C:D',30)
                        sheet1.set_column('B:B',15)
                        sheet1.set_column('E:K',10)
                        font = workbook.add_format({'font_name': '等线', 'font_size': 12,'align': 'center'})
                        for k in range(len(headnameList)):
                            sheet1.write(0, k, headnameList[k],font)
                        self.row = 1
                        self.SN = 0
                        for i in range(1, len(datalist) + 1):
                            if 'event' in datalist[i-1][0] and datalist[i-1][5].startswith('0x7'):
                                hex_number = datalist[i-1][5][-2:]
                                decimal_number = int(hex_number, 16)
                                self.SN = self.SN + 1
                                sheet1.write(self.row, 0, str(self.SN), font)
                                for j in range(len(datalist[i - 1])):
                                    sheet1.write(self.row, j+1,datalist[i - 1][j],font)
                                self.row=self.row+1
                                # formlist = [[str(self.SN+1),'check','单条测试用例处理状态1','0',datalist[i - 1][3],'100','0x710','0','8','--','Motorola'],
                                #             [str(self.SN+1),'check','单条测试用例处理状态2','7',datalist[i - 1][3],'100','0x710','8','8','--','Motorola'],
                                #             [str(self.SN+1),'check','单条测试用例处理状态3',decimal_number,datalist[i - 1][3],'100','0x710','16','8','--','Motorola']]
                                formlist = [
                                    [str(self.SN + 1), 'check', '单条测试用例处理状态', decimal_number, datalist[i - 1][3], '100',
                                     '0x710', '16', '8', '--', 'Motorola']]
                                for l in range(len(formlist)):
                                    for n in range(len(formlist[l])):
                                        sheet1.write(self.row,n,formlist[l][n],font)
                                    self.row=self.row+1
                                self.SN=self.SN+1
                            else:
                                self.SN = self.SN + 1
                                sheet1.write(self.row,0,str(self.SN),font)
                                for j in range(len(datalist[i - 1])):
                                    sheet1.write(self.row, j+1,datalist[i - 1][j],font)
                                self.row=self.row+1
                        workbook.close()
                        msg_box = QMessageBox(QMessageBox.Information, "标题", "导出成功!")
                        msg_box.exec_()
                    except:
                        msg_box = QMessageBox(QMessageBox.Information, "提示", "自动测试表格正在打开，请关闭！")
                        msg_box.exec_()
            elif self.tabWidget.currentIndex()==1:
                aa= InterfaceTest()
                aa.elfAnalysis= self.elfAnalysis
                aa.readMesseage(self.filename,self.elfpath)
            elif self.tabWidget.currentIndex()==2:
                aa= RoutingTest()
                aa.readMesseage(self.filename)
            elif self.tabWidget.currentIndex()==3:
                aa = SignalRoutingTest()
                aa.readMesseage(self.filename, './信号路由自动测试.xlsx', './硬件输入的信号路由测试.xlsx')
        except:
            logging.error(traceback.format_exc() + "\n")

    def merge_cells_by_coords(self,ws, start_row, start_column, end_row, end_column,data):
        cell_range = f'{get_column_letter(start_column)}{start_row}:{get_column_letter(end_column)}{end_row}'
        ws.merge_cells(cell_range)
        ws['A'+str(start_row)]= data


    def datatotable(self):
        try:
            testdata = []
            for i in range(0, self.tableWidget.rowCount()):
                datalist = []
                if self.tableWidget.cellWidget(i, 0) and self.tableWidget.cellWidget(i, 0).currentText() != "":
                    for j in range(0, self.tableWidget.columnCount() - 1):
                        try:
                            data = self.tableWidget.cellWidget(i, j).currentText()
                        except:
                            try:
                                data = self.tableWidget.cellWidget(i, j).text()
                            except:
                                try:
                                    data = self.tableWidget.item(i, j).text()
                                except:
                                    data = ''
                        datalist.append(data)
                elif self.tableWidget.item(i, 0) and self.tableWidget.item(i, 0).text().startswith('章节'):
                    datalist.append(self.tableWidget.item(i, 0).text())
                    datalist.append(self.getTabelCellData(i, 1))
                elif self.tableWidget.item(i, 0):
                    datalist.append(self.tableWidget.item(i, 0).text())
                testdata.append(datalist)
            if testdata!=[]:
                self.tableWidget.setRowCount(0)
                row = self.tableWidget.rowCount()
                for k in range(len(testdata)):
                    values = testdata[k]
                    if values and values[0].startswith('章节'):
                        self.addchapter(0)
                        item = self.tableWidget.cellWidget(row + k, 1)
                        item.setEditText(QtCore.QCoreApplication.translate("MainWindow", values[1]))
                    elif values and values[0] not in self.comboBoxList1:
                        note = values[0]
                        self.addnote(0)
                        self.tableWidget.setItem(row + k , 0, QTableWidgetItem(note))
                        brush = QBrush(QColor(230, 230, 230))
                        self.tableWidget.item(row + k , 0).setBackground(brush)
                    elif values==[]:
                        self.addrow(0)
                    else:
                        self.addrow(0)
                        if testdata[k][1] in self.CAN:
                            comboBox1 = QComboboxEx()
                            comboBox1.index = row + k
                            self.add_comobox(row + k, 2, self.datalist, comboBox1)
                            comboBox1.currentTextChanged.connect(lambda: self.comboBoxChange1(comboBox1))
                        elif testdata[k][1] in self.LIN:
                            comboBox2 = QComboboxEx()
                            comboBox2.index = row + k
                            self.add_comobox(row + k, 2, self.linlist, comboBox2)
                            comboBox2.currentTextChanged.connect(lambda: self.comboBoxChange1(comboBox2))
                        for j in range(len(testdata[k])):
                            try:
                                item = self.tableWidget.cellWidget(row + k , j)
                                item.setEditText(QtCore.QCoreApplication.translate("MainWindow", values[j]))
                            except:
                                try:
                                    item = self.tableWidget.cellWidget(row + k, j)
                                    item.setText(values[j])
                                except:
                                    try:
                                        self.tableWidget.setItem(row + k, j, QTableWidgetItem(values[j]))
                                    except:
                                        pass
        except:
            logging.error(traceback.format_exc() + "\n")
    # 使用函数合并指定的单元格范围

    def tableToexcel(self,filepath):
        headnameList = ['步骤', '类别', '项目', '动作', '参数', '持续', '等待', '报文解析', '结果', '操作']
        testdata = []
        for i in range(0, self.tableWidget.rowCount()):
            datalist = []
            if self.tableWidget.cellWidget(i, 0) and self.tableWidget.cellWidget(i, 0).currentText() != "":
                for j in range(0, self.tableWidget.columnCount() - 1):
                    try:
                        data = self.tableWidget.cellWidget(i, j).currentText()
                    except:
                        try:
                            data = self.tableWidget.cellWidget(i, j).text()
                        except:
                            try:
                                data = self.tableWidget.item(i, j).text()
                            except:
                                data = ''
                    datalist.append(data)
            elif self.tableWidget.item(i, 0) and self.tableWidget.item(i, 0).text().startswith('章节'):
                datalist.append(self.tableWidget.item(i, 0).text() + ':' + self.getTabelCellData(i, 1))
            elif self.tableWidget.item(i, 0) and self.tableWidget.item(i, 0).text() != "":
                datalist.append(self.tableWidget.item(i, 0).text())
                # if '备注' in self.tableWidget.item(i,0).text():
                #     datalist.append(self.tableWidget.item(i,0).text())
                # else:
                #     datalist.append('备注：'+self.tableWidget.item(i,0).text())
            if datalist != []:
                testdata.append(datalist)

        # 测试用例.xlsx
        if testdata != []:
            wb = openpyxl.Workbook()
            wb.create_sheet(title='CAN', index=0)
            sheet1 = wb.active
            font = Font(name='等线', size=12)
            alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet1.column_dimensions['B'].width = 15
            sheet1.column_dimensions['C'].width = 50
            sheet1.column_dimensions['H'].width = 50
            sheet1.row_dimensions[1].height = 40
            for k in range(len(headnameList) - 1):
                sheet1.cell(row=1, column=k + 1, value=headnameList[k])
            for i in range(1, len(testdata) + 1):
                sheet1.row_dimensions[i + 1].height = 40
                if testdata[i - 1][0] not in self.comboBoxList1:
                    self.merge_cells_by_coords(sheet1, i + 1, 1, i + 1, 9, testdata[i - 1][0])
                else:
                    for j in range(len(testdata[i - 1])):
                        sheet1.cell(row=i + 1, column=j + 1, value=testdata[i - 1][j])
            for row in sheet1.iter_rows():
                for cell in row:
                    cell.font = font
                    cell.alignment = alignment
            wb.save(filepath)
    def saveToexcel(self):
        try:
            filename=QFileDialog.getSaveFileName(self,'导出测试用例','','Excel File(*.xlsx)')
            if filename[0]:
                self.tableToexcel(filename[0])
                msg_box = QMessageBox(QMessageBox.Information, "标题", "导出成功!")
                msg_box.exec_()
        except:
            logging.error(traceback.format_exc() + "\n")
        # print(tabledata)

    def excelTotable(self):
        try:
            self.openfile = self.excelTotable or os.getcwd()
            path = QFileDialog.getOpenFileName(self, "选择文件", '/', "xlsx Files(*.xlsx)")
            if path[0]:
                book = xlrd2.open_workbook(path[0])
                sheet = book.sheet_by_name('CAN')
                row = self.tableWidget.rowCount()
                for i in range(1,sheet.nrows):
                    values =sheet.row_values(i)
                    if values and values[0].startswith('章节'):
                        self.addchapter(0)
                        item = self.tableWidget.cellWidget(row + i-1, 1)
                        item.setEditText(QtCore.QCoreApplication.translate("MainWindow", values[0].split('章节:')[1]))
                    elif values[0] not in self.comboBoxList1:
                        note = values[0]
                        self.addnote(0)
                        self.tableWidget.setItem(row+i-1,0,QTableWidgetItem(note))
                        brush = QBrush(QColor(230, 230, 230))
                        self.tableWidget.item(row+i-1, 0).setBackground(brush)
                    else:
                        self.addrow(0)
                        for j in range(len(values)):
                            if values[j] in self.CAN:
                                comboBox1 = QComboboxEx()
                                comboBox1.index = row + i - 1
                                self.add_comobox(row + i-1, 2, self.datalist, comboBox1)
                                comboBox1.currentTextChanged.connect(lambda :self.comboBoxChange1(comboBox1))
                            elif values[j] in self.LIN:
                                comboBox2 = QComboboxEx()
                                comboBox2.index = row + i - 1
                                self.add_comobox(row + i-1, 2, self.linlist, comboBox2)
                                comboBox2.currentTextChanged.connect(lambda: self.comboBoxChange1(comboBox2))
                            try:
                                item = self.tableWidget.cellWidget(row+i-1, j)
                                item.setEditText(QtCore.QCoreApplication.translate("MainWindow", values[j]))
                            except:
                                try:
                                    item = self.tableWidget.cellWidget(row + i - 1, j)
                                    item.setText(values[j])
                                except:
                                    try:
                                        self.tableWidget.setItem(row + i - 1, j, QTableWidgetItem(values[j]))
                                    except:
                                        pass
        except:
            logging.error(traceback.format_exc() + "\n")

    #用于刷新数据的绑定事件，仅限第三列can/lin项目
    def comboBoxChange1(self, comboBox):
        select = comboBox.currentText()
        index = comboBox.index
        if select in self.datalist:
            for j in range(len(self.signalD)):
                if select == self.signalD[j][0] and self.signalD[j][1:]!='':
                    comboBox5 = QComboboxEx()
                    self.add_comobox(index, 4, self.signalD[j][1:], comboBox5)
                    break
                else:
                    self.tableWidget.removeCellWidget(index, 4)
        elif select in self.linlist:
            for j in range(len(self.linD)):
                if select == self.linD[j][0] and self.linD[j][1:]!='':
                    comboBox5 = QComboboxEx()
                    self.add_comobox(index, 4, self.linD[j][1:], comboBox5)
                    break
                else:
                    self.tableWidget.removeCellWidget(index, 4)

    def addchapter(self,row):
        if row==0:
            rowPosition = self.tableWidget.rowCount()
        else:
            rowPosition=row
        comboBox = QComboboxEx()
        palette= QPalette()
        palette.setColor(QPalette.Base,QColor(230,230,230))
        comboBox.setPalette(palette)
        col = self.tableWidget.columnCount()
        self.tableWidget.insertRow(rowPosition)
        self.tableWidget.setSpan(rowPosition,1,1,8)
        self.add_comobox(rowPosition,1,self.chapterList,comboBox)
        deleteButton = QPushButton("删除".format(rowPosition))
        deleteButton.clicked.connect(self.delete_clicked)
        self.tableWidget.setCellWidget(rowPosition, 9, deleteButton)
        self.tableWidget.setItem(rowPosition,0,QTableWidgetItem("章节"))
        brush = QBrush(QColor(230, 230, 230))
        for i in range(col):
            item = self.tableWidget.item(rowPosition,i)
            if item:
                item.setBackground(brush)
            else:
                self.tableWidget.setItem(rowPosition,i,QTableWidgetItem(""))
                self.tableWidget.item(rowPosition, i).setBackground(brush)

    def addchapter1(self, row):
        try:
            self.addchapter(row)
            self.datatotable()
            self.tableWidget.verticalScrollBar().setSliderPosition(row)
        except:
            logging.error(traceback.format_exc() + "\n")
    def addnote(self,row):
        if row==0:
            rowPosition = self.tableWidget.rowCount()
        else:
            rowPosition=row
        brush = QBrush(QColor(230,230,230))
        col = self.tableWidget.columnCount()
        self.tableWidget.insertRow(rowPosition)
        for i in range(col):
            item = self.tableWidget.item(rowPosition,i)
            if item:
                item.setBackground(brush)
            else:
                self.tableWidget.setItem(rowPosition,i,QTableWidgetItem(""))
                self.tableWidget.item(rowPosition, i).setBackground(brush)
        self.tableWidget.setSpan(rowPosition,0,1,9)
        deleteButton = QPushButton("删除".format(rowPosition))
        deleteButton.clicked.connect(self.delete_clicked)
        self.tableWidget.setCellWidget(rowPosition, 9, deleteButton)

    def addnote1(self,row):
        try:
            if row == 0:
                rowPosition = self.tableWidget.rowCount()
            else:
                rowPosition = row
            self.addnote(row)
            self.datatotable()
            self.tableWidget.verticalScrollBar().setSliderPosition(rowPosition)
        except:
            logging.error(traceback.format_exc() + "\n")

    def addrow1(self,row):
        try:
            self.addrow(row)
            self.datatotable()
            self.tableWidget.verticalScrollBar().setSliderPosition(row)
        except:
            logging.error(traceback.format_exc() + "\n")

    def addrow(self,row):
        try:
            LineEdit=QLineEdit()
            LineEdit.setStyleSheet("QLineEdit {"
                                    "border: none;"  # 无边框
                                    "font: 10pt 'kaiti';"  # 设置字体大小和类型
                                    "}"
                                    )
            comboBox1 = QComboboxEx()
            comboBox2 = QComboboxEx()
            comboBox3 = QComboboxEx()
            self.comboBox4 = QComboboxEx()
            self.comboBox5 = QComboboxEx()
            self.comboBoxList1 =['前提','触发','期望','配置']
            comboBoxList2 = self.datalist
            self.comboBoxList4 =['CAN','LIN','温度传感器','继电器输出','电压传感器','PWM输出','软件接口测试','半桥输出','电压输入','开关输入','PWM输入','诊断服务']
            if row ==0 :
                rowPosition = self.tableWidget.rowCount()
            else:
                rowPosition = row
            self.tableWidget.insertRow(rowPosition)
            self.tableWidget.setRowHeight(rowPosition,40)
            self.tableWidget.setCellWidget(rowPosition, 5, LineEdit)
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(QtCore.Qt.ItemIsEnabled)
            self.tableWidget.setItem(rowPosition, 7, item)
            deleteButton = QPushButton("删除".format(rowPosition))
            deleteButton.clicked.connect(self.delete_clicked)
            self.tableWidget.setCellWidget(rowPosition, 9, deleteButton)
            self.add_comobox(rowPosition,0,self.comboBoxList1,comboBox1)
            self.add_comobox(rowPosition, 1, self.comboBoxList4,comboBox2)
            self.add_comobox(rowPosition,3,self.comboBoxList3,comboBox3)
            # self.add_comobox(rowPosition, 4,self.signalD, self.comboBox5)
            self.tableWidget.cellWidget(rowPosition,0).currentTextChanged.connect(
                lambda: self.comboboxIndex1(rowPosition))
            self.tableWidget.cellWidget(rowPosition, 1).currentTextChanged.connect(
                lambda: self.comboBoxChange(rowPosition, self.tableWidget.cellWidget(rowPosition, 1)))
            self.tableWidget.verticalScrollBar().setSliderPosition(rowPosition)
        except:
            logging.error(traceback.format_exc() + "\n")

    #用第一列触发绑定事件（用于添加行）
    def comboboxIndex1(self,index):
        comboBox4 = QComboboxEx()
        qiwanngList = ['电压输入', '开关输入', 'PWM输入', 'CAN', 'LIN', '软件接口测试','诊断服务']
        select = self.tableWidget.cellWidget(index,0).currentText()
        if select=='期望':
            self.tableWidget.removeCellWidget(index, 1)
            comboBox = QComboboxEx()
            self.add_comobox(index, 1, qiwanngList, comboBox)
            comboBox.currentTextChanged.connect(lambda: self.comboBoxChange(index, comboBox))
        elif select == '配置' and self.tableWidget.cellWidget(index, 1).currentText() == 'CAN':
            self.add_comobox(index, 2, sorted(self.candiclist1[0][3:], key=str.lower), comboBox4)
            comboBox4.currentTextChanged.connect(lambda: self.comboBoxChange(index, comboBox4))
        elif select != '配置' and self.tableWidget.cellWidget(index, 1).currentText() == 'CAN':
            self.add_comobox(index, 2, self.datalist, comboBox4)
            comboBox4.currentTextChanged.connect(lambda: self.comboBoxChange(index, comboBox4))
        elif select=='配置':
            self.tableWidget.removeCellWidget(index, 1)
            comboBox = QComboboxEx()
            self.add_comobox(index, 1, ['CAN'], comboBox)
            comboBox.currentTextChanged.connect(lambda: self.comboBoxChange(index, comboBox))
        else:
            self.tableWidget.removeCellWidget(index, 1)
            comboBox = QComboboxEx()
            self.add_comobox(index, 1, self.comboBoxList4, comboBox)
            comboBox.currentTextChanged.connect(lambda: self.comboBoxChange(index, comboBox))



    def open_file(self):
        try:
            # 检查是否是第一次调用
            if self.first_open_file_call:
                QMessageBox.information(self, "温馨提示！",
                                        "导入协议文件后，别忘了导入map文件，最后才是导入对应的测试文件（仅第一次导入文件时提醒）")
                self.first_open_file_call = False  # 设置标志变量为False，表示已经显示过了

            self.openfile = self.open_file or os.getcwd()
            fname = QFileDialog.getOpenFileName(self, "选择文件", '/', "xlsx Files(*.xlsx)")
            self.filename = fname[0]

            if not self.filename:
                QMessageBox.warning(self, "提示", "未选择任何文件！")
                return

            book = xlrd2.open_workbook(self.filename)
            required_sheets = [
                'CAN用例解析矩阵',
                'LIN用例解析矩阵',
                '测试盒CAN通讯矩阵',
                '电阻模拟输出表',
                '硬件配置表',
                '软件接口定义表'
            ]

            # 检查所有必需的工作表是否存在
            missing_sheets = [sheet for sheet in required_sheets if sheet not in book.sheet_names()]
            if missing_sheets:
                error_message = f"文件 '{self.filename}' 缺少以下工作表: {', '.join(missing_sheets)}\n导入失败！"
                logging.error(error_message)
                QMessageBox.critical(self, "导入失败", error_message)
                return

            # 如果所有必需的工作表都存在，则继续处理
            sheet = book.sheet_by_name('CAN用例解析矩阵')
            datalist = []
            for i in range(1, sheet.nrows):
                values = sheet.row_values(i)
                if values[25] != '':
                    item = values[25].replace("：", ':').replace('\r', '').replace(';', '').replace('；', '')
                    item = item.split('\n')
                    signalname = values[8].replace('\n', '').replace('\r', '')
                    item2 = [signalname] + item
                    self.signalD1.append(item2)
                    for j in range(len(item)):
                        if ':' in item[j]:
                            item[j] = item[j].split(":")[1]
                        else:
                            item.remove(item[j])
                    item1 = [signalname] + item
                    self.signalD.append(item1)
                if values[8] != '' and values[0] != '' and values[1] == '':
                    canN = values[0].replace('\r', '').replace('\n', '')
                    signalN = values[8].replace('\r', '').replace('\n', '')
                    data1 = [canN, signalN, values[15], values[16], values[11], values[13], values[10]]
                    datalist.append(signalN)
                    self.diclist.append(data1)
            datalist = list(set(datalist))
            self.datalist = sorted(datalist, key=str.lower)
            data1 = []
            for k in range(sheet.nrows - 1, 0, -1):
                values = sheet.row_values(k)
                if values[5] == '' and values[8] != '':
                    data1.append(values[8].replace('\r', '').replace('\n', ''))
                elif values[5] != "":
                    self.Canname = values[5]
                    if data1 != [] and self.Canname != '':
                        self.diclist1.append([self.Canname, values[3], values[2]] + data1)
                        data1 = []

            sheet5 = book.sheet_by_name('LIN用例解析矩阵')
            linlist = []
            for i in range(1, sheet5.nrows):
                values = sheet5.row_values(i)
                if values[25] != '':
                    item = values[25].replace("：", ':').replace('\r', '').replace(';', '').replace('；', '')
                    item = item.split('\n')
                    signalname = values[8].replace('\n', '').replace('\r', '')
                    item2 = [signalname] + item
                    self.linD1.append(item2)
                    for j in range(len(item)):
                        if ':' in item[j]:
                            item[j] = item[j].split(":")[1]
                        else:
                            item.remove(item[j])
                    item1 = [signalname] + item
                    self.linD.append(item1)
                if values[8] != '' and values[0] != '' and values[1] == '':
                    linN = values[0].replace('\r', '').replace('\n', '')
                    signalN = values[8].replace('\r', '').replace('\n', '')
                    data1 = [linN, signalN, values[15], values[16], values[11], values[13], values[10]]
                    linlist.append(signalN)
                    self.lindic.append(data1)
            linlist = list(set(linlist))
            self.linlist = sorted(linlist, key=str.lower)
            data2 = []
            for k in range(sheet5.nrows - 1, 0, -1):
                values = sheet5.row_values(k)
                if values[6] == '' and values[8] != '':
                    data2.append(values[8].replace('\r', '').replace('\n', ''))
                elif values[6] != "":
                    self.Linname = values[6]
                    if data2 != [] and self.Linname != '':
                        self.lindic1.append([self.Linname, values[3], values[2]] + data2)
                        data2 = []

            sheet2 = book.sheet_by_name('测试盒CAN通讯矩阵')
            self.candata = []
            self.cansignal1 = []
            self.cansignal = []
            self.candiclist = []
            self.candiclist1 = []
            candata1 = []
            Vinput1 = []
            for m in range(sheet2.nrows - 1, 0, -1):
                values = sheet2.row_values(m)
                if values[5] == '' and values[8] != '':
                    candata1.append(values[8].replace('\r', '').replace('\n', ''))
                elif values[5] != "":
                    self.Canname = values[5]
                    if candata1 != [] and self.Canname != '':
                        self.candiclist1.append([values[0], self.Canname, values[3]] + candata1)
                        candata1 = []
            for n in range(1, sheet2.nrows):
                value = sheet2.row_values(n)
                if value[8] in self.candiclist1[0]:
                    self.candata.append(value[8])
                    canlist = value[25].replace('\r', '').split('\n')
                    data = [value[8]]
                    data2 = [value[8]]
                    for l in canlist:
                        if ":" in l and l != '':
                            data.append(l)
                            param = l.split(':')[1]
                            data2.append(param)
                    self.cansignal1.append(data)
                    self.cansignal.append(data2)
                    signalN = value[8].replace('\r', '').replace('\n', '')
                    data1 = [signalN, value[10], value[11], value[13]]
                    self.candiclist.append(data1)
            self.candata = list(set(self.candata))

            sheet6 = book.sheet_by_name('电阻模拟输出表')
            for k in range(0, sheet6.ncols, 2):
                dic = {}
                temp = sheet6.col_values(k)[4:]
                ch = sheet6.col_values(k + 1)[4:]
                for u in range(len(temp)):
                    dic[temp[u]] = ch[u]
                self.RTlist.append(dic)

            sheet3 = book.sheet_by_name('硬件配置表')
            self.ID = sheet3.row_values(2)[4:12]
            for l in range(2, 12):
                lst = sheet3.col_values(l)
                result = list(filter(None, lst))
                if l == 2:
                    self.CAN = result[2:]
                elif l == 3:
                    self.LIN = result[2:]
                elif l == 9:
                    self.openclose = result[2:-1]
                elif l == 4:
                    self.temp = result[2:-1]
                elif l == 5:
                    self.CV = result[2:-1]
                elif l == 6:
                    self.Vinput = result[2:-1]
                elif l == 7:
                    self.PWM = result[2:-1]
                elif l == 10:
                    self.outputPWM = result[2:-1]
                elif l == 11:
                    self.inputPWM = result[2:-1]
                elif l == 8:
                    self.banqiao = result[2:-1]
            lst = sheet3.col_values(18)
            result = list(filter(None, lst))
            self.comboBoxList3 = result[2:]

            sheet4 = book.sheet_by_name('软件接口定义表')
            self.interfaceList = [s.replace('\n', '').replace('\t', '').strip() for s in sheet4.col_values(2)[1:]]
            for k in range(1, sheet4.nrows):
                values = [str(s).replace('\n', '').replace('\t', '').strip() for s in sheet4.row_values(k)]
                self.interfaceDic.append(values)

            # 成功导入后的提示框
            QMessageBox.information(self, "导入成功", f"文件 '{self.filename}' 导入成功！")

        except Exception as e:
            logging.error(f"文件导入失败：{traceback.format_exc()}")
            QMessageBox.critical(self, "导入失败", f"文件 '{self.filename}' 导入失败！\n错误详情：{str(e)}")



    def clear_row(self):
        try:
            response = QMessageBox.question(self,'确认','确定清空数据吗？')
            if response==QMessageBox.Yes:
                self.tableWidget.setRowCount(0)
                self.tableWidget.clearContents()
            else:
                pass
        except:
            logging.error(traceback.format_exc() + "\n")


    def delete_clicked(self):
        button = self.sender()
        if button:
            row = self.tableWidget.indexAt(button.pos()).row()
            self.tableWidget.removeRow(row)
            self.datatotable()
            self.tableWidget.verticalScrollBar().setSliderPosition(row)
            # self.tableWidget.hideRow(row)
            # item = self.tableWidget.cellWidget(row , 0)
            # item.setEditText(QtCore.QCoreApplication.translate("MainWindow", ''))


    def add_comobox(self, i,j, comboBoxList,comboBox):
        try:
            comboBox.setEditable(True)
            self.completer = QCompleter(comboBoxList)
            self.completer.setFilterMode(Qt.MatchContains)
            self.completer.setCompletionMode(QCompleter.PopupCompletion)
            comboBox.addItems(comboBoxList)
            comboBox.setCompleter(self.completer)
            # comboBox.setLineEdit(QLineEdit(comboBox))
            # comboBox.lineEdit().setReadOnly(True)
            self.tableWidget.setCellWidget(i, j, comboBox)
            comboBox.setCurrentIndex(-1)
        except:
            logging.error(traceback.format_exc() + "\n")

        # comboBox.currentTextChanged.connect(lambda :self.comboBoxChange(comboBox,i))
    # 第二列和第三列对应触发事件（用于添加行）
    def comboBoxChange(self,index,comboBox):
        try:
            qiwanngList=['电压输入','开关输入','PWM输入','CAN','LIN','软件接口测试']
            comboBox4 = QComboboxEx()
            canname = ''
            LINname=''
            select = comboBox.currentText()
            # index = self.tableWidget.currentRow(comboBox)
            if select == "":
                return
            elif select in self.datalist:
                self.tableWidget.removeCellWidget(index, 1)
                for j in range(len(self.signalD)):
                    if select==self.signalD[j][0] and self.signalD[j][1:]!='':
                        comboBox5 = QComboboxEx()
                        self.add_comobox(index, 4, self.signalD[j][1:], comboBox5)
                        break
                    else:
                        self.tableWidget.removeCellWidget(index, 4)
                for i in range(len(self.diclist)):
                    if self.diclist[i][1] == select:
                        canname= self.diclist[i][0]
                comboBox2 = QComboboxEx()
                self.add_comobox(index, 1, self.comboBoxList4, comboBox2)
                comboBox2.currentTextChanged.connect(lambda: self.comboBoxChange(index,comboBox2))
                comboBox2.setCurrentText(canname)
            elif select in self.linlist:
                self.tableWidget.removeCellWidget(index, 1)
                for j in range(len(self.linD)):
                    if select==self.linD[j][0] and self.linD[j][1:]!='':
                        comboBox5 = QComboboxEx()
                        self.add_comobox(index, 4, self.linD[j][1:], comboBox5)
                        break
                    else:
                        self.tableWidget.removeCellWidget(index, 4)
                for i in range(len(self.lindic)):
                    if self.lindic[i][1] == select:
                        LINname= self.lindic[i][0]
                comboBox2 = QComboboxEx()
                self.add_comobox(index, 1, self.comboBoxList4, comboBox2)
                comboBox2.currentTextChanged.connect(lambda: self.comboBoxChange(index,comboBox2))
                comboBox2.setCurrentText(LINname)
            elif select in self.candiclist1[0]:
                for i in range(len(self.cansignal)):
                    if select == self.cansignal[i][0]:
                        comboBox5 = QComboboxEx()
                        self.add_comobox(index, 4, self.cansignal[i][1:], comboBox5)
            elif select =='CAN'and self.tableWidget.cellWidget(index,0).currentText()!='配置':
                self.add_comobox(index, 2, self.datalist,comboBox4)
                comboBox4.currentTextChanged.connect(lambda: self.comboBoxChange(index,comboBox4))
            elif select =='CAN' and self.tableWidget.cellWidget(index, 0).currentText() == '配置':
                self.add_comobox(index, 2, sorted(self.candiclist1[0][3:],key=str.lower), comboBox4)
                comboBox4.currentTextChanged.connect(lambda :self.comboBoxChange(index,comboBox4))
            # elif select=='配置' and self.tableWidget.cellWidget(index,1).currentText() =='CAN':
            #     self.add_comobox(index,2,sorted(self.candiclist1[0][3:],key=str.lower),comboBox4)
            #     comboBox4.currentTextChanged.connect(lambda: self.comboBoxChange(index,comboBox4 ))
            # elif select!='配置' and self.tableWidget.cellWidget(index,1).currentText() =='CAN' :
            #     self.add_comobox(index, 2, self.datalist, comboBox4)
            #     comboBox4.currentTextChanged.connect(lambda: self.comboBoxChange(index,comboBox4))
            elif select =='LIN':
                self.add_comobox(index, 2, self.linlist, comboBox4)
                comboBox4.currentTextChanged.connect(lambda: self.comboBoxChange(index, comboBox4))
            elif select=='继电器输出':
                comboBox5 = QComboboxEx()
                state = ['ON','OFF']
                self.add_comobox(index, 2, self.PWM, comboBox4)
                self.add_comobox(index,4,state,comboBox5)
            elif select=='半桥输出':
                comboBox5 = QComboboxEx()
                switchInput = ['LSD', 'HSD', '高阻']
                self.add_comobox(index, 2, self.banqiao, comboBox4)
                self.add_comobox(index,4,switchInput,comboBox5)
            elif select=='温度传感器':
                self.add_comobox(index, 2, self.temp, comboBox4)
                self.tableWidget.removeCellWidget(index, 4)
                self.tableWidget.setItem(index, 4, QTableWidgetItem(''))
            elif select == '电压传感器':
                self.add_comobox(index, 2, self.CV, comboBox4)
                self.tableWidget.removeCellWidget(index, 4)
                self.tableWidget.setItem(index, 4, QTableWidgetItem(''))
            elif select=='电压输入':
                self.add_comobox(index, 2, self.Vinput, comboBox4)
                self.tableWidget.removeCellWidget(index, 4)
                self.tableWidget.setItem(index, 4, QTableWidgetItem(''))
            elif select=='开关输入':
                comboBox5 = QComboboxEx()
                switchInput = ['LSD', 'HSD', '高阻']
                self.add_comobox(index, 2, self.openclose, comboBox4)
                self.add_comobox(index, 4, switchInput, comboBox5)
            elif select=='PWM输出':
                self.add_comobox(index, 2, self.outputPWM, comboBox4)
                self.tableWidget.removeCellWidget(index, 4)
                self.tableWidget.setItem(index, 4, QTableWidgetItem(''))
            elif select=='PWM输入':
                self.add_comobox(index, 2, self.inputPWM, comboBox4)
                self.tableWidget.removeCellWidget(index, 4)
                self.tableWidget.setItem(index, 4, QTableWidgetItem(''))
            elif select=='软件接口测试':
                self.add_comobox(index,2,self.interfaceList,comboBox4)
                self.tableWidget.removeCellWidget(index, 4)
                self.tableWidget.setItem(index,4,QTableWidgetItem(''))
            elif select=='诊断服务':
                self.tableWidget.removeCellWidget(index, 2)
                self.tableWidget.removeCellWidget(index, 4)

        except:
            logging.error(traceback.format_exc() + "\n")

    def settable(self):
        headnameList =[' 步骤 ','         类别         ','项目','    动作    ','            参数             ','    持续    ',
                       '  等待  ','报文解析','    结果   ', '    操作   ']
        self.tableWidget.setColumnCount(len(headnameList))
        self.tableWidget.setRowCount(0)
        self.tableWidget.setHorizontalHeaderLabels(headnameList)
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # 根据空格设置列宽
        self.tableWidget.horizontalHeader().setSectionResizeMode(0,QHeaderView.Fixed)
        self.tableWidget.horizontalHeader().setSectionResizeMode(1, QHeaderView.Fixed)
        self.tableWidget.horizontalHeader().setSectionResizeMode(5, QHeaderView.Fixed)
        self.tableWidget.verticalHeader().setVisible(False)
        # self.tableWidget.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # self.tableWidget.setColumnWidth(6, 200)
        # self.tableWidget.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        # self.tableWidget.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        # 根据实际输入内容调整列宽 剩下两列平均分配（项目和报文解析）
        self.tableWidget.horizontalHeader().setSectionResizeMode(9, QHeaderView.ResizeToContents)
        self.tableWidget.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.tableWidget.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        # self.tableWidget.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
        self.tableWidget.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeToContents)
        self.tableWidget.horizontalHeader().setSectionResizeMode(8, QHeaderView.ResizeToContents)
        # self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)



if __name__ == '__main__':
    # main = MainWindow()
    # def handle_exception(exception):
    #     # 这里可以添加你想要执行的代码，比如日志记录
    #     main.tableToexcel(filepath='./测试用例（自动保存）.xlsx')
    # # 固定的，PyQt5程序都需要QApplication对象。sys.argv是命令行参数列表，确保程序可以双击运行
    # try:
    app = QApplication(sys.argv)
    # 初始化
    mainwindow = QMainWindow()
    ui_components = Ui_MainWindow()
    ui_components.setupUi(mainwindow)
    aa = MainWindow()
    aa.show()
    sys.exit(app.exec_())
    # except Exception as e:
    #     handle_exception(e)
