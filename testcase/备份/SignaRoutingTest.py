"""
#!/usr/bin/env python3.9
#-*- coding:utf-8 -*-
@Project:testcase
@File:SignaRoutingTest.py
@Author:XU AO
@Time:2025/4/22 17:48
"""
import os
import logging
import sys
import traceback

import openpyxl
import xlrd2
import csv

from PyQt5.QtCore import Qt
from PyQt5.QtGui import QCursor
from PyQt5 import QtWidgets, QtCore
from PyQt5.QtWidgets import QMessageBox, QApplication, QMainWindow, QFileDialog, QTableWidgetItem, QHeaderView, \
    QPushButton, QAction, QMenu, QComboBox
from xlsxwriter import Workbook

from Function import *
# from CanLinConfig import CanLinConfig
import tkinter as tk
from tkinter import messagebox

from Function.CanLinConfig import CanLinConfig


class SignalRoutingTest:
    config = CanLinConfig()
    routingList = []
    CANList = []
    banqiaoList = []
    logging.basicConfig(
        filename="log.log",
        filemode="w",
        datefmt="%Y-%m-%d %H:%M:%S %p",
        format="%(asctime)s - %(name)s - %(levelname)s - %(module)s: %(message)s",
        level=logging.DEBUG
    )

    def __init__(self, table_widget):
        self.table = table_widget

    def settable(self):
        headnameList = ['   Name    ', 'Hyt', 'CANFD', 'ID', 'dlc', 'cycle', 'segment', 'startbit', 'length',
                        '预留','预留','预留','    Name    ','Hyt','CANFD','ID','dlc','cycle','segment','startbit',
                        'length','预留','操作']
        self.table.setColumnCount(len(headnameList))
        self.table.setRowCount(0)
        self.table.verticalHeader().setVisible(False)
        self.table.setHorizontalHeaderLabels(headnameList)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(12, QHeaderView.ResizeToContents)

    def exceltoTable(self):
        try:
            self.openfile = self.exceltoTable or os.getcwd()
            path = QFileDialog.getOpenFileName(None, "选择文件", '/', "csv Files(*.csv)")
            if path[0]:
                row = self.table.rowCount()
                with open(path[0], 'r') as file:
                    reader = csv.reader(file)
                    for i, val in enumerate(reader):
                        values = list(val)
                        if values != '':
                            self.add_row()
                        for j in range(len(values)):
                            try:
                                text = str(int(values[j]))
                            except:
                                text = str(values[j])
                            self.table.setItem(row + i - 1, j, QTableWidgetItem(text))
                self.delete_last_row()
        except:
            logging.error(traceback.format_exc() + "\n")

    def delete_last_row(self):
        row_count = self.table.rowCount()
        if row_count > 0:
            self.table.removeRow(row_count - 1)

    def add_row(self):
        try:
            current_row_count = self.table.rowCount()
            self.table.insertRow(current_row_count)
            deleteButton = QPushButton("删除")
            deleteButton.clicked.connect(lambda: self.delete_clicked(current_row_count))  # 传递当前行号
            self.table.setCellWidget(current_row_count, 22, deleteButton)
        except Exception as e:
            QMessageBox.critical(None, "错误", f"添加行时发生错误: {str(e)}")

    # 修改删除函数
    def delete_clicked(self, row):
        if row >= 0:  # 确保行号有效
            self.table.removeRow(row)  # 删除行
            self.table.verticalScrollBar().setSliderPosition(row)  # 滚动条调整位置
        else:
            QMessageBox.warning(None, "警告", "删除失败，无法找到该行。")

    def clear_table(self):
        """清空表格数据和 routingList"""
        try:
            response = QMessageBox.question(None, '确认', '确定清空数据吗？')
            if response == QMessageBox.Yes:
                self.table.setRowCount(0)  # 清空表格行
                self.routingList.clear()  # 清空 routingList
                self.settable()  # 重新设置表格头
                # QMessageBox.information(self, "提示", "数据已清空！")
            else:
                pass

        except Exception as e:
            logging.error(f"清空数据时出错: {traceback.format_exc()}")
            QMessageBox.warning(None, "警告", f"清空数据时出错: {str(e)}")

    def export_to_excel(self):
        try:
            # 打开文件保存对话框，让用户选择保存路径
            filename, _ = QFileDialog.getSaveFileName(None, "导出测试用例", "", "Excel Files (*.xlsx)")
            if not filename:
                return  # 如果用户取消选择，直接返回

            # 创建 Excel 工作簿和工作表
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "导出数据"

            # 写入表头
            headers = ['   Name    ', 'Hyt', 'CANFD', 'ID', 'dlc', 'cycle', 'segment', 'startbit', 'length',
                        '预留','预留','预留','    Name    ','Hyt','CANFD','ID','dlc','cycle','segment','startbit',
                        'length','预留','操作']
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)

            # 写入表格数据
            for row_num in range(self.table.rowCount()):
                for col_num in range(self.table.columnCount()):
                    item = self.table.item(row_num, col_num)
                    if item is not None:
                        sheet.cell(row=row_num + 2, column=col_num + 1, value=item.text())

            # 保存文件
            workbook.save(filename)
            QMessageBox.information(None, "成功", f"数据已成功导出到 {filename}")
        except Exception as e:
            logging.error(f"导出到Excel时发生错误: {traceback.format_exc()}")
            QMessageBox.critical(None, "错误", "导出失败，请查看日志！")

    def show_success_dialog(self, state, message):
        # 创建一个简单的Tkinter窗口
        success_dialog = tk.Tk()
        success_dialog.withdraw()  # 隐藏主窗口

        # 显示成功消息
        messagebox.showinfo(state, message)

    def readMesseage(self, path2, path3, path4):
        try:
            for i in range(0, self.table.rowCount()):
                datalist = []
                for j in range(0, self.table.columnCount() - 1):
                    data = self.table.item(i, j).text()
                    datalist.append(data)
                self.routingList.append(datalist)
            book2 = xlrd2.open_workbook(path2)
            sheetbook2 = book2.sheet_by_name('硬件配置表')
            self.LINList = sheetbook2.col_values(3)[3:]
            self.CANList = sheetbook2.col_values(2)[3:]
            self.banqiaoList = sheetbook2.col_values(8)[3:]
            if self.genautotable(path3, path4):
                self.show_success_dialog('FAIL', '生成失败，没有找到相关属性！')
            else:
                self.show_success_dialog('Success', '生成成功！')
        except:
            self.show_success_dialog('FAIL', '生成失败，查看log！')
            logging.error(traceback.format_exc() + "\n")

    def genautotable(self, path3, path4):
        dataList = []
        dataListhard = []
        # 从 tableWidget 中读取数据
        for row in range(self.table.rowCount()):
            item = []
            for col in range(self.table.columnCount() - 1):  # 去掉最后一列（操作按钮）
                cell = self.table.item(row, col)
                if cell is not None:
                    item.append(cell.text())
                else:
                    item.append("")  # 如果单元格为空，填充空字符串
            if item[0] or item[1]:  # 如果 Name 或 Hyt 不为空，则处理该行
                if item[1] != '' and item[0]:
                    preview = []
                    length = '1' * (int(item[8]))
                    num = int(length, 2)
                    if int(item[2]) == 2:
                        try:
                            string = self.config.getMessage(item[7], item[8], num, int(item[4]), 'intel')
                        except:
                            logging.info('\n此条数据可能存在异常：\nname:' + item[0] + '     startbit:' + item[
                                7] + '     length:' + item[8] + '       value:' + str(num) + '     dlc:' + item[4] + '      encoding_order:intel\n')
                        try:
                            data1 = '0' + str(self.LINList.index(item[1])) + '01'
                        except:
                            data1 = '0000'
                        ID = '4' + item[3].replace('0x', '').replace('0X', '').zfill(3)
                    elif int(item[2]) == 1 or int(item[2]) == 0:
                        try:
                            string = self.config.getMessage(item[7], item[8], num, int(item[4]), 'Motorola')
                        except:
                            logging.info('\n此条数据可能存在异常：\nname:' + item[0] + '     startbit:' + item[
                                7] + '     length:' + item[8] + '      value:' + str(num) + '      dlc:' + item[4] + '     encoding_order:Motorola\n')
                        try:
                            data1 = '0' + str(self.CANList.index(item[1] + '_CAN')) + '00'
                        except:
                            data1 = '0000'
                        ID = '8' + item[3].replace('0x', '').replace('0X', '').zfill(3)
                    else:
                        try:
                            string = self.config.getMessage(item[7], item[8], num, int(item[4]), 'Motorola')
                        except:
                            logging.info('\n此条数据可能存在异常：\nname:' + item[0] + '     startbit:' + item[
                                7] + '     length:' + item[8] + '      value:' + str(num) + '      dlc:' + item[4] + '     encoding_order:Motorola\n')
                        try:
                            data1 = '0' + str(self.CANList.index(item[1] + '_CAN')) + '00'
                        except:
                            data1 = '0000'
                        ID = item[3].replace('0x', '').replace('0X', '').zfill(4)

                    if int(item[5]) < 20:
                        cycle = '20'
                    else:
                        cycle = item[5]

                    if int(item[2]) == 2:
                        name = item[1] + "_" + item[0]
                        jiange = '200'
                        newcycle = '0014'
                    else:
                        if item[0] != '/':
                            name = item[1] + '_CAN_' + item[3] + "_" + item[0]
                        else:
                            name = item[1] + '_CAN_' + item[3] + "_"
                        newcycle = str(hex(int(cycle) * 5)).upper()[2:].zfill(4)
                        jiange = int(cycle) * 5

                    data2 = ID + '0014'
                    data3 = str(hex(int(item[4]))).upper()[2:].zfill(4)
                    data4 = str(hex(int(item[7]))).upper()[2:].zfill(4)
                    data5 = str(hex(int(item[8]))).upper()[2:].zfill(4)
                    message1 = data1 + data2 + data3 + data4 + data5
                    Mac = ' '.join(message1[i:i + 2] for i in range(0, len(message1), 2)).upper()
                    data = ['event', name, Mac, '20', '3', '0x700', '', '', '--', 'Motorola']
                    dataList.append(data)
                    preview.append(data)
                    seconditem = ['check', '单条测试用例处理状态', '0', '20', '100', '0x710', '16', '8', '--','Motorola']
                    dataList.append(seconditem)
                    preview.append(seconditem)

                    t = '00' * (int(item[4]) + 1)
                    string1 = ' '.join(t[i:i + 2] for i in range(2, len(t), 2)).upper()
                    thirditem = ['event', name, string, '20', '3', item[3], '', '', '--', 'Motorola']
                    anthirditem = ['event', name, string1, '20', '3', item[3], '', '', '--', 'Motorola']
                    dataList.append(thirditem)
                    preview.append(anthirditem)

                    if int(item[14]) == 2:
                        name = item[13] + "_" + item[12]
                        try:
                            channel = '0' + str(self.LINList.index(item[13])) + '01'
                        except:
                            channel = '0000'
                        ID = '4' + item[15].replace('0x', '').replace('0X', '').zfill(3)
                    elif int(item[14]) == 1:
                        if item[12] != '/':
                            name = item[13] + '_CAN_' + item[15] + "_" + item[12]
                        else:
                            name = item[13] + '_CAN_' + item[15] + "_"
                        try:
                            channel = '0' + str(self.CANList.index(item[13] + '_CAN')) + '00'
                        except:
                            channel = '0000'
                        ID = '8' + item[15].replace('0x', '').replace('0X', '').zfill(3)
                    else:
                        if item[12] != '/':
                            name = item[13] + '_CAN_' + item[15] + "_" + item[12]
                        else:
                            name = item[13] + '_CAN_' + item[15] + "_"
                        try:
                            channel = '0' + str(self.CANList.index(item[13] + '_CAN')) + '00'
                        except:
                            channel = '0000'
                        ID = item[15].replace('0x', '').replace('0X', '').zfill(4)

                    if int(item[17]) < 20:
                        data3 = '0014'
                        cycle2 = '20'
                    else:
                        data3 = str(hex(int(item[17]))).upper()[2:].zfill(4)
                        cycle2 = item[17]

                    if int(item[14]) == 2:
                        jiange = '100'
                        data2 = '0014'
                    else:
                        data2 = str(hex(int(cycle2) * 5)).upper()[2:].zfill(4)
                        jiange = int(cycle2)

                    data4 = str(hex(int(item[19]))).upper()[2:].zfill(4)
                    data5 = str(hex(int(item[20]))).upper()[2:].zfill(4)
                    message2 = channel + ID + data2 + data3 + data4 + data5
                    Mac2 = ' '.join(message2[i:i + 2] for i in range(0, len(message2), 2)).upper()
                    forthitem = ['event', name, Mac2, '20', '3', '0x701', '', '', '--', 'Motorola']
                    dataList.append(forthitem)
                    preview.append(forthitem)

                    fifthitem = ['check', '单条测试用例处理状态', '1', jiange, '100', '0x710', '16', '8', '--','Motorola']
                    dataList.append(fifthitem)
                    preview.append(fifthitem)

                    relength = '1' * (int(item[20]))
                    renum = int(relength, 2)
                    sixth = ['check', name, renum, '20', '100', item[15], item[19], item[20], '--', 'Motorola']
                    ansixth = ['check', name, '0', '20', '100', item[15], item[19], item[20], '--', 'Motorola']
                    dataList.append(sixth)
                    preview.append(ansixth)

                    dataList = dataList + preview
                    text = item[0] + ' ' + item[1] + ' ' + item[3] + '      ' + item[12] + ' ' + item[13] + ' ' + item[
                        15]
                    note = ['测试用例说明', text, '12', '', '', '', '', '', '', '']
                    dataList.append(note)
                elif item[1] == '' and item[0] != '':
                    name = item[0]
                    data1 = "0" + str(self.banqiaoList.index(item[0])) + '01'
                    if item[9] == '1':
                        data2 = '0001'
                        check = '0'
                    else:
                        data2 = '0000'
                        check = '1'
                    data3 = 'FFFF0000'
                    message1 = data1 + data2 + data3
                    Mac = ' '.join(message1[i:i + 2] for i in range(0, len(message1), 2)).upper()
                    data = ['event', name, Mac, '100', '1', '0x705', '', '', '--', 'Motorola']
                    dataListhard.append(data)
                    seconditem = ['check', '单条测试用例处理状态', '5', '100', '100', '0x710', '16', '8', '--','Motorola']
                    dataListhard.append(seconditem)

                    if int(item[14]) == 2:
                        name = item[13] + "_" + item[12]
                        try:
                            channel = '0' + str(self.LINList.index(item[13])) + '01'
                        except:
                            channel = '0000'
                        ID = '4' + item[15].replace('0x', '').replace('0X', '').zfill(3)
                    elif int(item[14]) == 1:
                        if item[12] != '/':
                            name = item[13] + '_CAN_' + item[15] + "_" + item[12]
                        else:
                            name = item[13] + '_CAN_' + item[15] + "_"
                        try:
                            channel = '0' + str(self.CANList.index(item[13] + '_CAN')) + '00'
                        except:
                            channel = '0000'
                        ID = '8' + item[15].replace('0x', '').replace('0X', '').zfill(3)
                    else:
                        if item[12] != '/':
                            name = item[13] + '_CAN_' + item[15] + "_" + item[12]
                        else:
                            name = item[13] + '_CAN_' + item[15] + "_"
                        try:
                            channel = '0' + str(self.CANList.index(item[13] + '_CAN')) + '00'
                        except:
                            channel = '0000'
                        ID = item[15].replace('0x', '').replace('0X', '').zfill(4)

                    if int(item[17]) < 20:
                        data3 = '0014'
                        cycle2 = '20'
                    else:
                        data3 = str(hex(int(item[17]))).upper()[2:].zfill(4)
                        cycle2 = item[17]

                    if int(item[14]) == 2:
                        jiange = '100'
                        data2 = '0014'
                    else:
                        data2 = str(hex(int(cycle2) * 5)).upper()[2:].zfill(4)
                        jiange = int(cycle2)

                    data4 = str(hex(int(item[19]))).upper()[2:].zfill(4)
                    data5 = str(hex(int(item[20]))).upper()[2:].zfill(4)
                    message2 = channel + ID + data2 + data3 + data4 + data5
                    Mac2 = ' '.join(message2[i:i + 2] for i in range(0, len(message2), 2)).upper()
                    forthitem = ['event', name, Mac2, '20', '1', '0x701', '', '', '--', 'Motorola']
                    dataListhard.append(forthitem)

                    fifthitem = ['check', '单条测试用例处理状态', '1', jiange, '100', '0x710', '16', '8', '--','Motorola']
                    dataListhard.append(fifthitem)

                    ansixth = ['check', name, check, '20', '100', item[15], item[19], item[20], '--', 'Motorola']
                    dataListhard.append(ansixth)

                    data1 = "0" + str(self.banqiaoList.index(item[0])) + '01'
                    if item[9] == '0':
                        data2 = '0001'
                        check = '0'
                    else:
                        data2 = '0000'
                        check = '1'
                    data3 = 'FFFF0000'
                    message1 = data1 + data2 + data3
                    Mac = ' '.join(message1[i:i + 2] for i in range(0, len(message1), 2)).upper()
                    sevenitem = ['event', item[0], Mac, '100', '1', '0x705', '', '', '--', 'Motorola']
                    sixth = ['check', name, check, '20', '100', item[15], item[19], item[20], '--', 'Motorola']
                    dataListhard.append(sevenitem)
                    dataListhard.append(seconditem)
                    dataListhard.append(forthitem)
                    dataListhard.append(fifthitem)
                    dataListhard.append(sixth)
                    text = item[0] + ' ' + item[1] + ' ' + item[3] + '      ' + item[12] + ' ' + item[13] + ' ' + item[
                        15]
                    note = ['测试用例说明', text, '10', '', '', '', '', '', '', '']
                    dataListhard.append(note)

        # 检查数据是否为空
        has_dataList = bool(dataList)
        has_dataListhard = bool(dataListhard)

        if not has_dataList and not has_dataListhard:
            QMessageBox.warning(None, "警告", "没有可保存的数据！")
            return True

        # 弹窗提醒用户需要保存的文件数量
        if has_dataList and has_dataListhard:
            QMessageBox.information(None, "提示", "需要保存两个文件：信号路由自动测试和硬件测试。")
        elif has_dataList:
            QMessageBox.information(None, "提示", "需要保存一个文件：信号路由自动测试。")
        else:
            QMessageBox.information(None, "提示", "需要保存一个文件：硬件测试。")

        # 导出信号路由自动测试
        if dataList:
            headnameList = ['序号', '操作类型', '操作名称', '操作值', '间隔（ms）', 'Cycle（ms）', 'canID', 'Start',
                            'Length',
                            'flag', 'format']
            file_path1, _ = QFileDialog.getSaveFileName(None, '导出信号路由自动测试', '', 'Excel File(*.xlsx)')
            if file_path1:
                workbook = Workbook(file_path1)
                sheet1 = workbook.add_worksheet('CAN')
                sheet1.set_column('C:D', 50)
                sheet1.set_column('B:B', 15)
                sheet1.set_column('E:K', 10)
                font = workbook.add_format({'font_name': '等线', 'font_size': 12, 'align': 'center'})
                for k in range(len(headnameList)):
                    sheet1.write(0, k, headnameList[k], font)
                row = 1
                for l in range(len(dataList)):
                    sheet1.write(row, 0, str(row), font)
                    for j in range(len(dataList[l])):
                        sheet1.write(row, j + 1, dataList[l][j], font)
                    row += 1
                workbook.close()
                dataList.clear()
        if dataListhard:
            headnameList = ['序号', '操作类型', '操作名称', '操作值', '间隔（ms）', 'Cycle（ms）', 'canID', 'Start',
                            'Length',
                            'flag', 'format']
            # workbook = Workbook(path4)
            # 获取用户选择的文件路径
            file_path2, _ = QFileDialog.getSaveFileName(None, '导出硬件测试', '', 'Excel File(*.xlsx)')
            workbook = Workbook(file_path2)
            sheet1 = workbook.add_worksheet('CAN')
            sheet1.set_column('C:D', 50)
            sheet1.set_column('B:B', 15)
            sheet1.set_column('E:K', 10)
            font = workbook.add_format({'font_name': '等线', 'font_size': 12, 'align': 'center'})
            for k in range(len(headnameList)):
                sheet1.write(0, k, headnameList[k], font)
            row = 1
            for l in range(len(dataListhard)):
                sheet1.write(row, 0, str(row), font)
                for j in range(len(dataListhard[l])):
                    sheet1.write(row, j + 1, dataListhard[l][j], font)
                row = row + 1
            workbook.close()
            dataListhard.clear()
