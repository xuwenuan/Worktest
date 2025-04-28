import os
import logging
import sys
import traceback

import openpyxl
import xlrd2
import csv

from PyQt5.QtCore import Qt
from PyQt5.QtGui import QCursor
from PyQt5 import QtWidgets, QtCore
from PyQt5.QtWidgets import QMessageBox, QApplication, QMainWindow, QFileDialog, QTableWidgetItem, QHeaderView, \
    QPushButton, QAction, QMenu, QComboBox
from xlsxwriter import Workbook

from Function import *
# from CanLinConfig import CanLinConfig
import tkinter as tk
from tkinter import messagebox

from Function.CanLinConfig import CanLinConfig


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1300, 600)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout.setObjectName("gridLayout")
        self.tableWidget = QtWidgets.QTableWidget(self.centralwidget)
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)
        self.gridLayout.addWidget(self.tableWidget, 0, 0, 1, 5)

        # 添加清空按钮
        self.pushButton_clear = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_clear.setObjectName("pushButton_clear")
        self.gridLayout.addWidget(self.pushButton_clear, 1, 3, 1, 1)

        # 导出到Excel按钮
        self.pushButton_export = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_export.setObjectName("pushButton_export")
        self.gridLayout.addWidget(self.pushButton_export, 1, 4, 1, 1)

        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setObjectName("pushButton")
        self.gridLayout.addWidget(self.pushButton, 1, 1, 1, 1)
        self.pushButton_2 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_2.setObjectName("pushButton_2")
        self.gridLayout.addWidget(self.pushButton_2, 1, 0, 1, 1)
        self.pushButton3 = QtWidgets.QPushButton(self.centralwidget)
        self.gridLayout.addWidget(self.pushButton3, 1, 2, 1, 1)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 800, 23))
        self.menubar.setObjectName("menubar")
        self.menu = QtWidgets.QMenu(self.menubar)
        self.menu.setObjectName("menu")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "空调测试用例编写工具"))
        self.pushButton.setText(_translate("MainWindow", "导出自动测试文件"))
        self.pushButton_2.setText(_translate("MainWindow", "导入文件"))
        self.pushButton3.setText('添加行')
        self.pushButton_clear.setText('清空数据')  # 设置清空按钮的文本
        self.pushButton_export.setText(_translate("MainWindow", "导出测试用例"))


class SignalRoutingTest(QMainWindow, Ui_MainWindow):
    config = CanLinConfig()
    logging.basicConfig(
        filename="log.log",
        filemode="w",
        datefmt="%Y-%m-%d %H:%M:%S %p",
        format="%(asctime)s - %(name)s - %(levelname)s - %(module)s: %(message)s",
        level=logging.DEBUG
    )
    routingList = []
    CANList = []
    banqiaoList = []

    def __init__(self, parent=None):
        super(SignalRoutingTest, self).__init__(parent)
        self.setupUi(self)
        self.pushButton_2.clicked.connect(self.exceltoTable)
        self.pushButton_clear.clicked.connect(self.clear_table)  # 连接清空按钮的点击事件
        self.pushButton_export.clicked.connect(self.export_to_excel)
        self.settable()
        self.pushButton3.clicked.connect(lambda: self.addrow(0))

    def clear_table(self):
        """清空表格数据和 routingList"""
        try:
            response = QMessageBox.question(self, '确认', '确定清空数据吗？')
            if response == QMessageBox.Yes:
                self.tableWidget.setRowCount(0)  # 清空表格行
                self.routingList.clear()  # 清空 routingList
                self.settable()  # 重新设置表格头
                # QMessageBox.information(self, "提示", "数据已清空！")
            else:
                pass

        except Exception as e:
            logging.error(f"清空数据时出错: {traceback.format_exc()}")
            QMessageBox.warning(self, "警告", f"清空数据时出错: {str(e)}")

    def export_to_excel(self):
        try:
            # 打开文件保存对话框，让用户选择保存路径
            filename, _ = QFileDialog.getSaveFileName(self, "导出测试用例", "", "Excel Files (*.xlsx)")
            if not filename:
                return  # 如果用户取消选择，直接返回

            # 创建 Excel 工作簿和工作表
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "导出数据"

            # 写入表头
            headers = ['   Name    ', 'Hyt', 'CANFD', 'ID', 'dlc', 'cycle', 'segment', 'startbit', 'length',
                        '预留','预留','预留','    Name    ','Hyt','CANFD','ID','dlc','cycle','segment','startbit',
                        'length','预留','操作']
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)

            # 写入表格数据
            for row_num in range(self.tableWidget.rowCount()):
                for col_num in range(self.tableWidget.columnCount()):
                    item = self.tableWidget.item(row_num, col_num)
                    if item is not None:
                        sheet.cell(row=row_num + 2, column=col_num + 1, value=item.text())

            # 保存文件
            workbook.save(filename)
            QMessageBox.information(self, "成功", f"数据已成功导出到 {filename}")
        except Exception as e:
            logging.error(f"导出到Excel时发生错误: {traceback.format_exc()}")
            QMessageBox.critical(self, "错误", "导出失败，请查看日志！")

    def exceltoTable(self):
        try:
            self.openfile = self.exceltoTable or os.getcwd()
            path = QFileDialog.getOpenFileName(self, "选择文件", '/', "csv Files(*.csv)")
            if path[0]:
                row = self.tableWidget.rowCount()
                with open(path[0], 'r') as file:
                    reader = csv.reader(file)
                    for i,val in enumerate(reader):
                        values = list(val)
                        if values!='':
                            self.addrow(0)
                        for j in range(len(values)):
                            try:
                                text = str(int(values[j]))
                            except:
                                text = str(values[j])
                            self.tableWidget.setItem(row + i - 1, j, QTableWidgetItem(text))
                self.delete_last_row()
        except:
            logging.error(traceback.format_exc() + "\n")

    def delete_last_row(self):
        row_count = self.tableWidget.rowCount()
        if row_count > 0:
            self.tableWidget.removeRow(row_count - 1)


    def addrow(self, row):
        try:
            if row == 0:
                rowPosition = self.tableWidget.rowCount()
            else:
                rowPosition = row
            self.tableWidget.insertRow(rowPosition)
            deleteButton = QPushButton("删除".format(rowPosition))
            deleteButton.clicked.connect(self.delete_clicked)
            self.tableWidget.setCellWidget(rowPosition, self.tableWidget.columnCount()-1, deleteButton)
        except:
            logging.error(traceback.format_exc() + "\n")

    def delete_clicked(self):
        button = self.sender()
        if button:
            row = self.tableWidget.indexAt(button.pos()).row()
            self.tableWidget.removeRow(row)
            self.tableWidget.verticalScrollBar().setSliderPosition(row)

    def settable(self):
        headnameList = ['   Name    ', 'Hyt', 'CANFD', 'ID', 'dlc', 'cycle', 'segment', 'startbit', 'length',
                        '预留','预留','预留','    Name    ','Hyt','CANFD','ID','dlc','cycle','segment','startbit',
                        'length','预留','操作']
        self.tableWidget.setColumnCount(len(headnameList))
        self.tableWidget.setRowCount(0)
        self.tableWidget.verticalHeader().setVisible(False)
        self.tableWidget.setHorizontalHeaderLabels(headnameList)
        # self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.tableWidget.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.tableWidget.horizontalHeader().setSectionResizeMode(12, QHeaderView.ResizeToContents)
        # self.tableWidget.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        # self.tableWidget.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeToContents)
        # self.tableWidget.horizontalHeader().setSectionResizeMode(9, QHeaderView.ResizeToContents)
        # self.tableWidget.horizontalHeader().setSectionResizeMode(11, QHeaderView.ResizeToContents)
        # self.tableWidget.horizontalHeader().setSectionResizeMode(18, QHeaderView.ResizeToContents)

    def show_success_dialog(self,state,message):
        # 创建一个简单的Tkinter窗口
        success_dialog = tk.Tk()
        success_dialog.withdraw()  # 隐藏主窗口

        # 显示成功消息
        messagebox.showinfo(state, message)



    #添加的复制粘贴模块
    def contextMenuEvent(self, event):
        try:
        # 创建右键菜单
            menu = QMenu(self)
            row = self.tableWidget.rowAt(event.y())

        # 添加复制和粘贴功能
            copy_action = QAction("复制 (Ctrl+C)", self)
            paste_action = QAction("粘贴 (Ctrl+V)", self)

            copy_action.triggered.connect(self.copy_selected_cells)
            paste_action.triggered.connect(self.paste_cells)

            menu.addAction(copy_action)
            menu.addAction(paste_action)

        # 显示菜单
            menu.exec_(QCursor.pos())

        except Exception as e:
            pass

    def copy_selected_cells(self):
        """复制选中的单元格内容到剪贴板"""
        selected_ranges = self.tableWidget.selectedRanges()
        if not selected_ranges:
            return

        clipboard_text = ""
        for selected_range in selected_ranges:
            top_row, bottom_row = selected_range.topRow(), selected_range.bottomRow()
            left_col, right_col = selected_range.leftColumn(), selected_range.rightColumn()

        # 构造复制内容
            for row in range(top_row, bottom_row + 1):
                row_data = []
                for col in range(left_col, right_col + 1):
                    item = self.tableWidget.item(row, col)
                    if item:
                        row_data.append(item.text())
                    else:
                        widget = self.tableWidget.cellWidget(row, col)
                        if isinstance(widget, QComboBox):
                            row_data.append(widget.currentText())
                        else:
                            row_data.append("")
                clipboard_text += "\t".join(row_data) + "\n"

    # 将内容复制到剪贴板
        clipboard = QApplication.clipboard()
        clipboard.setText(clipboard_text.strip())

    def paste_cells(self):
        """将剪贴板内容粘贴到选中的单元格"""
        clipboard = QApplication.clipboard()
        clipboard_text = clipboard.text()
        if not clipboard_text:
            return

    # 解析剪贴板内容为二维数组
        rows = clipboard_text.split("\n")
        data = [row.split("\t") for row in rows]

    # 获取目标起始位置
        selected_ranges = self.tableWidget.selectedRanges()
        if not selected_ranges:
            return
        start_row = selected_ranges[0].topRow()
        start_col = selected_ranges[0].leftColumn()

    # 填充表格
        for i, row_data in enumerate(data):
            for j, cell_value in enumerate(row_data):
                target_row = start_row + i
                target_col = start_col + j

            # 动态扩展表格
                if target_row >= self.tableWidget.rowCount():
                    self.tableWidget.insertRow(target_row)
                if target_col >= self.tableWidget.columnCount():
                    self.tableWidget.insertColumn(target_col)

            # 设置单元格内容
                widget = self.tableWidget.cellWidget(target_row, target_col)
                if isinstance(widget, QComboBox):
                    widget.setCurrentText(cell_value)
                else:
                    item = QTableWidgetItem(cell_value)
                    self.tableWidget.setItem(target_row, target_col, item)

    def readMesseage(self, path2, path3, path4):
        try:
            for i in range(0, self.tableWidget.rowCount()):
                datalist = []
                for j in range(0, self.tableWidget.columnCount() - 1):
                    data = self.tableWidget.item(i, j).text()
                    datalist.append(data)
                self.routingList.append(datalist)
            book2 = xlrd2.open_workbook(path2)
            sheetbook2 = book2.sheet_by_name('硬件配置表')
            self.LINList=sheetbook2.col_values(3)[3:]
            self.CANList = sheetbook2.col_values(2)[3:]
            self.banqiaoList = sheetbook2.col_values(8)[3:]
            if self.genautotable(path3, path4):
                self.show_success_dialog('FAIL', '生成失败，没有找到相关属性！')
            else:
                self.show_success_dialog('Success', '生成成功！')
        except:
            self.show_success_dialog('FAIL', '生成失败，查看log！')
            logging.error(traceback.format_exc() + "\n")

    def genautotable(self, path3, path4):
        dataList = []
        dataListhard = []
        # 从 tableWidget 中读取数据
        for row in range(self.tableWidget.rowCount()):
            item = []
            for col in range(self.tableWidget.columnCount() - 1):  # 去掉最后一列（操作按钮）
                cell = self.tableWidget.item(row, col)
                if cell is not None:
                    item.append(cell.text())
                else:
                    item.append("")  # 如果单元格为空，填充空字符串
            if item[0] or item[1]:  # 如果 Name 或 Hyt 不为空，则处理该行
                if item[1] != '' and item[0]:
                    preview = []
                    length = '1' * (int(item[8]))
                    num = int(length, 2)
                    if int(item[2]) == 2:
                        try:
                            string = self.config.getMessage(item[7], item[8], num, int(item[4]), 'intel')
                        except:
                            logging.info('\n此条数据可能存在异常：\nname:' + item[0] + '     startbit:' + item[7] + '     length:' + item[8] + '       value:' + str(num) + '     dlc:' + item[4] + '      encoding_order:intel\n')
                        try:
                            data1 = '0' + str(self.LINList.index(item[1])) + '01'
                        except:
                            data1 = '0000'
                        ID = '4' + item[3].replace('0x', '').replace('0X', '').zfill(3)
                    elif int(item[2]) == 1 or int(item[2]) == 0:
                        try:
                            string = self.config.getMessage(item[7], item[8], num, int(item[4]), 'Motorola')
                        except:
                            logging.info('\n此条数据可能存在异常：\nname:' + item[0] + '     startbit:' + item[7] + '     length:' + item[8] + '      value:' + str(num) + '      dlc:' + item[4] + '     encoding_order:Motorola\n')
                        try:
                            data1 = '0' + str(self.CANList.index(item[1] + '_CAN')) + '00'
                        except:
                            data1 = '0000'
                        ID = '8' + item[3].replace('0x', '').replace('0X', '').zfill(3)
                    else:
                        try:
                            string = self.config.getMessage(item[7], item[8], num, int(item[4]), 'Motorola')
                        except:
                            logging.info('\n此条数据可能存在异常：\nname:' + item[0] + '     startbit:' + item[7] + '     length:' + item[8] + '      value:' + str(num) + '      dlc:' + item[4] + '     encoding_order:Motorola\n')
                        try:
                            data1 = '0' + str(self.CANList.index(item[1] + '_CAN')) + '00'
                        except:
                            data1 = '0000'
                        ID = item[3].replace('0x', '').replace('0X', '').zfill(4)

                    if int(item[5]) < 20:
                        cycle = '20'
                    else:
                        cycle = item[5]

                    if int(item[2]) == 2:
                        name = item[1] + "_" + item[0]
                        jiange = '200'
                        newcycle = '0014'
                    else:
                        if item[0] != '/':
                            name = item[1] + '_CAN_' + item[3] + "_" + item[0]
                        else:
                            name = item[1] + '_CAN_' + item[3] + "_"
                        newcycle = str(hex(int(cycle) * 5)).upper()[2:].zfill(4)
                        jiange = int(cycle) * 5

                    data2 = ID + '0014'
                    data3 = str(hex(int(item[4]))).upper()[2:].zfill(4)
                    data4 = str(hex(int(item[7]))).upper()[2:].zfill(4)
                    data5 = str(hex(int(item[8]))).upper()[2:].zfill(4)
                    message1 = data1 + data2 + data3 + data4 + data5
                    Mac = ' '.join(message1[i:i + 2] for i in range(0, len(message1), 2)).upper()
                    data = ['event', name, Mac, '20', '3', '0x700', '', '', '--', 'Motorola']
                    dataList.append(data)
                    preview.append(data)
                    seconditem = ['check', '单条测试用例处理状态', '0', '20', '100', '0x710', '16', '8', '--', 'Motorola']
                    dataList.append(seconditem)
                    preview.append(seconditem)

                    t = '00' * (int(item[4]) + 1)
                    string1 = ' '.join(t[i:i + 2] for i in range(2, len(t), 2)).upper()
                    thirditem = ['event', name, string, '20', '3', item[3], '', '', '--', 'Motorola']
                    anthirditem = ['event', name, string1, '20', '3', item[3], '', '', '--', 'Motorola']
                    dataList.append(thirditem)
                    preview.append(anthirditem)

                    if int(item[14]) == 2:
                        name = item[13] + "_" + item[12]
                        try:
                            channel = '0' + str(self.LINList.index(item[13])) + '01'
                        except:
                            channel = '0000'
                        ID = '4' + item[15].replace('0x', '').replace('0X', '').zfill(3)
                    elif int(item[14]) == 1:
                        if item[12] != '/':
                            name = item[13] + '_CAN_' + item[15] + "_" + item[12]
                        else:
                            name = item[13] + '_CAN_' + item[15] + "_"
                        try:
                            channel = '0' + str(self.CANList.index(item[13] + '_CAN')) + '00'
                        except:
                            channel = '0000'
                        ID = '8' + item[15].replace('0x', '').replace('0X', '').zfill(3)
                    else:
                        if item[12] != '/':
                            name = item[13] + '_CAN_' + item[15] + "_" + item[12]
                        else:
                            name = item[13] + '_CAN_' + item[15] + "_"
                        try:
                            channel = '0' + str(self.CANList.index(item[13] + '_CAN')) + '00'
                        except:
                            channel = '0000'
                        ID = item[15].replace('0x', '').replace('0X', '').zfill(4)

                    if int(item[17]) < 20:
                        data3 = '0014'
                        cycle2 = '20'
                    else:
                        data3 = str(hex(int(item[17]))).upper()[2:].zfill(4)
                        cycle2 = item[17]

                    if int(item[14]) == 2:
                        jiange = '100'
                        data2 = '0014'
                    else:
                        data2 = str(hex(int(cycle2) * 5)).upper()[2:].zfill(4)
                        jiange = int(cycle2)

                    data4 = str(hex(int(item[19]))).upper()[2:].zfill(4)
                    data5 = str(hex(int(item[20]))).upper()[2:].zfill(4)
                    message2 = channel + ID + data2 + data3 + data4 + data5
                    Mac2 = ' '.join(message2[i:i + 2] for i in range(0, len(message2), 2)).upper()
                    forthitem = ['event', name, Mac2, '20', '3', '0x701', '', '', '--', 'Motorola']
                    dataList.append(forthitem)
                    preview.append(forthitem)

                    fifthitem = ['check', '单条测试用例处理状态', '1', jiange, '100', '0x710', '16', '8', '--', 'Motorola']
                    dataList.append(fifthitem)
                    preview.append(fifthitem)

                    relength = '1' * (int(item[20]))
                    renum = int(relength, 2)
                    sixth = ['check', name, renum, '20', '100', item[15], item[19], item[20], '--', 'Motorola']
                    ansixth = ['check', name, '0', '20', '100', item[15], item[19], item[20], '--', 'Motorola']
                    dataList.append(sixth)
                    preview.append(ansixth)

                    dataList = dataList + preview
                    text = item[0] + ' ' + item[1] + ' ' + item[3] + '      ' + item[12] + ' ' + item[13] + ' ' + item[15]
                    note = ['测试用例说明', text, '12', '', '', '', '', '', '', '']
                    dataList.append(note)
                elif item[1] == '' and item[0] != '':
                    name = item[0]
                    data1 = "0" + str(self.banqiaoList.index(item[0])) + '01'
                    if item[9] == '1':
                        data2 = '0001'
                        check = '0'
                    else:
                        data2 = '0000'
                        check = '1'
                    data3 = 'FFFF0000'
                    message1 = data1 + data2 + data3
                    Mac = ' '.join(message1[i:i + 2] for i in range(0, len(message1), 2)).upper()
                    data = ['event', name, Mac, '100', '1', '0x705', '', '', '--', 'Motorola']
                    dataListhard.append(data)
                    seconditem = ['check', '单条测试用例处理状态', '5', '100', '100', '0x710', '16', '8', '--', 'Motorola']
                    dataListhard.append(seconditem)

                    if int(item[14]) == 2:
                        name = item[13] + "_" + item[12]
                        try:
                            channel = '0' + str(self.LINList.index(item[13])) + '01'
                        except:
                            channel = '0000'
                        ID = '4' + item[15].replace('0x', '').replace('0X', '').zfill(3)
                    elif int(item[14]) == 1:
                        if item[12] != '/':
                            name = item[13] + '_CAN_' + item[15] + "_" + item[12]
                        else:
                            name = item[13] + '_CAN_' + item[15] + "_"
                        try:
                            channel = '0' + str(self.CANList.index(item[13] + '_CAN')) + '00'
                        except:
                            channel = '0000'
                        ID = '8' + item[15].replace('0x', '').replace('0X', '').zfill(3)
                    else:
                        if item[12] != '/':
                            name = item[13] + '_CAN_' + item[15] + "_" + item[12]
                        else:
                            name = item[13] + '_CAN_' + item[15] + "_"
                        try:
                            channel = '0' + str(self.CANList.index(item[13] + '_CAN')) + '00'
                        except:
                            channel = '0000'
                        ID = item[15].replace('0x', '').replace('0X', '').zfill(4)

                    if int(item[17]) < 20:
                        data3 = '0014'
                        cycle2 = '20'
                    else:
                        data3 = str(hex(int(item[17]))).upper()[2:].zfill(4)
                        cycle2 = item[17]

                    if int(item[14]) == 2:
                        jiange = '100'
                        data2 = '0014'
                    else:
                        data2 = str(hex(int(cycle2) * 5)).upper()[2:].zfill(4)
                        jiange = int(cycle2)

                    data4 = str(hex(int(item[19]))).upper()[2:].zfill(4)
                    data5 = str(hex(int(item[20]))).upper()[2:].zfill(4)
                    message2 = channel + ID + data2 + data3 + data4 + data5
                    Mac2 = ' '.join(message2[i:i + 2] for i in range(0, len(message2), 2)).upper()
                    forthitem = ['event', name, Mac2, '20', '1', '0x701', '', '', '--', 'Motorola']
                    dataListhard.append(forthitem)

                    fifthitem = ['check', '单条测试用例处理状态', '1', jiange, '100', '0x710', '16', '8', '--', 'Motorola']
                    dataListhard.append(fifthitem)

                    ansixth = ['check', name, check, '20', '100', item[15], item[19], item[20], '--', 'Motorola']
                    dataListhard.append(ansixth)

                    data1 = "0" + str(self.banqiaoList.index(item[0])) + '01'
                    if item[9] == '0':
                        data2 = '0001'
                        check = '0'
                    else:
                        data2 = '0000'
                        check = '1'
                    data3 = 'FFFF0000'
                    message1 = data1 + data2 + data3
                    Mac = ' '.join(message1[i:i + 2] for i in range(0, len(message1), 2)).upper()
                    sevenitem = ['event', item[0], Mac, '100', '1', '0x705', '', '', '--', 'Motorola']
                    sixth = ['check', name, check, '20', '100', item[15], item[19], item[20], '--', 'Motorola']
                    dataListhard.append(sevenitem)
                    dataListhard.append(seconditem)
                    dataListhard.append(forthitem)
                    dataListhard.append(fifthitem)
                    dataListhard.append(sixth)
                    text = item[0] + ' ' + item[1] + ' ' + item[3] + '      ' + item[12] + ' ' + item[13] + ' ' + item[15]
                    note = ['测试用例说明', text, '10', '', '', '', '', '', '', '']
                    dataListhard.append(note)

        # 检查数据是否为空
        has_dataList = bool(dataList)
        has_dataListhard = bool(dataListhard)

        if not has_dataList and not has_dataListhard:
            QMessageBox.warning(self, "警告", "没有可保存的数据！")
            return True

        # 弹窗提醒用户需要保存的文件数量
        if has_dataList and has_dataListhard:
            QMessageBox.information(self, "提示", "需要保存两个文件：信号路由自动测试和硬件测试。")
        elif has_dataList:
            QMessageBox.information(self, "提示", "需要保存一个文件：信号路由自动测试。")
        else:
            QMessageBox.information(self, "提示", "需要保存一个文件：硬件测试。")

        # 导出信号路由自动测试
        if dataList:
            headnameList = ['序号', '操作类型', '操作名称', '操作值', '间隔（ms）', 'Cycle（ms）', 'canID', 'Start', 'Length',
                            'flag', 'format']
            file_path1, _ = QFileDialog.getSaveFileName(self, '导出信号路由自动测试', '', 'Excel File(*.xlsx)')
            if file_path1:
                workbook = Workbook(file_path1)
                sheet1 = workbook.add_worksheet('CAN')
                sheet1.set_column('C:D', 50)
                sheet1.set_column('B:B', 15)
                sheet1.set_column('E:K', 10)
                font = workbook.add_format({'font_name': '等线', 'font_size': 12, 'align': 'center'})
                for k in range(len(headnameList)):
                    sheet1.write(0, k, headnameList[k], font)
                row = 1
                for l in range(len(dataList)):
                    sheet1.write(row, 0, str(row), font)
                    for j in range(len(dataList[l])):
                        sheet1.write(row, j + 1, dataList[l][j], font)
                    row += 1
                workbook.close()
                dataList.clear()
        if dataListhard:
            headnameList = ['序号', '操作类型', '操作名称', '操作值', '间隔（ms）', 'Cycle（ms）', 'canID', 'Start',
                            'Length',
                            'flag', 'format']
            # workbook = Workbook(path4)
            # 获取用户选择的文件路径
            file_path2, _ = QFileDialog.getSaveFileName(self, '导出硬件测试', '', 'Excel File(*.xlsx)')
            workbook = Workbook(file_path2)
            sheet1 = workbook.add_worksheet('CAN')
            sheet1.set_column('C:D', 50)
            sheet1.set_column('B:B', 15)
            sheet1.set_column('E:K', 10)
            font = workbook.add_format({'font_name': '等线', 'font_size': 12, 'align': 'center'})
            for k in range(len(headnameList)):
                sheet1.write(0, k, headnameList[k], font)
            row = 1
            for l in range(len(dataListhard)):
                sheet1.write(row, 0, str(row), font)
                for j in range(len(dataListhard[l])):
                    sheet1.write(row, j + 1, dataListhard[l][j], font)
                row = row + 1
            workbook.close()
            dataListhard.clear()


if __name__ == '__main__':
    # path1= r"C:\Users\pc\Downloads\SinaRoutingTable.CSV"
    # path2 = r"C:\Users\pc\Downloads\自动化测试盒协议_BDM_KQC2_R.xlsx"
    # # path3 = './信号路由自动测试.xlsx'
    # # path4='./硬件测试.xlsx'

    app = QApplication(sys.argv)
    # 初始化
    mainwindow = QMainWindow()
    ui_components = Ui_MainWindow()
    ui_components.setupUi(mainwindow)
    aa = SignalRoutingTest()
    aa.show()
    sys.exit(app.exec_())




















