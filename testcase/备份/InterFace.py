"""
#!/usr/bin/env python3.9
#-*- coding:utf-8 -*-
@Project:testcase
@File:InterFace.py
@Author:XU AO
@Time:2025/4/22 08:36
"""
import configparser
import os
import json
import logging
import sys
import traceback

import openpyxl
import xlrd2
import csv

from PyQt5.QtGui import QKeySequence, QCursor, QFont
from PyQt5.QtWidgets import QShortcut, QComboBox, QAction, QMenu

from PyQt5 import QtWidgets, QtCore
from PyQt5.QtGui import QKeySequence
from PyQt5.QtWidgets import QMessageBox, QApplication, QMainWindow, QFileDialog, QTableWidgetItem, QHeaderView, \
    QPushButton
from xlsxwriter import Workbook

from Function import *
# from CanLinConfig import CanLinConfig
# from ELFAnalysis import ELFAnalysis
import tkinter as tk
from tkinter import messagebox

from UI2 import Ui_MainWindow
from Function.CanLinConfig import CanLinConfig
from Function.ELFAnalysis import ELFAnalysis

class InterFace:
    config = CanLinConfig()
    elfAnalysis = ELFAnalysis()

    def __init__(self, table_widget):
        self.table = table_widget  # 主窗口传入的表格控件
        self.hardwaredic = {}
        self.CANdic = {}
        self.LINdic = {}
        self.interfacedic = {}
        self.InterfaceList = []
        self.elfAnalysis = ELFAnalysis()

    def settable(self):
        headnameList = ['接口类型', '接口名称（代码中的全局变量）', '信号描述（用例显示名称）', '信号归属模块（ARXML名称）',
                        'map地址', '信号方向', '   信号长度(bit)   ', '正向测试值   ', '关联信号',
                        '关联属性', '操作']
        self.table.horizontalHeader().setVisible(True)
        self.table.setColumnCount(len(headnameList))
        self.table.setRowCount(0)
        self.table.verticalHeader().setVisible(False)
        self.table.setHorizontalHeaderLabels(headnameList)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(8, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(9, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(10, QHeaderView.ResizeToContents)
    def add_row(self):
        try:
            current_row_count = self.table.rowCount()
            self.table.insertRow(current_row_count)
            deleteButton = QPushButton("删除")
            deleteButton.clicked.connect(lambda: self.delete_clicked(current_row_count))  # 传递当前行号
            self.table.setCellWidget(current_row_count, 10, deleteButton)
        except Exception as e:
            QMessageBox.critical(None, "错误", f"添加行时发生错误: {str(e)}")

    # 修改删除函数
    def delete_clicked(self, row):
        if row >= 0:  # 确保行号有效
            self.table.removeRow(row)  # 删除行
            self.table.verticalScrollBar().setSliderPosition(row)  # 滚动条调整位置
        else:
            QMessageBox.warning(None, "警告", "删除失败，无法找到该行。")

    def clear_table(self):
        try:
            response = QMessageBox.question(None, '确认', '确定清空数据吗？')
            if response == QMessageBox.Yes:
                self.table.setRowCount(0)
                # 清空接口列表
                self.InterfaceList.clear()
                # self.settable()  # 重新设置表格头
                # QMessageBox.information(self, "提示", "数据已清空！")
            else:
                pass
        except Exception as e:
            logging.error(f"清空表格时发生错误: {traceback.format_exc()}")



    def excel_toTable(self):
        try:
            path = QFileDialog.getOpenFileName(None,"选择文件", '/', "xlsx Files(*.xlsx)")
            sheet_name = '软件接口定义表'
            if not path[0]:
                QMessageBox.information(None,"温馨提示", "未选择文件！")
                return
            if path[0]:
                book = xlrd2.open_workbook(path[0])
                if sheet_name not in book.sheet_names():
                    QMessageBox.information(None, "温馨提示", "导入文件不正确！")
                    return
                sheet = book.sheet_by_name('软件接口定义表')
                row = self.table.rowCount()
                for i in range(1, sheet.nrows):
                    values = sheet.row_values(i)
                    self.add_row()
                    for j in range(len(values)):
                        try:
                            text = str(int(values[j]))
                        except:
                            text = str(values[j])
                        self.table.setItem(row + i - 1, j, QTableWidgetItem(text))
        except:
            logging.error(traceback.format_exc() + "\n")


    def export_to_excel(self):
        try:
            # 打开文件保存对话框，让用户选择保存路径
            filename, _ = QFileDialog.getSaveFileName(None, "导出测试用例", "", "Excel Files (*.xlsx)")
            if not filename:
                return  # 如果用户取消选择，直接返回

            # 创建 Excel 工作簿和工作表
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "软件接口定义表"

            # 写入表头
            headers = ["接口类型", "接口名称", "信号描述", "信号归属模块", "map地址", "信号方向", "信号长度","正向测试值", "关联信号", "关联属性"]
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)

            # 写入表格数据
            for row_num in range(self.table.rowCount()):
                for col_num in range(self.table.columnCount()):
                    item = self.table.item(row_num, col_num)
                    if item is not None:
                        sheet.cell(row=row_num + 2, column=col_num + 1, value=item.text())

            # 保存文件
            workbook.save(filename)
            QMessageBox.information(None, "成功", f"数据已成功导出到 {filename}")
        except Exception as e:
            logging.error(f"导出到Excel时发生错误: {traceback.format_exc()}")
            QMessageBox.critical(None, "错误", "导出失败，请查看日志！")

    def show_success_dialog(self, state, message):
        # 创建一个简单的Tkinter窗口
        success_dialog = tk.Tk()
        # 隐藏主窗口
        success_dialog.withdraw()
        # 显示成功消息
        messagebox.showinfo(state, message)

    def readMesseage(self, path1, path2):
        try:
            for i in range(0, self.table.rowCount()):
                datalist = []
                for j in range(0, self.table.columnCount() - 1):
                    data = self.table.item(i, j).text()
                    datalist.append(data)
                self.InterfaceList.append(datalist)
            self.elfAnalysis.loadPath(path2)
            if path1:
                book1 = xlrd2.open_workbook(path1)
                sheetbook1 = book1.sheet_by_name('软件接口定义表')
                for i in range(1, sheetbook1.nrows):
                    values = sheetbook1.row_values(i)
                    # self.InterfaceList.append(values)
                    self.interfacedic[values[1]] = values[2]
                sheetbook2 = book1.sheet_by_name('硬件配置表')
                for i in range(2, 12):
                    values = sheetbook2.col_values(i)
                    CANlist = [str(values[2])] + values[4:]
                    CANlist = [item for item in CANlist if item and item.strip()]
                    self.hardwaredic[values[3]] = CANlist
                sheetbook3 = book1.sheet_by_name('LIN用例解析矩阵')
                dataList = []
                for i in range(sheetbook3.nrows - 1, 0, -1):
                    values = sheetbook3.row_values(i)
                    if values[3] == '':
                        # 信号描述、起始位、信号长度
                        data = [values[8], values[11], values[13]]
                        dataList.append(data)
                    elif values[3] != [] and dataList != []:
                        self.LINdic[values[3]] = dataList
                        dataList = []
                dataList = []
                sheetbook4 = book1.sheet_by_name('CAN用例解析矩阵')
                for i in range(sheetbook4.nrows - 1, 0, -1):
                    values = sheetbook4.row_values(i)
                    if values[3] == '':
                        # 信号描述、起始位、信号长度、周期
                        data = [values[8], values[11], values[13]]
                        dataList.append(data)
                    elif values[3] != [] and dataList != []:
                        self.CANdic[values[3] + '/' + str(values[5])] = dataList
                        dataList = []
                if self.genautotable():
                    self.show_success_dialog('Success', '生成成功！')
                else:
                    self.show_success_dialog('FAIL', '生成失败，没有找到相关属性！')
        except:
            self.show_success_dialog('FAIL', '生成失败，查看log！')
            logging.error(traceback.format_exc() + "\n")

    def genautotable(self):
        dataList = []
        # 遍历 tableWidget 中的每一行，动态生成 dataList
        for row in range(self.table.rowCount()):
            # 提取每一列的值
            item = []
            for col in range(self.table.columnCount() - 1):  # 排除最后一列的操作按钮
                cell_item = self.table.item(row, col)
                if cell_item is not None:
                    item.append(cell_item.text())
                else:
                    item.append("")  # 如果单元格为空，使用空字符串

            # 以下逻辑保持不变，仅将数据来源从 InterfaceList 替换为 tableWidget
            channel = '00'
            type = ''
            CANID = ''
            for key, values in self.hardwaredic.items():
                for value in values:
                    if value in item[8]:  # 关联信号列
                        CANID = str(values[0])
                        channel = values.index(value)
                        type = key

            # 对首行的 CANID 处理
            if CANID == '700/701':
                CANID = '0x700'
            elif '.' in CANID:
                CANID = '0x' + CANID.split('.')[0]
            else:
                CANID = '0x000'

            var_name = item[1].strip().upper()  # 去除空格并转换为大写
            try:
                inaddress = self.elfAnalysis.getAddressWithName(var_name).split('x')[1]
            except KeyError:
                logging.error(f"Variable not found in ELF file: {var_name}")
                inaddress = "00000000"  # 使用默认值或跳过该条目

            if type == 'LIN':
                thirddata1 = [key for key, values in self.LINdic.items() for val in values if item[8] in val]
                thirdstartbit = [val[1] for key, values in self.LINdic.items() for val in values if item[8] in val]
                thirdlength = [val[2] for key, values in self.LINdic.items() for val in values if item[8] in val][0]
                startbit = str(hex(int(thirdstartbit[0]))).upper()[2:].zfill(4)
                shuxing = '01'
                cycle = '000A'
            elif type == 'CAN':
                thirddata1 = [key for key, values in self.CANdic.items() for val in values if item[8] in val]
                thirdstartbit = [val[1] for key, values in self.CANdic.items() for val in values if item[8] in val]
                thirdlength = [val[2] for key, values in self.CANdic.items() for val in values if item[8] in val][0]
                startbit = str(hex(int(thirdstartbit[0]))).upper()[2:].zfill(4)
                shuxing = '00'
                cycle = str(hex(int(thirddata1[0].split('/')[1]))).upper()[2:].zfill(4)
            else:
                thirddata1 = '00'
                thirdstartbit = '0'
                startbit = '0000'
                shuxing = '00'
                cycle = '0000'
                thirdlength = '0'

            # 根据信号方向（IN 或 OUT）生成测试用例
            if item[5] == 'IN':  # 信号方向为 IN
                preview = []
                if item[9] != 'HWA':
                    length = str(hex(int(item[6]))).upper()[2:].zfill(4)
                    data1 = thirddata1[0].split('x')[1].split('/')[0].zfill(3)
                    text = str(channel).zfill(2) + shuxing + '4' + data1 + '0064' + cycle + startbit + length
                    text = ' '.join(text[i:i + 2] for i in range(0, len(text), 2)).upper()
                    firsttext = ['event', item[8], text, '21', '1', CANID, '', '', '--', 'Motorola']
                    secondtext = ['check', '单条测试用例处理状态', CANID[-1], '22', '100', '0x710', '16', '8', '--','Motorola']
                else:
                    if item[7] == '1':
                        data2 = '0001'
                    else:
                        data2 = '0000'
                    shuxing = '01'
                    message1 = str(channel).zfill(2) + shuxing + data2 + 'FFFF0000'
                    message1 = ' '.join(message1[i:i + 2] for i in range(0, len(message1), 2)).upper()
                    firsttext = ['event', item[8], message1, '20', '1', '0x705', '', '', '--', 'Motorola']
                    secondtext = ['check', '单条测试用例处理状态', CANID[-1], '21', '100', '0x710', '16', '8', '--','Motorola']

                dataList.append(firsttext)
                preview.append(firsttext)
                dataList.append(secondtext)
                preview.append(secondtext)

                if item[9] != 'HWA':
                    string = self.config.getMessage(thirdstartbit[0], item[6], item[7], 8, '')
                    string11 = self.config.getMessage(thirdstartbit[0], item[6], '0', 8, '')
                    thirdtext = ['event', item[8], string, '100', '1', thirddata1[0], '', '', '--', 'Motorola']
                    dataList.append(thirdtext)
                    threetext = ['event', item[8], string11, '100', '1', thirddata1[0], '', '', '--', 'Motorola']
                    preview.append(threetext)

                num = '01' if item[0] in ('UInt8', 'Boolean') else '81' if item[0] == 'SInt8' else '02' if item[0] == 'UInt16' else '82' if \
                item[0] == 'SInt16' else '00'
                message = '2EF6E9' + inaddress + num + '00000000'
                message = ' '.join(message[i:i + 2] for i in range(0, len(message), 2)).upper()

                if item[9] != 'HWA':
                    forthtext = ['event', 'Rte_Read_' + item[2], message, '24', '1', '0x7F0', '', '', '--', 'Motorola']
                    fifthtext = ['check', '单条测试用例处理状态', '240', '25', '100', '0x710', '16', '8', '--','Motorola']
                    sixthtext = ['event', 'Rte_Read_' + item[2], '22 F6 E9', '26', '1', '0x7F0', '', '', '--','Motorola']
                    seventhtext = ['check', '单条测试用例处理状态', '240', '27', '100', '0x710', '16', '8', '--','Motorola']
                else:
                    forthtext = ['event', 'Rte_Read_' + item[2], message, '22', '1', '0x7F0', '', '', '--', 'Motorola']
                    fifthtext = ['check', '单条测试用例处理状态', '240', '23', '100', '0x710', '16', '8', '--', 'Motorola']
                    sixthtext = ['event', 'Rte_Read_' + item[2], '22 F6 E9', '24', '1', '0x7F0', '', '', '--','Motorola']
                    seventhtext = ['check', '单条测试用例处理状态', '240', '25', '100', '0x710', '16', '8', '--','Motorola']

                dataList.append(forthtext)
                preview.append(forthtext)
                dataList.append(fifthtext)
                preview.append(fifthtext)
                dataList.append(sixthtext)
                preview.append(sixthtext)
                dataList.append(seventhtext)
                preview.append(seventhtext)

                zhengxiang = str(hex(int(item[7]))).upper()[2:].zfill(8)
                fanxiang = '00000000' if zhengxiang != '00000000' else str(hex(int(item[7]))).upper()[2:].zfill(8)
                if fanxiang == '00000000':
                    fanxiang = '00000001'

                message8 = '62F6E9' + inaddress + num + zhengxiang
                message8 = ' '.join(message8[i:i + 2] for i in range(0, len(message8), 2)).upper()
                message16 = '62F6E9' + inaddress + num + fanxiang
                message16 = ' '.join(message16[i:i + 2] for i in range(0, len(message16), 2)).upper()

                if item[9] != 'HWA':
                    eighthtext = ['routercheck', 'Rte_Read_' + item[2], message8, '28', '100', '0x7F1', '', '', '--','Motorola']
                    eighttext = ['routercheck', 'Rte_Read_' + item[2], message16, '28', '100', '0x7F1', '', '', '--','Motorola']
                else:
                    eighthtext = ['routercheck', 'Rte_Read_' + item[2], message8, '26', '100', '0x7F1', '', '', '--','Motorola']
                    eighttext = ['routercheck', 'Rte_Read_' + item[2], message16, '26', '100', '0x7F1', '', '', '--','Motorola']

                dataList.append(eighthtext)
                preview.append(eighttext)
                dataList = dataList + preview
                note = ['测试用例说明', '由' + item[8] + '   测：Rte_Read_' + item[2]]
                dataList.append(note)

            elif item[5] == 'OUT':  # 信号方向为 OUT
                preview = []
                zhengxiang = str(hex(int(item[7]))).upper()[2:].zfill(8)
                fanxiang = '00000000' if zhengxiang != '00000000' else str(hex(int(item[7]))).upper()[2:].zfill(8)
                caozuozhi = str(hex(int(item[7]))).upper()[2:] if zhengxiang != '00000000' else '0'
                fanxiangnum = '0' if zhengxiang != '00000000' else str(hex(int(item[7]))).upper()[2:]

                num = '11' if item[0] in ('UInt8', 'Boolean') else '91' if item[0] == 'SInt8' else '12' if item[0] == 'UInt16' else '92' if \
                item[0] == 'SInt16' else '00'
                text = '2EF6E9' + inaddress + num + zhengxiang
                text = ' '.join(text[i:i + 2] for i in range(0, len(text), 2)).upper()
                text1 = '2EF6E9' + inaddress + num + fanxiang
                text1 = ' '.join(text1[i:i + 2] for i in range(0, len(text1), 2)).upper()

                firsttext = ['event', 'Rte_Write_' + item[2], text, '20', '1', '0x7F0', '', '', '--', 'Motorola']
                dataList.append(firsttext)
                firsttext1 = ['event', 'Rte_Write_' + item[2], text1, '20', '1', '0x7F0', '', '', '--', 'Motorola']
                preview.append(firsttext1)

                secondtext = ['check', '单条测试用例处理状态', '240', '21', '100', '0x710', '16', '8', '--', 'Motorola']
                dataList.append(secondtext)
                preview.append(secondtext)

                if '开关输入保留' in item[8]:
                    CANID = '0x706'

                if item[9] != 'HWA' and item[9] != 'APP':
                    data = thirddata1[0].split('/')[0][2:].zfill(4)
                    message3 = str(channel).zfill(2) + shuxing + data + '0064' + cycle + self.config.getStartandLengthHex(thirdstartbit[0],thirdlength).replace(' ', '')
                    message3 = ' '.join(message3[i:i + 2] for i in range(0, len(message3), 2)).upper()
                    thirdtext = ['event', item[8], message3, '22', '1', '0x701', '', '', '--', 'Motorola']
                    forthtext = ['check', '单条测试用例处理状态', '1', '100', '100', '0x710', '16', '8', '--','Motorola']
                    dataList.append(thirdtext)
                    preview.append(thirdtext)
                    dataList.append(forthtext)
                    preview.append(forthtext)

                    fifthtext = ['check', item[8], caozuozhi, '23', '100', CANID, '', '', '--', 'Motorola']
                    dataList.append(fifthtext)
                    fivetext = ['check', item[8], fanxiangnum, '23', '100', CANID, '', '', '--', 'Motorola']
                    preview.append(fivetext)

                    note = ['测试用例说明', '由Rte_Write_' + item[2] + '   测：' + item[8]]

                elif item[9] == 'HWA':
                    fifthtext = ['check', item[8], caozuozhi, '22', '100', CANID, '', '', '--', 'Motorola']
                    dataList.append(fifthtext)
                    fivetext = ['check', item[8], fanxiangnum, '22', '100', CANID, '', '', '--', 'Motorola']
                    preview.append(fivetext)

                    note = ['测试用例说明', '由Rte_Write_' + item[2] + '   测：' + item[8]]

                else:
                    num = '01' if item[0] in ('UInt8', 'Boolean') else '81' if item[0] == 'SInt8' else '02' if item[0] == 'UInt16' else '82' if \
                    item[0] == 'SInt16' else '00'
                    try:
                        name = [key for key, values in self.interfacedic.items() if item[8] == values][0]
                    except:
                        self.show_success_dialog('FAIL', 'APP属性的关联信号无法找到Rte接口！')
                        sys.exit(1)

                    outaddress = self.elfAnalysis.getAddressWithName(name).split('x')[1]
                    data = '2EF6E9' + outaddress + num + fanxiang
                    data = ' '.join(data[i:i + 2] for i in range(0, len(data), 2)).upper()

                    thirdtext = ['event', 'Rte_Read_' + item[8], data, '22', '1', '0x7F0', '', '', '--', 'Motorola']
                    dataList.append(thirdtext)
                    preview.append(thirdtext)

                    forthtext = ['check', '单条测试用例处理状态', '240', '23', '100', '0x710', '16', '8', '--','Motorola']
                    dataList.append(forthtext)
                    preview.append(forthtext)

                    fifthtext = ['event', 'Rte_Read_' + item[8], '22 F6 E9', '24', '1', '0x7F0', '', '', '--', 'Motorola']
                    dataList.append(fifthtext)
                    preview.append(fifthtext)

                    sixthtext = ['check', '单条测试用例处理状态', '240', '25', '100', '0x710', '16', '8', '--','Motorola']
                    dataList.append(sixthtext)
                    preview.append(sixthtext)

                    message7 = '62F6E9' + outaddress + num + zhengxiang
                    message7 = ' '.join(message7[i:i + 2] for i in range(0, len(message7), 2)).upper()
                    message14 = '62F6E9' + outaddress + num + fanxiang
                    message14 = ' '.join(message14[i:i + 2] for i in range(0, len(message14), 2)).upper()

                    seventhtext = ['routercheck', 'Rte_Read_' + item[8], message7, '26', '100', '0x7F1', '', '', '--','Motorola']
                    seventext = ['routercheck', 'Rte_Read_' + item[8], message14, '26', '100', '0x7F1', '', '', '--','Motorola']

                    dataList.append(seventhtext)
                    preview.append(seventext)

                    note = ['测试用例说明', '由Rte_Write_' + item[2] + '   测：Rte_Read_' + item[8]]

                dataList = dataList + preview
                dataList.append(note)

        has_dataList = bool(dataList)
        if not has_dataList:
            QMessageBox.warning(None, "警告", "没有可保存的数据！")
            return False
        else:
            # 写入 Excel 文件
            headnameList = ['序号', '操作类型', '操作名称', '操作值', '间隔（ms）', 'Cycle（ms）', 'canID', 'Start',
                            'Length', 'flag', 'format']
            file_path1, _ = QFileDialog.getSaveFileName(None, '导出id路由自动测试', '', 'Excel File(*.xlsx)')
            workbook = Workbook(file_path1)
            # workbook = Workbook(path3)
            sheet1 = workbook.add_worksheet('CAN')
            sheet1.set_column('C:D', 50)
            sheet1.set_column('B:B', 15)
            sheet1.set_column('E:K', 10)
            font = workbook.add_format({'font_name': '等线', 'font_size': 12, 'align': 'center'})

            for k in range(len(headnameList)):
                sheet1.write(0, k, headnameList[k], font)

            row = 1
            for l in range(len(dataList)):
                sheet1.write(row, 0, str(row), font)
                for j in range(len(dataList[l])):
                    sheet1.write(row, j + 1, dataList[l][j], font)
                row += 1
            workbook.close()
            return True

