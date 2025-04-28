"""
#!/usr/bin/env python3.9
#-*- coding:utf-8 -*-
@Project:testcase
@File:RoutingTest.py
@Author:XU AO
@Time:2025/4/22 17:20
"""
import logging
import traceback
import openpyxl
import xlrd2

from PyQt5.QtWidgets import QMessageBox, QFileDialog, QTableWidgetItem, QHeaderView, QPushButton, QInputDialog
from xlsxwriter import Workbook
import tkinter as tk
from tkinter import messagebox

class RoutingTest:
    logging.basicConfig(
        filename="log.log",
        filemode="w",
        datefmt="%Y-%m-%d %H:%M:%S %p",
        format="%(asctime)s - %(name)s - %(levelname)s - %(module)s: %(message)s",
        level=40
    )
    routingList = []
    CANList = []
    logging.basicConfig(
        filename="log.log",
        filemode="w",
        datefmt="%Y-%m-%d %H:%M:%S %p",
        format="%(asctime)s - %(name)s - %(levelname)s - %(module)s: %(message)s",
        level=logging.DEBUG
    )

    def __init__(self, table_widget):
        self.table = table_widget


    def settable(self):
        headnameList = ['源网段', '源节点', '源ID', '源端是否CANFD', '源DLC', '源周期', '源字段', '目标网段', '目标节点',
                        '目标ID','目标端是否CANFD','目标DLC','目标周期','目标字段','报文路由类型','MsgRoutType','测试使能','可选项1','操作']
        self.table.setColumnCount(len(headnameList))
        self.table.setRowCount(0)
        self.table.verticalHeader().setVisible(False)
        self.table.setHorizontalHeaderLabels(headnameList)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(11, QHeaderView.ResizeToContents)


    def excel_toTable(self):
        try:
            path = QFileDialog.getOpenFileName(None, "选择文件", '/', "xlsx Files(*.xlsx)")
            if path[0]:
                book = xlrd2.open_workbook(path[0])
                sheet = (book.sheets()[0])
                row = self.table.rowCount()
                for i in range(1, sheet.nrows):
                    values = sheet.row_values(i)
                    self.add_row()
                    for j in range(len(values)):
                        try:
                            text = str(int(values[j]))
                        except:
                            text = str(values[j])
                        self.table.setItem(row + i - 1, j, QTableWidgetItem(text))
        except:
            logging.error(traceback.format_exc() + "\n")

    def add_row(self):
        try:
            current_row_count = self.table.rowCount()

            self.table.insertRow(current_row_count)
            deleteButton = QPushButton("删除")
            deleteButton.clicked.connect(self.delete_clicked)  # 传递当前行号
            self.table.setCellWidget(current_row_count, 18, deleteButton)
        except Exception as e:
            QMessageBox.critical(None, "错误", f"添加行时发生错误: {str(e)}")

    def add_rows(self):
        try:
            # 弹出对话框让用户输入需要添加的行数
            row_count, ok = QInputDialog.getInt(
                None,
                "添加行数",
                "请输入要添加的行数:",
                value=1,  # 默认值为1
                min=1     # 最小值为1
            )

            if not ok:  # 如果用户取消输入
                return

            current_row_count = self.table.rowCount()

            for _ in range(row_count):
                self.table.insertRow(current_row_count)
                delete_button = QPushButton("删除")
                delete_button.clicked.connect(self.delete_clicked)  # 绑定删除按钮事件
                self.table.setCellWidget(current_row_count, 18, delete_button)
                current_row_count += 1

        except Exception as e:
            QMessageBox.critical(None, "错误", f"添加行时发生错误: {str(e)}")

    def delete_clicked(self):
        current_row = self.table.currentRow()
        self.table.removeRow(current_row)  # 删除行
        self.table.verticalScrollBar().setSliderPosition(current_row)  # 滚动条调整位置

    def clear_table(self):
        """
        清空表格中的所有数据和行
        """
        try:
            response = QMessageBox.question(None, '确认', '确定清空数据吗？')
            if response == QMessageBox.Yes:
                self.table.clearContents()  # 清空表格内容
                self.table.setRowCount(0)  # 清空表格行数
                self.settable()  # 重新设置表格头
                # 清空外部缓存数据（如果有）
                self.routingList.clear()
                self.CANList.clear()

                # QMessageBox.information(self, "提示", "数据已清空！")
            else:
                pass
        except Exception as e:
            logging.error(traceback.format_exc() + "\n")

    def export_to_excel(self):
        try:
            # 打开文件保存对话框，让用户选择保存路径
            filename, _ = QFileDialog.getSaveFileName(None, "导出测试用例", "", "Excel Files (*.xlsx)")
            if not filename:
                return  # 如果用户取消选择，直接返回

            # 创建 Excel 工作簿和工作表
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "导出数据"

            # 写入表头
            headers = ['源网段', '源节点', '源ID', '源端是否CANFD', '源DLC', '源周期', '源字段', '目标网段', '目标节点',
                       '目标ID', '目标端是否CANFD', '目标DLC', '目标周期', '目标字段', '报文路由类型', 'MsgRoutType',
                       '测试使能', '可选项1', '操作']
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)

            # 写入表格数据
            for row_num in range(self.table.rowCount()):
                for col_num in range(self.table.columnCount()):
                    item = self.table.item(row_num, col_num)
                    if item is not None:
                        sheet.cell(row=row_num + 2, column=col_num + 1, value=item.text())

            # 保存文件
            workbook.save(filename)
            QMessageBox.information(None, "成功", f"数据已成功导出到 {filename}")
        except Exception as e:
            logging.error(f"导出到Excel时发生错误: {traceback.format_exc()}")
            QMessageBox.critical(None, "错误", "导出失败，请查看日志！")

    def show_success_dialog(self,state,message):
        # 创建一个简单的Tkinter窗口
        success_dialog = tk.Tk()
        # 隐藏主窗口
        success_dialog.withdraw()
        # 显示成功消息
        messagebox.showinfo(state, message)

    def readMesseage(self, path2):
        try:
            filename = QFileDialog.getSaveFileName(None, '导出测试用例', '', 'Excel File(*.xlsx)')
            if filename[0]:
                path3 = filename[0]
                for i in range(0, self.table.rowCount()):
                    datalist = []
                    for j in range(0, self.table.columnCount() - 1):
                        data = self.table.item(i, j).text()
                        datalist.append(data)
                    self.routingList.append(datalist)
                if path2:
                    book2 = xlrd2.open_workbook(path2)
                    sheetbook2 = book2.sheet_by_name('硬件配置表')
                    self.LINList = sheetbook2.col_values(3)[3:]
                    self.CANList = sheetbook2.col_values(2)[3:]
                    if self.genautotable(path3):
                        self.show_success_dialog('Success', '生成成功！')
                    else:
                        self.show_success_dialog('FAIL', '生成失败，没有找到相关属性！')
        except:
            logging.error(traceback.format_exc() + "\n")
            self.show_success_dialog('FAIL', '生成失败，查看log！')

    def is_remark_row(self, row):
        """判断某一行是否为备注行或空白行"""
        cell_item_0 = self.table.item(row, 0)
        if cell_item_0:
            # 判断是否为备注行或空白行
            return "备注" in cell_item_0.text() or cell_item_0.text().strip() == ""
        return True  # 如果单元格不存在，也视为需要跳过的行


    def genautotable(self, path3):
        dataList = []

        # 遍历 tableWidget 中的每一行，动态生成 dataList
        for row in range(self.table.rowCount()):
            if self.is_remark_row(row):
                continue  # 跳过备注行
            item = []

            # 遍历当前行的其他列
            for col in range(self.table.columnCount() - 1):  # 排除最后一列的操作按钮
                cell_item = self.table.item(row, col)
                if cell_item is not None:
                    item.append(cell_item.text())
                else:
                    item.append("")  # 如果单元格为空，使用空字符串作为默认值

            # 以下逻辑保持不变，仅将数据来源从 routingList 替换为 tableWidget
            preview = []
            try:
                data1 = '0' + str(self.LINList.index(item[0])) + '01'
            except ValueError:
                data1 = '0000'
            if int(item[3]) == 2:
                ID = '4' + item[2].replace('0x', '').replace('0X', '').zfill(3)
            else:
                ID = item[2].replace('0x', '').replace('0X', '').zfill(4)
            if int(item[5]) < 20:
                data3 = '0014'
                cycle = '20'
            else:
                data3 = str(hex(int(item[5]))).upper()[2:].zfill(4)
                cycle = item[5]
            data2 = ID + str(hex(int(cycle) * 5)).upper()[2:].zfill(4)
            data4 = str(hex((int(item[4]) - 1) * 8)).upper()[2:].zfill(4)
            data5 = str(hex(int(item[4]) * 8)).upper()[2:].zfill(4)
            message1 = data1 + data2 + data3 + data4 + data5
            Mac = ' '.join(message1[i:i + 2] for i in range(0, len(message1), 2)).upper()
            if item[1] != '/':
                name = item[0] + '_' + item[2] + "_" + item[1]
            else:
                name = item[0] + '_' + item[2] + "_"
            data = ['event', name, Mac, cycle, '3', '0x700', '', '', '--', 'Motorola']
            dataList.append(data)
            preview.append(data)

            seconditem = ['check', '单条测试用例处理状态', '0', cycle, '100', '0x710', '16', '8', '--', 'Motorola']
            dataList.append(seconditem)
            preview.append(seconditem)

            s = 'FF' * (int(item[4]) + 1)
            string = ' '.join(s[i:i + 2] for i in range(2, len(s), 2)).upper()
            thirditem = ['event', name, string, cycle, '3', item[2], '', '', '--', 'Motorola']
            anthirditem = ['event', name, string.replace('FF', '00'), cycle, '3', item[2], '', '', '--', 'Motorola']
            dataList.append(thirditem)
            preview.append(anthirditem)

            if item[8] != '/':
                name = item[7] + 'CAN_' + item[9] + "_" + item[8]
            else:
                name = item[7] + 'CAN_' + item[9] + "_"
            try:
                channel = '0' + str(self.CANList.index(item[7] + '_CAN')) + '00'
            except ValueError:
                channel = '0000'
            if int(item[10]) == 1:
                ID = '8' + item[9].replace('0x', '').replace('0X', '').zfill(3)
            else:
                ID = item[9].replace('0x', '').replace('0X', '').zfill(4)
            if int(item[12]) < 20:
                data3 = '0014'
                cycle2 = '20'
            else:
                data3 = str(hex(int(item[12]))).upper()[2:].zfill(4)
                cycle2 = item[12]

            data4 = str(hex((int(item[11]) - 1) * 8)).upper()[2:].zfill(4)
            data5 = str(hex(int(item[11]) * 8)).upper()[2:].zfill(4)
            message2 = channel + ID + str(hex(int(cycle2) * 5)).upper()[2:].zfill(4) + data3 + data4 + data5
            Mac2 = ' '.join(message2[i:i + 2] for i in range(0, len(message2), 2)).upper()
            forthitem = ['event', name, Mac2, cycle2, '3', '0x701', '', '', '--', 'Motorola']
            dataList.append(forthitem)
            preview.append(forthitem)

            fifthitem = ['check', '单条测试用例处理状态', '1', int(cycle2) * 5, '100', '0x710', '16', '8', '--',
                         'Motorola']
            dataList.append(fifthitem)
            preview.append(fifthitem)

            t = 'FF' * (int(item[11]) + 1)
            text = ' '.join(t[i:i + 2] for i in range(2, len(t), 2)).upper()
            sixth = ['routercheck', name, text, cycle2, '100', item[9], '', '', '--', 'Motorola']
            ansixth = ['routercheck', name, text.replace('FF', '00'), cycle2, '100', item[9], '', '', '--', 'Motorola']
            dataList.append(sixth)
            preview.append(ansixth)

            dataList = dataList + preview
            text = item[0] + ' ' + item[1] + ' ' + item[2] + '      ' + item[7] + ' ' + item[8] + ' ' + item[9]
            note = ['测试用例说明', text, '12', '', '', '', '', '', '', '']
            dataList.append(note)

        has_dataList = bool(dataList)
        if not has_dataList:
            QMessageBox.warning(None, "警告", "没有可保存的数据！")
            return False
        else:
            headnameList = ['序号', '操作类型', '操作名称', '操作值', '间隔（ms）', 'Cycle（ms）', 'canID', 'Start',
                            'Length',
                            'flag', 'format']
            workbook = Workbook(path3)
            sheet1 = workbook.add_worksheet('CAN')
            sheet1.set_column('C:D', 50)
            sheet1.set_column('B:B', 15)
            sheet1.set_column('E:K', 10)
            font = workbook.add_format({'font_name': '等线', 'font_size': 12, 'align': 'center'})
            for k in range(len(headnameList)):
                sheet1.write(0, k, headnameList[k], font)
            row = 1
            for l in range(len(dataList)):
                sheet1.write(row, 0, str(row), font)
                for j in range(len(dataList[l])):
                    sheet1.write(row, j + 1, dataList[l][j], font)
                row = row + 1
            workbook.close()
            return True

