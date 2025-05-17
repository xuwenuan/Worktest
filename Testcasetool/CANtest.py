"""
#!/usr/bin/env python3.9
#-*- coding:utf-8 -*-
@Project:TSMasterAPI-main
@File:CANtest.py
@Author:XU AO
@Time:2025/5/6 08:59
"""
import cantools
import openpyxl
from PyQt5.QtWidgets import QMainWindow, QApplication, QFileDialog, QMessageBox
from TSMasterAPI import *
from ctypes import *

from CANUI1  import Ui_MainWindow
import can
import sys

msg = TLIBCAN()
msg.FIdxChn = 0
msg.FIdentifier = 0x100
msg.FProperties = 5  # 表示为扩展帧
msg.FDLC = 8
FData = [0x10, 0x11, 0x12, 0x13, 0x14, 0x15, 0x16, 0x17]
for i in range(len(FData)):
    msg.FData[i] = FData[i]
msg1 = TLIBCAN(FIdentifier = 0x111,FData = [10,11,12,13,14,15,16,17])

FDmsg = TLIBCANFD()
FDmsg.FIdxChn = 0
FDmsg.FIdentifier = 0x101
FDmsg.FProperties = 5
FDmsg.FFDProperties = 0x1
FDmsg.FDLC = 9
FData0 = [0x10, 0x11, 0x12, 0x13, 0x14, 0x15, 0x16, 0x17, 0x18, 0x19, 0x1A, 0x20]
for i in range(len(FData0)):
    FDmsg.FData[i] = FData0[i]

def On_CAN_EVENT(OBJ, ACAN):
    if (ACAN.contents.FIdentifier == 0x111 and ACAN.contents.FIdxChn == 0):
        ACAN.contents.FData[0] += 1

OnCANevent = TCANQueueEvent_Win32(On_CAN_EVENT)
obj = c_int32(0)
id1 = c_int32(0)  # 加载dbc句柄

AppName = b'TSMaster'

def get_enumerate_hw_devices():
    ACount = c_int32(0)
    r = tsapp_enumerate_hw_devices(ACount)
    return r, ACount


class OutputRedirector:
    def __init__(self, text_edit):
        self.text_edit = text_edit

    def write(self, message):
        self.text_edit.append(message)

    def flush(self):
        pass


class Ui_Main(QMainWindow, Ui_MainWindow):

    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.pushButton_5.clicked.connect(self.refresh_connected_devices)
        self.comboBox.currentIndexChanged.connect(self.on_combobox_changed)

        # self.pushButton.clicked.connect(self.SendCANFD_CAN_Message)
        self.pushButton.clicked.connect(self.test_sendmessage)
        # 重定向标准输出
        sys.stdout = OutputRedirector(self.textEdit)

    def on_combobox_changed(self):
        """返回当前选中的文本"""
        return self.comboBox.currentText()

    def refresh_connected_devices(self):
        # 设置CAN通道数
        if (tsapp_set_can_channel_count(1) == 0):
            print("CAN通道设置成功")
        else:
            print("CAN通道设置失败", tsapp_set_can_channel_count(1))

        # 设置LIN通道数
        if (tsapp_set_lin_channel_count(0) == 0):
            print("LIN通道设置成功")
        else:
            print("LIN通道设置失败")

        # 定义一个字典用于将字符串映射到对应的枚举值
        device_sub_type_map = {
            "TC1011": TLIB_TS_Device_Sub_Type.TC1011,
            "TC1013": TLIB_TS_Device_Sub_Type.TC1013,
            "TC1016": TLIB_TS_Device_Sub_Type.TC1016,
            # 可以继续添加其他型号...
        }

        # 获取用户输入（改为从 QComboBox 获取）
        device_sub_type_input = self.on_combobox_changed()  # <-- 修改点：使用 combobox 的当前文本

        # 查找对应的枚举值
        device_sub_type = device_sub_type_map.get(device_sub_type_input.strip())

        if device_sub_type is None:
            print(f"错误：不支持的设备子类型 {device_sub_type_input}")
        else:
            # 硬件通道映射至软件通道
            # tosun其他硬件只需修改第6个参数，找到对应型号即可
            if 0 == tsapp_set_mapping_verbose(AppName,
                                             TLIBApplicationChannelType.APP_CAN,
                                             CHANNEL_INDEX.CHN1,
                                             device_sub_type_input.encode("UTF8"),  # 使用选择的值作为设备名称
                                             TLIBBusToolDeviceType.TS_USB_DEVICE,
                                             device_sub_type, 0, 0, True):
                print("1通道映射成功")
            else:
                print("1通道映射失败")

        # 设置canfd波特率
        if 0 == tsapp_configure_baudrate_canfd(CHANNEL_INDEX.CHN1, 500.0, 2000.0,
                                               TLIBCANFDControllerType.lfdtISOCAN,
                                               TLIBCANFDControllerMode.lfdmNormal, True):
            print("1通道canfd波特率成功")
        else:
            print("1通道canfd波特率失败")

        if 0 == tsapp_register_pretx_event_can(obj, OnCANevent):
            print("回调事件注册成功")
        else:
            print("回调事件注册失败")

        if 0 == tsapp_connect():  # 连接 CAN 工具
            print("can工具连接成功")
            r = tsfifo_enable_receive_fifo()
            print("tsfifo_enable_receive_fifo() = ", r)
        else:
            print("can工具连接失败")


    # 发送can canfd报文
    def test_sendmessage(self):
        # 初始化标志变量，标记 CAN FD 报文发送是否成功
        canfd_success = True
        try:
            # 导入所需模块
            from PyQt5.QtWidgets import QFileDialog, QMessageBox
            import openpyxl

            # 选择文件对话框，限制为 .xlsx 文件
            file_info = QFileDialog.getOpenFileName(None, "选择文件", '/', "xlsx Files(*.xlsx)")
            if not file_info[0]:  # 未选择文件时给出提示
                QMessageBox.information(None, "温馨提示", "未选择文件！")
                return

            # 检查文件扩展名是否合法
            if not file_info[0].endswith('.xlsx'):
                QMessageBox.warning(None, "错误", "请选择有效的 .xlsx 文件")
                return

            # 加载 Excel 工作簿并获取活动工作表
            workbook = openpyxl.load_workbook(file_info[0])
            sheet = workbook.active

            # 用于存储成功创建并发送的 TLIBCANFD 对象
            t_msg_fd = []

            # 遍历 Excel 中的每一行，从第二行开始（第一行为标题）
            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                # 提取需要的列内容：操作类型、名称、操作值、CAN ID
                operation_type, operation_name, operation_value, can_id = row[1], row[2], row[3], row[6]

                # 处理 'event' 和 'check' 类型，且操作值和 CAN ID 均存在
                if operation_type in ("event", "check") and operation_value and can_id:
                    try:
                        # 将操作值解析为字节列表，假设输入格式为十六进制字符串
                        data_bytes = [int(b, 16) for b in operation_value.split()]
                        # 创建 TLIBCANFD 对象
                        msg_fd = TLIBCANFD(
                            FIdentifier=int(can_id, 16),  # 将 CAN ID 转换为十六进制整数
                            FDLC=len(data_bytes),          # 数据长度
                            FData=data_bytes[:]           # 仅填充实际长度的数据
                        )
                        # 如果是 'event' 类型，则发送 CAN FD 报文
                        if operation_type == "event":
                            canfd_result = tsapp_transmit_canfd_async(msg_fd)
                        # 检查发送结果
                        if canfd_result != 0:
                            print(f"第 {idx} 次 CAN FD 报文发送失败，错误码：{canfd_result}")
                            canfd_success = False  # 发送失败，标志置为 False
                            break  # 中断循环
                        # 发送成功或检查模式，将对象加入列表
                        t_msg_fd.append(msg_fd)
                        print(f"成功创建并发送 CAN FD 报文: {msg_fd}")
                    except Exception as e:
                        # 捕获并输出异常信息
                        print(f"解析 {operation_name} 失败: {e}")
                        canfd_success = False
                        break

            # 如果存在 'check' 类型报文，进行接收校验
            for msg_fd in t_msg_fd:
                if msg_fd.FProperties & 1 == 0:  # 仅处理 RX 报文
                    try:
                        # 接收报文
                        received_msg = tsfifo_receive_canfd_msgs(msg_fd.FIdentifier)
                        # 校验接收的数据
                        expected_data = list(msg_fd.FData[:msg_fd.FDLC])
                        received_data = list(received_msg.FData[:received_msg.FDLC])
                        if expected_data == received_data:
                            print(f"Check 成功: ID 0x{msg_fd.FIdentifier:03X}, 数据匹配")
                        else:
                            print(f"Check 失败: ID 0x{msg_fd.FIdentifier:03X}, 期望 {expected_data}, 实际 {received_data}")
                    except Exception as e:
                        print(f"Check 失败: ID 0x{msg_fd.FIdentifier:03X}, 错误: {e}")

            # 如果所有报文发送成功且列表中存在有效报文，则进行周期发送
            if canfd_success and t_msg_fd:
                for msg_fd in t_msg_fd:
                    # 设置周期发送，间隔 100ms
                    ret2 = tsapp_add_cyclic_msg_canfd(msg_fd, 100)
                    if ret2 == 0:
                        print(f"CAN FD 周期发送成功: ID 0x{msg_fd.FIdentifier:03X}")  # 十六进制格式输出 ID
                    else:
                        print(f"CAN FD 周期发送失败: ID 0x{msg_fd.FIdentifier:03X}")
            elif not t_msg_fd:
                # 如果没有有效的 CAN FD 报文
                print("No valid CANFD messages found in the file.")

        except Exception as e:
            # 捕获并输出文件读取和处理过程中的异常
            QMessageBox.critical(None, "错误", f"发生异常：{str(e)}")



if __name__ == '__main__':
    initialize_lib_tsmaster("TSMaster".encode("utf8"))
    ret, ACount = get_enumerate_hw_devices()
    print("ret = ", ret)
    print("在线硬件数量有%#d个" % (ACount.value - 1))
    PTLIBHWInfo = TLIBHWInfo()
    for i in range(ACount.value):
        tsapp_get_hw_info_by_index(i, PTLIBHWInfo)
        print(PTLIBHWInfo.FDeviceType, PTLIBHWInfo.FDeviceIndex, PTLIBHWInfo.FVendorName.decode("utf8"),
              PTLIBHWInfo.FDeviceName.decode("utf8"),
              PTLIBHWInfo.FSerialString.decode("utf8"))
    app = QApplication([])
    window = Ui_Main()
    window.show()
    app.exec_()



#测试封存
    # def test_sendmessage(self):
    #     # 初始化标志变量，标记 CAN FD 报文发送是否成功
    #     canfd_success = True
    #     try:
    #         # 导入所需模块
    #         from PyQt5.QtWidgets import QFileDialog, QMessageBox
    #         import openpyxl
    #
    #         # 选择文件对话框，限制为 .xlsx 文件
    #         file_info = QFileDialog.getOpenFileName(None, "选择文件", '/', "xlsx Files(*.xlsx)")
    #         if not file_info[0]:  # 未选择文件时给出提示
    #             QMessageBox.information(None, "温馨提示", "未选择文件！")
    #             return
    #
    #         # 检查文件扩展名是否合法
    #         if not file_info[0].endswith('.xlsx'):
    #             QMessageBox.warning(None, "错误", "请选择有效的 .xlsx 文件")
    #             return
    #
    #         # 加载 Excel 工作簿并获取活动工作表
    #         workbook = openpyxl.load_workbook(file_info[0])
    #         sheet = workbook.active
    #
    #         # 用于存储成功创建并发送的 TLIBCANFD 对象
    #         t_msg_fd = []
    #
    #         # 遍历 Excel 中的每一行，从第二行开始（第一行为标题）
    #         for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
    #             # 提取需要的列内容：操作类型、名称、操作值、CAN ID
    #             operation_type, operation_name, operation_value, can_id = row[1], row[2], row[3], row[6]
    #
    #             # 仅处理 'event' 类型，且操作值和 CAN ID 均存在
    #             if operation_type == "event" and operation_value and can_id:
    #                 try:
    #                     # 将操作值解析为字节列表，假设输入格式为十六进制字符串
    #                     data_bytes = [int(b, 16) for b in operation_value.split()]
    #                     # 创建 TLIBCANFD 对象
    #                     msg_fd = TLIBCANFD(
    #                         FIdentifier=int(can_id, 16),  # 将 CAN ID 转换为十六进制整数
    #                         FDLC=len(data_bytes),  # 数据长度
    #                         FData=data_bytes[:]  # 仅填充实际长度的数据
    #                     )
    #                     # 发送 CAN FD 报文
    #                     canfd_result = tsapp_transmit_canfd_async(msg_fd)
    #                     # 检查发送结果
    #                     if canfd_result != 0:
    #                         print(f"第 {idx} 次 CAN FD 报文发送失败，错误码：{canfd_result}")
    #                         canfd_success = False  # 发送失败，标志置为 False
    #                         break  # 中断循环
    #                     # 发送成功，将对象加入列表
    #                     t_msg_fd.append(msg_fd)
    #                     print(f"成功创建并发送 CAN FD 报文: {msg_fd}")
    #                 except Exception as e:
    #                     # 捕获并输出异常信息
    #                     print(f"解析 {operation_name} 失败: {e}")
    #                     canfd_success = False
    #                     break
    #
    #         # 如果所有报文发送成功且列表中存在有效报文，则进行周期发送
    #         if canfd_success and t_msg_fd:
    #             for msg_fd in t_msg_fd:
    #                 # 设置周期发送，间隔 100ms
    #                 ret2 = tsapp_add_cyclic_msg_canfd(msg_fd, 100)
    #                 if ret2 == 0:
    #                     print(f"CAN FD 周期发送成功: ID 0x{msg_fd.FIdentifier:03X}")  # 十六进制格式输出 ID
    #                 else:
    #                     print(f"CAN FD 周期发送失败: ID 0x{msg_fd.FIdentifier:03X}")
    #         elif not t_msg_fd:
    #             # 如果没有有效的 CAN FD 报文
    #             print("No valid CANFD messages found in the file.")
    #
    #     except Exception as e:
    #         # 捕获并输出文件读取和处理过程中的异常
    #         QMessageBox.critical(None, "错误", f"发生异常：{str(e)}")