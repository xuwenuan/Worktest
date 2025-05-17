"""
#!/usr/bin/env python3.9
#-*- coding:utf-8 -*-
@Project:testcase
@File:InterFace.py
@Author:XU AO
@Time:2025/4/22 08:36
"""
import logging
import os
import sys
import traceback

import openpyxl
import xlrd2
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QCursor
from PyQt5.QtWidgets import QMessageBox, QFileDialog, QTableWidgetItem, QHeaderView, QPushButton, QMenu, QAction, \
    QComboBox, QApplication, QInputDialog, QLineEdit, QCompleter
from xlsxwriter import Workbook
import tkinter as tk
from tkinter import messagebox
from Function.CanLinConfig import CanLinConfig
from Function.ELFAnalysis import ELFAnalysis

from PyQt5.QtWidgets import QComboBox
from PyQt5.QtCore import QEvent

class NoWheelComboBox(QComboBox):
    def __init__(self, parent=None):
        super().__init__(parent)

    def wheelEvent(self, event):
        # 转发滚轮事件给父控件（即 QTableWidget）
        if self.parent():
            parent_widget = self.parent()
            # 将滚轮事件转发给父控件
            QEvent.ignore(event)
            parent_widget.wheelEvent(event)


class DataImportThread(QThread):
    data_signal = pyqtSignal(list, list)  # 发送已解析数据和第8列数据列表
    error_signal = pyqtSignal(str)  # 发送错误信息

    def __init__(self, file_path, sheet_name):
        super().__init__()
        self.file_path = file_path
        self.sheet_name = sheet_name

    def run(self):
        try:
            book = xlrd2.open_workbook(self.file_path)
            if self.sheet_name not in book.sheet_names():
                raise ValueError(f"工作表 '{self.sheet_name}' 不存在")

            sheet = book.sheet_by_name(self.sheet_name)
            data = []
            combined_list = []

            for i in range(1, sheet.nrows):
                row_data = sheet.row_values(i)
                data.append(row_data)

                # 收集第8列数据
                if len(row_data) > 8:
                    value = str(row_data[8]).strip()
                    if value and value not in combined_list:
                        combined_list.append(value)

            # 发送数据
            self.data_signal.emit(data, combined_list)

        except Exception as e:
            error_message = f"数据导入过程中发生异常：{traceback.format_exc()}"
            logging.error(error_message)
            self.error_signal.emit(str(e))


class QComboboxEx(QComboBox):
    comboBoxID = ""
    canOrLinValue = ""
    index = 0

    def wheelEvent(self, QWheelEvent) -> None:
        pass

class InterFace:
    config = CanLinConfig()

    def __init__(self, table_widget,elfAnalysis):
        self.table = table_widget  # 主窗口传入的表格控件
        self.hardwaredic = {}
        self.CANdic = {}
        self.LINdic = {}
        self.interfacedic = {}
        self.InterfaceList = []
        self.elfAnalysis = elfAnalysis
        self.combined_list = []  # 初始化为空列表，避免访问未定义属性的问题


    def settable(self):
        headnameList = ['接口类型', '接口名称（代码中的全局变量）', '信号描述（用例显示名称）', '信号归属模块（ARXML名称）',
                        'map地址', '信号方向', '   信号长度(bit)   ', '正向测试值   ', '关联信号',
                        '关联属性', '操作']
        self.table.horizontalHeader().setVisible(True)
        self.table.setColumnCount(len(headnameList))
        self.table.setRowCount(0)
        self.table.verticalHeader().setVisible(False)
        self.table.setHorizontalHeaderLabels(headnameList)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(8, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(9, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(10, QHeaderView.ResizeToContents)

    def add_row(self):
        try:
            current_row_count = self.table.rowCount()

            self.table.insertRow(current_row_count)
            deleteButton = QPushButton("删除")
            deleteButton.clicked.connect(self.delete_clicked)  # 传递当前行号
            self.table.setCellWidget(current_row_count, 10, deleteButton)

        except Exception as e:
            QMessageBox.critical(None, "错误", f"添加行时发生错误: {str(e)}")

    def add_row_2(self, default_value=None):
        try:
            current_row_count = self.table.rowCount()

            # 插入新行
            self.table.insertRow(current_row_count)

            # 添加删除按钮
            deleteButton = QPushButton("删除")
            deleteButton.clicked.connect(self.delete_clicked)
            self.table.setCellWidget(current_row_count, 10, deleteButton)

            # 添加下拉框到第8列
            combo_box = QComboBox()
            self.add_comobox(current_row_count, 8, self.combined_list, combo_box, default_value=default_value)

        except Exception as e:
            logging.error(f"add_row_2 方法执行异常: {str(e)}")
            QMessageBox.critical(None, "错误", f"添加行时发生错误: {str(e)}")

    def add_row_3(self, insert_row):
        try:

            # 插入一行
            self.table.insertRow(insert_row)

            print(f"插入后行数: {self.table.rowCount()}")

            # 添加删除按钮到第10列
            delete_button = QPushButton("删除")
            delete_button.clicked.connect(self.delete_clicked)
            self.table.setCellWidget(insert_row, 10, delete_button)

            # 添加下拉框到第8列
            combo_box = QComboBox()
            self.add_comobox(insert_row, 8, self.combined_list, combo_box)

        except Exception as e:
            QMessageBox.critical(None, "错误", f"在插入行 {insert_row} 时发生错误: {str(e)}")

    def copy_datalist(self, datalist, linlist, Canlength, Linlength):
        try:
            # 复制 CAN 和 LIN 数据列表
            self.LINlist = linlist[:]
            self.CANlist = datalist[:]

            # 合并 CAN 和 LIN 列表并去重
            self.combined_list = sorted(set(datalist + linlist), key=str.lower)

            # 初始化新字典
            merged_dict = Canlength.copy()

            # 合并 Linlength 中的键值对（跳过异常）
            for key, value in Linlength.items():
                try:
                    if key not in merged_dict:
                        merged_dict[key] = value
                except Exception as e:
                    logging.warning(f"合并键 '{key}' 时发生异常: {str(e)}")
                    continue

            # 保存合并后的字典
            self.SignalLengthDict = merged_dict

        except Exception as e:
            logging.error(f"copy_datalist 方法执行异常: {str(e)}")
            QMessageBox.critical(None, "错误", f"数据列表复制过程中发生异常: {str(e)}")

    def add_rows(self):
        try:
            row_count, ok = QInputDialog.getInt(
                None,
                "添加行数",
                "请输入要添加的行数:",
                value=1,
                min=1,
                max=1000  # 添加 max 限制
            )

            if not ok:
                return

            current_row_count = self.table.rowCount()

            # 提高性能：禁用 UI 刷新直到插入完成
            self.table.setUpdatesEnabled(False)

            for _ in range(row_count):
                self.table.insertRow(current_row_count)

                # 添加删除按钮到第10列
                delete_button = QPushButton("删除")
                delete_button.clicked.connect(self.delete_clicked)
                self.table.setCellWidget(current_row_count, 10, delete_button)

                # 添加下拉框到第8列
                combo_box = QComboBox()
                self.add_comobox(current_row_count, 8, self.combined_list , combo_box)

                current_row_count += 1

        except Exception as e:
            QMessageBox.critical(None, "错误", f"添加行时发生错误: {str(e)}")
        finally:
            self.table.setUpdatesEnabled(True)  # 确保最终恢复刷新

    def add_comobox(self, i, j, comboBoxList, comboBox,default_value=None):
        try:
            # 使用自定义的 NoWheelComboBox
            comboBox = NoWheelComboBox()

            comboBox.setEditable(True)

            # 初始化 completer
            completer = QCompleter(comboBoxList)
            completer.setFilterMode(Qt.MatchContains)
            completer.setCompletionMode(QCompleter.PopupCompletion)
            comboBox.setCompleter(completer)

            # 清空之前的内容并添加新项
            comboBox.clear()
            comboBox.addItems(comboBoxList)

            # 插入到表格中的第 j 列
            self.table.setCellWidget(i, j, comboBox)
            # 设置默认值
            if default_value and default_value in comboBoxList:
                comboBox.setCurrentText(default_value)
            else:
                comboBox.setCurrentIndex(-1)  # 没有默认值时设为空

            # 添加信号连接
            comboBox.currentTextChanged.connect(lambda text, idx=i: self.combox_changed(idx, comboBox))

        except Exception as e:
            logging.error(f"add_comobox 发生错误: {str(e)}")


    def combox_changed(self, row, comboBox):
        try:
            # 检查 table 是否存在且 row 合法
            if not hasattr(self, 'table') or self.table is None:
                logging.warning("Table 未初始化")
                return

            if row < 0 or row >= self.table.rowCount():
                logging.warning(f"行号 {row} 超出范围")
                return

            # 检查 SignalLengthDict 字典是否存在
            if not hasattr(self, 'SignalLengthDict') or self.SignalLengthDict is None:
                logging.warning("SignalLengthDict 未初始化")
                return
            # 检查是否是滚轮事件触发
            modifiers = QApplication.keyboardModifiers()
            mouse_buttons = QApplication.mouseButtons()

            # 如果没有按下任何鼠标按钮且没有按下键盘按键，则认为是滚轮触发
            if not mouse_buttons and not modifiers:
                logging.info("滚轮事件触发，忽略")
                return

            selected_text = comboBox.currentText()
            # print(f"Selected Text: {selected_text}")

            # 初始值
            signal_length_int = "N/A"
            decimal_value = "N/A"

            # 获取信号长度值
            signal_length = self.SignalLengthDict.get(selected_text, "N/A")
            # print(f"Signal Length (Raw): {signal_length}")

            # 检查并转换为整数
            try:
                # 转换为整数
                signal_length_int = int(float(signal_length))
                # 计算 2^signal_length - 1
                decimal_value = (2 ** signal_length_int) - 1
            except (ValueError, TypeError):
                logging.warning(f"无法处理信号长度: {signal_length}")

            # print(f"Signal Length: {signal_length_int}, Decimal: {decimal_value}")

            # 设置第 6 列：位长度
            self.table.setItem(row, 6, QTableWidgetItem(str(signal_length_int)))

            # 设置第 7 列：十进制最大值
            self.table.setItem(row, 7, QTableWidgetItem(str(decimal_value)))

        except Exception as e:
            logging.error(f"combox_changed 发生错误: {str(e)}")

    def delete_clicked(self):
        try:
            # print(f"删除接口测试用例行: ")
            current_row = self.table.currentRow()
            self.table.removeRow(current_row)  # 删除行
            self.table.verticalScrollBar().setSliderPosition(current_row)  # 滚动条调整位置
        except Exception as e:
            # print(f"删除接口测试用例行过程中发生错误: {str(e)}")
            logging.error(f"删除接口测试用例行过程中发生错误: {str(e)}")
            msg_box = QMessageBox(QMessageBox.Critical, "删除错误", f"删除接口测试用例行过程中发生错误: {str(e)}")
            msg_box.exec_()

    def clear_table(self):
        try:
            response = QMessageBox.question(None, '确认', '确定清空数据吗？')
            if response == QMessageBox.Yes:
                self.table.setRowCount(0)
                # 清空接口列表
                self.InterfaceList.clear()
                # self.settable()  # 重新设置表格头
                # QMessageBox.information(self, "提示", "数据已清空！")
            else:
                pass
        except Exception as e:
            logging.error(f"清空表格时发生错误: {traceback.format_exc()}")

    def excel_toTable(self):
        try:
            file_info = QFileDialog.getOpenFileName(None, "选择文件", '/', "xlsx Files(*.xlsx)")
            sheet_name = '软件接口定义表'

            if not file_info[0]:
                QMessageBox.information(None, "温馨提示", "未选择文件！")
                return

            file_path = file_info[0]

            # 启动数据导入线程
            self.import_thread = DataImportThread(file_path, sheet_name)
            self.import_thread.data_signal.connect(self.on_data_imported)
            self.import_thread.error_signal.connect(self.on_import_error)
            self.import_thread.start()


        except Exception as e:
            logging.error(f"文件导入异常：{str(e)}")
            QMessageBox.critical(None, "导入失败", f"文件导入失败：{str(e)}")

    def on_data_imported(self, data, combined_list):
        """
        数据导入完成后的处理方法
        """
        start_row = self.table.rowCount()

        # 更新 combined_list
        self.combined_list = sorted(set(self.combined_list + combined_list), key=str.lower)

        # 批量插入数据
        for i, row_data in enumerate(data):
            default_value = row_data[8] if len(row_data) > 8 else ""
            self.add_row_2(default_value)

            # 填充数据到表格
            for j, value in enumerate(row_data):
                try:
                    text = str(int(value))
                except ValueError:
                    text = str(value)
                self.table.setItem(start_row + i, j, QTableWidgetItem(text))
        self.on_import_finished()

    def on_import_finished(self):
        """
        数据导入完成后的处理方法
        """
        # 关闭进度对话框
        if hasattr(self, 'progress_dialog') and self.progress_dialog.isVisible():
            self.progress_dialog.close()

        # 弹出导入成功提示框
        reply = QMessageBox.information(
            None,
            "导入成功",
            "数据导入完成，点击 '确定' 关闭窗口。"
        )

        # 确保关闭所有相关窗口
        if reply == QMessageBox.Ok:
            if hasattr(self, 'import_thread') and self.import_thread.isRunning():
                self.import_thread.terminate()

    def on_import_error(self, error_message):
        """
        文件导入过程中发生异常时的处理方法
        """
        QMessageBox.critical(None, "导入失败", f"数据导入失败：{error_message}")

    def export_to_excel(self):
        try:
            # 判断表格是否为空
            if self.table.rowCount() == 0 or self.table.columnCount() == 0:
                QMessageBox.warning(None, "警告", "表格中没有可导出的数据！")
                return

            # 打开文件保存对话框，让用户选择保存路径
            filename, _ = QFileDialog.getSaveFileName(None, "导出测试用例", "", "Excel Files (*.xlsx)")
            if not filename:
                return  # 用户取消选择

            # 检查表格是否为空
            is_empty = True
            for row_num in range(self.table.rowCount()):
                for col_num in range(self.table.columnCount()):
                    item = self.table.item(row_num, col_num)
                    if item and item.text().strip() != "":
                        is_empty = False
                        break
                if not is_empty:
                    break

            if is_empty:
                QMessageBox.warning(None, "警告", "表格中没有可导出的有效数据！")
                return

            # 创建 Excel 工作簿和工作表
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "软件接口定义表"

            # 写入表头
            headers = ["接口类型", "接口名称", "信号描述", "信号归属模块", "map地址", "信号方向", "信号长度",
                       "正向测试值", "关联信号", "关联属性"]
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)

            # 写入表格数据
            for row_num in range(self.table.rowCount()):
                for col_num in range(self.table.columnCount()):
                    # 优先检查 QComboBox
                    cell_widget = self.table.cellWidget(row_num, col_num)
                    if isinstance(cell_widget, QComboBox):
                        # 获取 QComboBox 的当前选择项
                        value = cell_widget.currentText()
                        print(f"导出行 {row_num + 1}，列 {col_num + 1} 的 QComboBox 值：{value}")
                    else:
                        # 获取普通单元格内容
                        item = self.table.item(row_num, col_num)
                        value = item.text() if item is not None else ""

                    # 写入到 Excel 单元格
                    sheet.cell(row=row_num + 2, column=col_num + 1, value=value)

            # 保存文件
            workbook.save(filename)
            QMessageBox.information(None, "成功", f"数据已成功导出到 {filename}")

        except Exception as e:
            logging.error(f"导出到Excel时发生错误: {traceback.format_exc()}")
            QMessageBox.critical(None, "错误", "导出失败，请查看日志！")

    def show_success_dialog(self, state, message):
        # 创建一个简单的Tkinter窗口
        success_dialog = tk.Tk()
        # 隐藏主窗口
        success_dialog.withdraw()
        # 显示成功消息
        messagebox.showinfo(state, message)

    def read_Messeage(self, path1):
        try:
            for i in range(0, self.table.rowCount()):
                datalist = []
                for j in range(0, self.table.columnCount() - 1):
                    data = self.table.item(i, j).text()
                    datalist.append(data)
                self.InterfaceList.append(datalist)
            # self.elfAnalysis.loadPath(path2)
            if path1:
                book1 = xlrd2.open_workbook(path1)
                sheetbook1 = book1.sheet_by_name('软件接口定义表')
                for i in range(1, sheetbook1.nrows):
                    values = sheetbook1.row_values(i)
                    # self.InterfaceList.append(values)
                    self.interfacedic[values[1]] = values[2]
                sheetbook2 = book1.sheet_by_name('硬件配置表')
                for i in range(2, 12):
                    values = sheetbook2.col_values(i)
                    CANlist = [str(values[2])] + values[4:]
                    CANlist = [item for item in CANlist if item and item.strip()]
                    self.hardwaredic[values[3]] = CANlist
                sheetbook3 = book1.sheet_by_name('LIN用例解析矩阵')
                dataList = []
                for i in range(sheetbook3.nrows - 1, 0, -1):
                    values = sheetbook3.row_values(i)
                    if values[3] == '':
                        # 信号描述、起始位、信号长度
                        data = [values[8], values[11], values[13]]
                        dataList.append(data)
                    elif values[3] != [] and dataList != []:
                        self.LINdic[values[3]] = dataList
                        dataList = []
                dataList = []
                sheetbook4 = book1.sheet_by_name('CAN用例解析矩阵')
                for i in range(sheetbook4.nrows - 1, 0, -1):
                    values = sheetbook4.row_values(i)
                    if values[3] == '':
                        # 信号描述、起始位、信号长度、周期
                        data = [values[8], values[11], values[13]]
                        dataList.append(data)
                    elif values[3] != [] and dataList != []:
                        self.CANdic[values[3] + '/' + str(values[5])] = dataList
                        dataList = []

                if self.genautotable():
                    self.show_success_dialog('Success', '生成成功！')
                else:
                    self.show_success_dialog('FAIL', '生成失败，没有找到相关属性！')
        except:
            self.show_success_dialog('FAIL', '生成失败，查看log！')
            logging.error(traceback.format_exc() + "\n")


    def genautotable(self):
        dataList = []
        # 遍历 tableWidget 中的每一行，动态生成 dataList
        for row in range(self.table.rowCount()):
            # 提取每一列的值
            item = []
            for col in range(self.table.columnCount() - 1):
                # 获取单元格控件
                cell_widget = self.table.cellWidget(row, col)
                cell_item = self.table.item(row, col)

                # 如果是 QComboBox，则获取其当前文本
                if isinstance(cell_widget, QComboBox):
                    item.append(cell_widget.currentText())
                elif cell_item is not None:
                    item.append(cell_item.text())
                else:
                    item.append("")

            # 以下逻辑保持不变，仅将数据来源从 InterfaceList 替换为 tableWidget
            channel = '00'
            type = ''
            CANID = ''
            for key, values in self.hardwaredic.items():
                for value in values:
                    if value in item[8]:  # 关联信号列
                        CANID = str(values[0])
                        channel = values.index(value)
                        type = key

            # 对首行的 CANID 处理
            if CANID == '700/701':
                CANID = '0x700'
            elif '.' in CANID:
                CANID = '0x' + CANID.split('.')[0]
            else:
                CANID = '0x000'

            var_name = item[1].strip().upper()  # 去除空格并转换为大写
            try:
                inaddress = self.elfAnalysis.getAddressWithName(var_name).split('x')[1]
            except KeyError:
                logging.error(f"Variable not found in ELF file: {var_name}")
                inaddress = "00000000"  # 使用默认值或跳过该条目

            if type == 'LIN':
                thirddata1 = [key for key, values in self.LINdic.items() for val in values if item[8] in val]
                thirdstartbit = [val[1] for key, values in self.LINdic.items() for val in values if item[8] in val]
                thirdlength = [val[2] for key, values in self.LINdic.items() for val in values if item[8] in val][0]
                startbit = str(hex(int(thirdstartbit[0]))).upper()[2:].zfill(4)
                shuxing = '01'
                cycle = '000A'
            elif type == 'CAN':
                thirddata1 = [key for key, values in self.CANdic.items() for val in values if item[8] in val]
                thirdstartbit = [val[1] for key, values in self.CANdic.items() for val in values if item[8] in val]
                thirdlength = [val[2] for key, values in self.CANdic.items() for val in values if item[8] in val][0]
                startbit = str(hex(int(thirdstartbit[0]))).upper()[2:].zfill(4)
                shuxing = '00'
                cycle = str(hex(int(thirddata1[0].split('/')[1]))).upper()[2:].zfill(4)
            else:
                thirddata1 = '00'
                thirdstartbit = '0'
                startbit = '0000'
                shuxing = '00'
                cycle = '0000'
                thirdlength = '0'

            # 根据信号方向（IN 或 OUT）生成测试用例
            if item[5] == 'IN':  # 信号方向为 IN
                preview = []
                if item[9] != 'HWA':
                    length = str(hex(int(item[6]))).upper()[2:].zfill(4)
                    data1 = thirddata1[0].split('x')[1].split('/')[0].zfill(3)
                    text = str(channel).zfill(2) + shuxing + '4' + data1 + '0064' + cycle + startbit + length
                    text = ' '.join(text[i:i + 2] for i in range(0, len(text), 2)).upper()
                    firsttext = ['event', item[8], text, '21', '1', CANID, '', '', '--', 'Motorola']
                    secondtext = ['check', '单条测试用例处理状态', CANID[-1], '22', '100', '0x710', '16', '8', '--','Motorola']
                else:
                    if item[7] == '1':
                        data2 = '0001'
                    else:
                        data2 = '0000'
                    shuxing = '01'
                    message1 = str(channel).zfill(2) + shuxing + data2 + 'FFFF0000'
                    message1 = ' '.join(message1[i:i + 2] for i in range(0, len(message1), 2)).upper()
                    firsttext = ['event', item[8], message1, '20', '1', '0x705', '', '', '--', 'Motorola']
                    secondtext = ['check', '单条测试用例处理状态', CANID[-1], '21', '100', '0x710', '16', '8', '--','Motorola']

                dataList.append(firsttext)
                preview.append(firsttext)
                dataList.append(secondtext)
                preview.append(secondtext)

                if item[9] != 'HWA':
                    string = self.config.getMessage(thirdstartbit[0], item[6], item[7], 8, '')
                    string11 = self.config.getMessage(thirdstartbit[0], item[6], '0', 8, '')
                    thirdtext = ['event', item[8], string, '100', '1', thirddata1[0], '', '', '--', 'Motorola']
                    dataList.append(thirdtext)
                    threetext = ['event', item[8], string11, '100', '1', thirddata1[0], '', '', '--', 'Motorola']
                    preview.append(threetext)

                num = '01' if item[0] in ('UInt8', 'Boolean') else '81' if item[0] == 'SInt8' else '02' if item[0] == 'UInt16' else '82' if \
                item[0] == 'SInt16' else '00'
                message = '2EF6E9' + inaddress + num + '00000000'
                message = ' '.join(message[i:i + 2] for i in range(0, len(message), 2)).upper()

                if item[9] != 'HWA':
                    forthtext = ['event', 'Rte_Read_' + item[2], message, '24', '1', '0x7F0', '', '', '--', 'Motorola']
                    fifthtext = ['check', '单条测试用例处理状态', '240', '25', '100', '0x710', '16', '8', '--','Motorola']
                    sixthtext = ['event', 'Rte_Read_' + item[2], '22 F6 E9', '26', '1', '0x7F0', '', '', '--','Motorola']
                    seventhtext = ['check', '单条测试用例处理状态', '240', '27', '100', '0x710', '16', '8', '--','Motorola']
                else:
                    forthtext = ['event', 'Rte_Read_' + item[2], message, '22', '1', '0x7F0', '', '', '--', 'Motorola']
                    fifthtext = ['check', '单条测试用例处理状态', '240', '23', '100', '0x710', '16', '8', '--', 'Motorola']
                    sixthtext = ['event', 'Rte_Read_' + item[2], '22 F6 E9', '24', '1', '0x7F0', '', '', '--','Motorola']
                    seventhtext = ['check', '单条测试用例处理状态', '240', '25', '100', '0x710', '16', '8', '--','Motorola']

                dataList.append(forthtext)
                preview.append(forthtext)
                dataList.append(fifthtext)
                preview.append(fifthtext)
                dataList.append(sixthtext)
                preview.append(sixthtext)
                dataList.append(seventhtext)
                preview.append(seventhtext)

                zhengxiang = str(hex(int(item[7]))).upper()[2:].zfill(8)
                fanxiang = '00000000' if zhengxiang != '00000000' else str(hex(int(item[7]))).upper()[2:].zfill(8)
                if fanxiang == '00000000':
                    fanxiang = '00000001'

                message8 = '62F6E9' + inaddress + num + zhengxiang
                message8 = ' '.join(message8[i:i + 2] for i in range(0, len(message8), 2)).upper()
                message16 = '62F6E9' + inaddress + num + fanxiang
                message16 = ' '.join(message16[i:i + 2] for i in range(0, len(message16), 2)).upper()

                if item[9] != 'HWA':
                    eighthtext = ['routercheck', 'Rte_Read_' + item[2], message8, '28', '100', '0x7F1', '', '', '--','Motorola']
                    eighttext = ['routercheck', 'Rte_Read_' + item[2], message16, '28', '100', '0x7F1', '', '', '--','Motorola']
                else:
                    eighthtext = ['routercheck', 'Rte_Read_' + item[2], message8, '26', '100', '0x7F1', '', '', '--','Motorola']
                    eighttext = ['routercheck', 'Rte_Read_' + item[2], message16, '26', '100', '0x7F1', '', '', '--','Motorola']

                dataList.append(eighthtext)
                preview.append(eighttext)
                dataList = dataList + preview
                note = ['测试用例说明', '由' + item[8] + '   测：Rte_Read_' + item[2]]
                dataList.append(note)

            elif item[5] == 'OUT':  # 信号方向为 OUT
                preview = []
                zhengxiang = str(hex(int(item[7]))).upper()[2:].zfill(8)
                fanxiang = '00000000' if zhengxiang != '00000000' else str(hex(int(item[7]))).upper()[2:].zfill(8)
                caozuozhi = str(hex(int(item[7]))).upper()[2:] if zhengxiang != '00000000' else '0'
                fanxiangnum = '0' if zhengxiang != '00000000' else str(hex(int(item[7]))).upper()[2:]

                num = '11' if item[0] in ('UInt8', 'Boolean') else '91' if item[0] == 'SInt8' else '12' if item[0] == 'UInt16' else '92' if \
                item[0] == 'SInt16' else '00'
                text = '2EF6E9' + inaddress + num + zhengxiang
                text = ' '.join(text[i:i + 2] for i in range(0, len(text), 2)).upper()
                text1 = '2EF6E9' + inaddress + num + fanxiang
                text1 = ' '.join(text1[i:i + 2] for i in range(0, len(text1), 2)).upper()

                firsttext = ['event', 'Rte_Write_' + item[2], text, '20', '1', '0x7F0', '', '', '--', 'Motorola']
                dataList.append(firsttext)
                firsttext1 = ['event', 'Rte_Write_' + item[2], text1, '20', '1', '0x7F0', '', '', '--', 'Motorola']
                preview.append(firsttext1)

                secondtext = ['check', '单条测试用例处理状态', '240', '21', '100', '0x710', '16', '8', '--', 'Motorola']
                dataList.append(secondtext)
                preview.append(secondtext)

                if '开关输入保留' in item[8]:
                    CANID = '0x706'

                if item[9] != 'HWA' and item[9] != 'APP':
                    data = thirddata1[0].split('/')[0][2:].zfill(4)
                    message3 = str(channel).zfill(2) + shuxing + data + '0064' + cycle + self.config.getStartandLengthHex(thirdstartbit[0],thirdlength).replace(' ', '')
                    message3 = ' '.join(message3[i:i + 2] for i in range(0, len(message3), 2)).upper()
                    thirdtext = ['event', item[8], message3, '22', '1', '0x701', '', '', '--', 'Motorola']
                    forthtext = ['check', '单条测试用例处理状态', '1', '100', '100', '0x710', '16', '8', '--','Motorola']
                    dataList.append(thirdtext)
                    preview.append(thirdtext)
                    dataList.append(forthtext)
                    preview.append(forthtext)

                    fifthtext = ['check', item[8], caozuozhi, '23', '100', CANID, '', '', '--', 'Motorola']
                    dataList.append(fifthtext)
                    fivetext = ['check', item[8], fanxiangnum, '23', '100', CANID, '', '', '--', 'Motorola']
                    preview.append(fivetext)

                    note = ['测试用例说明', '由Rte_Write_' + item[2] + '   测：' + item[8]]

                elif item[9] == 'HWA':
                    fifthtext = ['check', item[8], caozuozhi, '22', '100', CANID, '', '', '--', 'Motorola']
                    dataList.append(fifthtext)
                    fivetext = ['check', item[8], fanxiangnum, '22', '100', CANID, '', '', '--', 'Motorola']
                    preview.append(fivetext)

                    note = ['测试用例说明', '由Rte_Write_' + item[2] + '   测：' + item[8]]

                else:
                    num = '01' if item[0] in ('UInt8', 'Boolean') else '81' if item[0] == 'SInt8' else '02' if item[0] == 'UInt16' else '82' if \
                    item[0] == 'SInt16' else '00'
                    try:
                        name = [key for key, values in self.interfacedic.items() if item[8] == values][0]
                    except:
                        self.show_success_dialog('FAIL', 'APP属性的关联信号无法找到Rte接口！')
                        sys.exit(1)

                    outaddress = self.elfAnalysis.getAddressWithName(name).split('x')[1]
                    data = '2EF6E9' + outaddress + num + fanxiang
                    data = ' '.join(data[i:i + 2] for i in range(0, len(data), 2)).upper()

                    thirdtext = ['event', 'Rte_Read_' + item[8], data, '22', '1', '0x7F0', '', '', '--', 'Motorola']
                    dataList.append(thirdtext)
                    preview.append(thirdtext)

                    forthtext = ['check', '单条测试用例处理状态', '240', '23', '100', '0x710', '16', '8', '--','Motorola']
                    dataList.append(forthtext)
                    preview.append(forthtext)

                    fifthtext = ['event', 'Rte_Read_' + item[8], '22 F6 E9', '24', '1', '0x7F0', '', '', '--', 'Motorola']
                    dataList.append(fifthtext)
                    preview.append(fifthtext)

                    sixthtext = ['check', '单条测试用例处理状态', '240', '25', '100', '0x710', '16', '8', '--','Motorola']
                    dataList.append(sixthtext)
                    preview.append(sixthtext)

                    message7 = '62F6E9' + outaddress + num + zhengxiang
                    message7 = ' '.join(message7[i:i + 2] for i in range(0, len(message7), 2)).upper()
                    message14 = '62F6E9' + outaddress + num + fanxiang
                    message14 = ' '.join(message14[i:i + 2] for i in range(0, len(message14), 2)).upper()

                    seventhtext = ['routercheck', 'Rte_Read_' + item[8], message7, '26', '100', '0x7F1', '', '', '--','Motorola']
                    seventext = ['routercheck', 'Rte_Read_' + item[8], message14, '26', '100', '0x7F1', '', '', '--','Motorola']

                    dataList.append(seventhtext)
                    preview.append(seventext)

                    note = ['测试用例说明', '由Rte_Write_' + item[2] + '   测：Rte_Read_' + item[8]]

                dataList = dataList + preview
                dataList.append(note)

        has_dataList = bool(dataList)
        if not has_dataList:
            QMessageBox.warning(None, "警告", "没有可保存的数据！")
            return False
        else:
            # 写入 Excel 文件
            headnameList = ['序号', '操作类型', '操作名称', '操作值', '间隔（ms）', 'Cycle（ms）', 'canID', 'Start',
                            'Length', 'flag', 'format']
            file_path1, _ = QFileDialog.getSaveFileName(None, '导出自动测试文件', '', 'Excel File(*.xlsx)')
            workbook = Workbook(file_path1)
            # workbook = Workbook(path3)
            sheet1 = workbook.add_worksheet('CAN')
            sheet1.set_column('C:D', 50)
            sheet1.set_column('B:B', 15)
            sheet1.set_column('E:K', 10)
            font = workbook.add_format({'font_name': '等线', 'font_size': 12, 'align': 'center'})

            for k in range(len(headnameList)):
                sheet1.write(0, k, headnameList[k], font)

            row = 1
            for l in range(len(dataList)):
                sheet1.write(row, 0, str(row), font)
                for j in range(len(dataList[l])):
                    sheet1.write(row, j + 1, dataList[l][j], font)
                row += 1
            workbook.close()
            return True

    def excel_to_excel(self, filepath):
        """将当前表格数据保存到指定路径的 Excel 文件"""
        try:
            if not self.table:
                logging.warning("表格未初始化，跳过导出")
                return False

            # 确保目录存在
            os.makedirs(os.path.dirname(filepath), exist_ok=True)

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "软件接口定义表"

            # 写入表头
            headers = ["接口类型", "接口名称", "信号描述", "信号归属模块", "map地址", "信号方向", "信号长度",
                       "正向测试值", "关联信号", "关联属性"]
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)

            # 写入表格数据
            for row_num in range(self.table.rowCount()):
                for col_num in range(self.table.columnCount()):
                    # 优先检查 QComboBox
                    cell_widget = self.table.cellWidget(row_num, col_num)
                    if isinstance(cell_widget, QComboBox):
                        # 获取 QComboBox 的当前选择项
                        value = cell_widget.currentText()
                        print(f"导出行 {row_num + 1}，列 {col_num + 1} 的 QComboBox 值：{value}")
                    else:
                        # 获取普通单元格内容
                        item = self.table.item(row_num, col_num)
                        value = item.text() if item is not None else ""

                    # 写入到 Excel 单元格
                    sheet.cell(row=row_num + 2, column=col_num + 1, value=value)

            workbook.save(filepath)
            return True

        except Exception as e:
            logging.error(f"导出到Excel时发生错误: {traceback.format_exc()}")
            return False



